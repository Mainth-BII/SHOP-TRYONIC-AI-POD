import os
import sys
import glob
import subprocess
import openpyxl

def main():
    print("🚀 Bắt đầu tiến trình BATCH JIRA SYNC...")
    
    # 1. Tìm file Test Report mới nhất trong hệ thống
    base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
    reports_dir = os.path.join(base_dir, "Test cases", "Test_Reports")
    
    if os.path.exists(reports_dir):
        report_files = glob.glob(os.path.join(reports_dir, "TestReport_FULL_*.xlsx"))
    else:
        # Fallback nếu script chạy ngoài thư mục chuẩn
        report_files = glob.glob(os.path.join(base_dir, "TestReport_FULL_*.xlsx"))
        
    if not report_files:
        print("❌ Không tìm thấy Test Report nào trong hệ thống!")
        sys.exit(0) # CI/CD không rớt nếu không có test chạy

    latest_report = max(report_files, key=os.path.getctime)
    print(f"📄 Tìm thấy Report mới nhất: {os.path.basename(latest_report)}")

    # 2. Quét Excel tìm các case FAIL
    try:
        wb = openpyxl.load_workbook(latest_report, data_only=True)
    except Exception as e:
        print(f"❌ Lỗi đọc file Excel: {e}")
        sys.exit(1)
        
    bugs_to_log = []
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Tìm cột Result dựa vào keyword
        headers = {}
        header_row = 1
        for row_idx in range(1, 4):
            row_vals = [str(cell.value).strip().lower() for cell in ws[row_idx] if cell.value]
            if "result" in row_vals or "kết quả" in row_vals or "status" in row_vals:
                header_row = row_idx
                for col_idx, cell in enumerate(ws[row_idx], 1):
                    if cell.value:
                        headers[str(cell.value).strip().lower()] = col_idx
                break
                
        if not headers:
            continue
            
        result_col = headers.get("result", headers.get("kết quả"))
        if not result_col:
            continue

        tc_id_col = headers.get("test case id", headers.get("mã test case", 1))
        desc_col = headers.get("description", headers.get("mô tả", 3))
        actual_col = headers.get("actual result", headers.get("hiện trạng", headers.get("kết quả thực tế")))
        
        # Quét các dòng
        for row in range(header_row + 1, ws.max_row + 1):
            res_val = str(ws.cell(row=row, column=result_col).value or "").strip().upper()
            if res_val == "FAIL":
                tc_id = str(ws.cell(row=row, column=tc_id_col).value or "UNKNOWN-TC").strip()
                desc = str(ws.cell(row=row, column=desc_col).value or "").strip()
                actual = str(ws.cell(row=row, column=actual_col).value or "Không có actual result").strip() if actual_col else ""
                
                # Check nếu cột Bug ID đã có link thì bỏ qua (Tránh spam log trùng)
                bug_id_col = headers.get("bug id", headers.get("jira id"))
                if bug_id_col:
                    existing_bug = str(ws.cell(row=row, column=bug_id_col).value or "").strip()
                    if "TAS-" in existing_bug:
                        continue # Đã log trước đó rồi
                        
                bugs_to_log.append({
                    "tc_id": tc_id,
                    "sheet": sheet_name,
                    "summary": f"[{sheet_name}] {desc[:80]}...",
                    "details": f"1. Run Test Case: {tc_id}\nLỗi hiển thị: {actual}",
                    "impact": f"TC {tc_id} bị block trên luồng {sheet_name}"
                })

    # 3. Kích hoạt Skill Jira Bug Logging
    print(f"🐞 Phát hiện {len(bugs_to_log)} FAILED Test Cases cần tạo Bug Jira.")
    
    cli_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "log_jira_ticket.py")
    
    success_count = 0
    for bug in bugs_to_log:
        cmd = [
            sys.executable, cli_path,
            "--summary", bug["summary"],
            "--desc", bug["details"],
            "--impact_scope", bug["impact"],
            "--blocked_tcs", bug["tc_id"],
            "--cause", "Unknown" # Mặc định để QA vào cập nhật sau
        ]
        
        print(f"⏳ Đang log: {bug['tc_id']}...")
        result = subprocess.run(cmd, capture_output=True, text=True)
        
        if result.returncode == 0:
            success_count += 1
            print(f"  👉 OK!")
        else:
            print(f"  ❌ Failed: {result.stderr}")
            
    print(f"✅ BATCH SYNC HOÀN TẤT. Successfully logged {success_count}/{len(bugs_to_log)} bugs.")

if __name__ == "__main__":
    main()
