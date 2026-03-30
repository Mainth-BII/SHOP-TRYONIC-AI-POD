"""
Add 'ARTWORK TEST DATA' sheet to v30 Excel.

Nghiên cứu các loại artwork trên thế giới để phục vụ testing pipeline:
  Artwork → In lên áo → AI Try-on (model mặc thử)

Tổ chức theo category, mỗi artwork type có metadata phục vụ QA.
"""
import openpyxl, os, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
SRC = os.path.join(BASE, "TC_POD-TShirt-Platform_ExecutionSummary_v30_2026-03-26.xlsx")
today = datetime.date.today().strftime("%Y-%m-%d")
OUTPUT = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork.xlsx")

FF = "Calibri"; SZ = 11
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
CAT_FILL = PatternFill("solid", fgColor="D6E4F0")
CAT_FONT = Font(name=FF, size=12, bold=True, color="1F4E79")
BODY_FONT = Font(name=FF, size=SZ)
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Color fills for test status
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
CATEGORY_COLORS = {
    "🎨 TRADITIONAL FINE ART": PatternFill("solid", fgColor="F2DCDB"),
    "💻 DIGITAL ART": PatternFill("solid", fgColor="DCE6F1"),
    "🤖 AI-GENERATED ART STYLES": PatternFill("solid", fgColor="E4DFEC"),
    "🌍 CULTURAL / REGIONAL ART": PatternFill("solid", fgColor="EBF1DE"),
    "👕 POD-SPECIFIC / T-SHIRT DESIGN": PatternFill("solid", fgColor="FDE9D9"),
    "⚠️ TECHNICAL CHALLENGE CASES": PatternFill("solid", fgColor="DAEEF3"),
}

# ─── Columns ───
COLUMNS = [
    "No.",
    "Category",
    "Artwork Type",
    "Sub-Style / Variant",
    "Description",
    "Origin / Culture",
    "Color Complexity",          # Low, Medium, High, Very High
    "Detail Level",              # Simple, Moderate, Complex, Ultra
    "Transparency / Alpha",      # Yes / No
    "Typical File Format",       # PNG, SVG, JPG, AI, PSD
    "Print Technique",           # DTG, Screen Print, Sublimation, Heat Transfer, Embroidery
    "Expected Print Quality",    # Excellent, Good, Acceptable, Challenging
    "AI Try-on Challenge",       # None, Low, Medium, High
    "Key Test Focus",
    "Sample Prompt (AI Gen)",
    "Test Priority",             # P0, P1, P2, P3
    "Test Status",               # Untested, Pass, Fail, N/A
    "Notes / Risk"
]
COL_WIDTHS = [5, 30, 22, 24, 40, 18, 16, 14, 16, 18, 22, 20, 18, 40, 40, 12, 12, 35]

# ═══════════════════════════════════════════════════════════════
# ARTWORK DATA — Organized by category
# ═══════════════════════════════════════════════════════════════
ARTWORK_DATA = {
    "🎨 TRADITIONAL FINE ART": [
        # (Artwork Type, Sub-Style, Description, Origin, Color Complexity, Detail Level, Transparency, File Format, Print Technique, Print Quality, AI Tryon Challenge, Key Test Focus, Sample Prompt, Priority, Notes)
        ("Oil Painting", "Impressionist", "Soft brushstrokes, light-focused, blended colors", "Europe (France)", "High", "Complex", "No", "JPG/PNG", "DTG, Sublimation", "Good", "Medium", "Color reproduction accuracy, brush texture visibility on fabric", "Impressionist oil painting of sunflowers at sunset", "P1", "Texture detail may be lost in small prints"),
        ("Oil Painting", "Hyperrealism", "Photo-realistic paintings with extreme detail", "Global", "Very High", "Ultra", "No", "JPG/PNG", "DTG", "Excellent", "Low", "Ultra-HD resolution handling, edge sharpness on fabric", "Hyperrealistic oil painting of water droplets on glass", "P1", "Requires high-res source (300+ DPI)"),
        ("Watercolor", "Wet-on-Wet", "Transparent washes, soft edges, organic bleeding", "Global", "Medium", "Moderate", "Yes", "PNG", "DTG, Sublimation", "Good", "Medium", "Transparency preservation, soft gradients on fabric", "Watercolor painting of cherry blossoms with wet wash", "P1", "Alpha channel handling critical"),
        ("Watercolor", "Botanical Illustration", "Precise plant drawings with watercolor rendering", "Europe (UK)", "Medium", "Complex", "Yes", "PNG", "DTG", "Excellent", "Low", "Fine line preservation + watercolor transparency blending", "Botanical watercolor illustration of orchid species", "P2", "White background removal important"),
        ("Acrylic Painting", "Abstract Expressionism", "Bold colors, thick textures, dynamic composition", "USA", "Very High", "Complex", "No", "JPG/PNG", "DTG, Sublimation", "Good", "Low", "Color vibrancy reproduction, texture rendering", "Abstract expressionist acrylic painting with bold drips", "P1", "Vibrant colors may shift in CMYK printing"),
        ("Charcoal Drawing", "Portrait", "Monochrome, high contrast, subtle gradations", "Global", "Low", "Complex", "No", "JPG/PNG", "DTG, Screen Print", "Excellent", "Low", "Grayscale accuracy, fine detail preservation", "Charcoal portrait drawing of elderly person", "P2", "Works well on light and dark garments"),
        ("Ink Drawing", "Line Art", "Clean lines, minimal shading, pen-and-ink style", "Global", "Low", "Simple", "Yes", "SVG/PNG", "Screen Print, DTG", "Excellent", "None", "Line crispness, scalability, transparency handling", "Ink line drawing of mountain landscape", "P0", "Most basic artwork type — baseline test"),
        ("Pencil Sketch", "Realistic Sketch", "Graphite on paper, soft shading, tonal range", "Global", "Low", "Moderate", "No", "JPG/PNG", "DTG", "Good", "Low", "Subtle tonal gradation, paper texture rendering", "Detailed pencil sketch of a cat", "P2", "Low contrast may be lost on colored garments"),
        ("Pastel Art", "Soft Pastel", "Powdery texture, vibrant soft colors, blending", "Europe", "High", "Moderate", "No", "JPG/PNG", "DTG, Sublimation", "Good", "Medium", "Soft gradient reproduction, pastel color accuracy", "Soft pastel landscape of lavender fields", "P2", "Delicate colors require accurate color profiles"),
        ("Printmaking", "Woodcut / Linocut", "Bold graphic shapes, high contrast, carved textures", "Japan/Europe", "Low", "Moderate", "No", "SVG/PNG", "Screen Print", "Excellent", "None", "Clean solid areas, sharp edges", "Woodcut print of a wolf in forest", "P1", "Ideal for screen printing"),
        ("Printmaking", "Lithography", "Tonal gradations, multi-layer colored prints", "Europe", "Medium", "Complex", "No", "JPG/PNG", "DTG", "Good", "Low", "Multi-tone reproduction, registration accuracy", "Lithographic print of vintage botanical", "P2", "Classic print aesthetic"),
        ("Mosaic Art", "Byzantine / Tile", "Pixelated pattern from small pieces, geometric", "Middle East/Mediterranean", "High", "Complex", "No", "JPG/PNG", "DTG, Sublimation", "Good", "Low", "Pattern clarity, color segmentation on fabric", "Byzantine mosaic of golden angel", "P3", "Zoom level testing important"),
    ],

    "💻 DIGITAL ART": [
        ("Vector Art", "Flat Design", "Clean geometric shapes, solid flat colors, no gradients", "Global", "Medium", "Simple", "Yes", "SVG/AI/PNG", "Screen Print, DTG, HTV", "Excellent", "None", "Scalability test, flat color accuracy, transparency", "Flat design vector illustration of cityscape", "P0", "Primary format for POD — must be perfect"),
        ("Vector Art", "Isometric", "3D-like views with isometric perspective", "Global", "Medium", "Moderate", "Yes", "SVG/PNG", "DTG", "Excellent", "Low", "Perspective accuracy, line alignment on garment curves", "Isometric illustration of a coffee shop", "P2", "Straight lines may distort on body curves"),
        ("Digital Painting", "Semi-Realistic", "Blended digital brushwork, moderate realism", "Global", "High", "Complex", "No", "PNG/PSD", "DTG, Sublimation", "Good", "Medium", "Brush texture rendering, color depth on fabric", "Semi-realistic digital painting of warrior character", "P1", "Standard digital art test"),
        ("Digital Painting", "Concept Art", "Character/environment design for games/film", "Global", "High", "Complex", "Yes", "PNG/PSD", "DTG", "Good", "Medium", "Complex composition cropping, BG transparency", "Concept art of sci-fi spaceship in nebula", "P1", "Often has transparent backgrounds"),
        ("Pixel Art", "Retro 8-bit", "Chunky pixels, limited color palette, retro gaming", "Japan/USA", "Low", "Simple", "Yes", "PNG/GIF", "Screen Print, DTG", "Excellent", "None", "Pixel crispness (no anti-aliasing), upscaling quality", "8-bit pixel art of retro game character", "P1", "Must NOT apply smoothing/interpolation"),
        ("Pixel Art", "Hi-Res Pixel", "Modern pixel art with larger canvas, more colors", "Global", "Medium", "Moderate", "Yes", "PNG", "DTG", "Excellent", "Low", "Maintain pixel grid at print size", "Hi-res pixel art landscape with parallax layers", "P2", "Scaling ratio must be integer multiple"),
        ("3D Render", "Photo-Realistic Render", "Ray-traced 3D scenes with lighting, shadows", "Global", "Very High", "Ultra", "No", "JPG/PNG", "DTG, Sublimation", "Excellent", "Low", "Shadow rendering, specular highlights on fabric", "Photorealistic 3D render of chrome robot", "P1", "High file size — performance test"),
        ("3D Render", "Low Poly", "Simplified 3D with visible polygon faces", "Global", "Medium", "Moderate", "Yes", "PNG", "DTG, Screen Print", "Excellent", "None", "Facet edges sharpness, gradient within faces", "Low poly 3D animal head portrait", "P1", "Popular POD style"),
        ("Digital Collage", "Photo Manipulation", "Composite of multiple photos with digital effects", "Global", "Very High", "Ultra", "Yes", "PNG/PSD", "DTG", "Good", "Medium", "Layer blending on fabric, edge artifact detection", "Surreal photo collage of floating islands", "P2", "Complex alpha channels"),
        ("Motion Graphics Frame", "Animated Poster Style", "Single frame from animation, dynamic composition", "Global", "High", "Complex", "Yes", "PNG", "DTG", "Good", "Medium", "Dynamic element clarity in static print", "Animated poster style with flowing particles", "P3", "May lose motion context in print"),
        ("Generative Art", "Algorithmic / Fractal", "Math-generated patterns, infinite detail", "Global", "Very High", "Ultra", "No", "PNG/JPG", "Sublimation", "Good", "Low", "Fractal detail resolution at print size", "Mandelbrot fractal with rainbow gradient", "P2", "Beautiful at high res, complex color mapping"),
    ],

    "🤖 AI-GENERATED ART STYLES": [
        ("AI Realistic", "Photorealistic Portrait", "AI-generated human portrait, near-photo quality", "AI (Midjourney/SD)", "High", "Ultra", "No", "PNG/JPG", "DTG", "Good", "High", "AI artifact detection on print, face accuracy on tryon", "Realistic portrait of woman with flowers in hair, dramatic lighting", "P0", "Critical: AI Try-on may confuse printed face with model"),
        ("AI Fantasy", "Epic Fantasy Scene", "Dramatic fantasy landscapes, character art", "AI (Midjourney/SD)", "Very High", "Ultra", "No", "PNG", "DTG, Sublimation", "Good", "Low", "Color vibrancy, detail preservation at garment scale", "Epic fantasy dragon flying over misty mountains", "P1", "High detail may compress poorly"),
        ("AI Anime", "Anime / Manga Style", "Japanese animation style with AI generation", "AI (Midjourney --niji)", "High", "Complex", "Yes", "PNG", "DTG", "Excellent", "Medium", "Anime line quality, cel-shading on fabric", "Anime-style cyber samurai with neon katana", "P0", "Very popular for POD market"),
        ("AI Abstract", "Abstract Fluid Art", "AI-generated flowing abstract patterns", "AI (Midjourney/SD)", "Very High", "Complex", "No", "PNG/JPG", "DTG, Sublimation", "Good", "Low", "Fluid color blending, seamless edges for allover", "Abstract fluid art with gold and teal marbling", "P1", "Test seamless tiling for allover prints"),
        ("AI Chibi", "Kawaii / Chibi Characters", "Cute, oversized-head characters", "AI (Midjourney --niji)", "Medium", "Moderate", "Yes", "PNG", "DTG, Screen Print", "Excellent", "Low", "Character cuteness preservation, clean alpha cut", "Kawaii chibi cat astronaut with star background", "P1", "Popular subgenre for casual wear"),
        ("AI Typography", "3D Text / Lettering", "AI-generated typographic artwork", "AI (Midjourney/SD)", "Medium", "Moderate", "Yes", "PNG", "DTG, Screen Print", "Good", "None", "Text readability, 3D effect on flat fabric", "3D chrome text saying 'DREAM BIG' with flames", "P1", "Text legibility must be verified"),
        ("AI Surreal", "Dreamcore / Surrealism", "Surreal dreamlike compositions", "AI (Midjourney/SD)", "High", "Ultra", "No", "PNG/JPG", "DTG", "Good", "Medium", "Surreal element clarity, object recognition on garment", "Surreal staircase leading to clouds with floating clocks", "P2", "Interesting but niche market"),
        ("AI Retro", "Retrowave / Synthwave", "80s-inspired neon aesthetics, retro-futuristic", "AI (Midjourney/SD)", "High", "Moderate", "No", "PNG", "DTG, Sublimation", "Good", "Low", "Neon glow reproduction on fabric, gradient banding", "Synthwave sunset with palm trees and DeLorean", "P1", "Strong POD seller — gradient testing critical"),
        ("AI Sticker", "Die-Cut Sticker Style", "Outlined sticker with white border, glossy look", "AI (Midjourney/SD)", "Medium", "Simple", "Yes", "PNG", "DTG, HTV", "Excellent", "None", "Border crispness, sticker illusion on fabric", "Cute sticker style corgi with thick white outline", "P0", "Very popular format — white outline test"),
        ("AI Watercolor", "AI Watercolor Simulation", "AI-generated watercolor effect art", "AI (Midjourney/SD)", "Medium", "Complex", "Yes", "PNG", "DTG", "Good", "Medium", "Watercolor bleeding effect, transparency edges", "AI watercolor painting of Parisian cafe in rain", "P2", "Compare with real watercolor scan"),
        ("AI Comic", "Comic Book / Graphic Novel", "AI comic-style with halftone, bold lines", "AI (Midjourney/SD)", "High", "Complex", "No", "PNG", "DTG, Screen Print", "Excellent", "Low", "Halftone pattern, speech bubble rendering", "Comic book superhero in dynamic action pose", "P2", "Halftone may cause moiré pattern"),
    ],

    "🌍 CULTURAL / REGIONAL ART": [
        ("Japanese Ukiyo-e", "Woodblock Print", "Edo-period style, flat colors, bold outlines, nature/figures", "Japan", "Medium", "Complex", "No", "JPG/PNG", "DTG, Screen Print", "Excellent", "Low", "Traditional color palette, woodgrain texture rendering", "Ukiyo-e style great wave with modern twist", "P1", "Classic art-on-clothing test"),
        ("Chinese Ink Wash", "Sumi-e / Shui-mo", "Minimalist brush painting, ink gradients, zen aesthetics", "China", "Low", "Moderate", "Yes", "PNG", "DTG", "Good", "Medium", "Ink transparency, brush stroke delicacy on fabric", "Chinese ink wash painting of bamboo in mist", "P1", "Subtle tones may wash out on colored garments"),
        ("Chinese Calligraphy", "Classical Script", "Artistic brushwork characters, cultural significance", "China", "Low", "Simple", "Yes", "SVG/PNG", "Screen Print, DTG", "Excellent", "None", "Character clarity, stroke precision", "Chinese calligraphy of proverb with red seal", "P2", "Cultural sensitivity check needed"),
        ("Indian Art", "Madhubani / Mithila", "Folk art with intricate patterns, bright colors, natural motifs", "India (Bihar)", "High", "Complex", "No", "JPG/PNG", "DTG, Sublimation", "Good", "Low", "Pattern density, color saturation, symmetry", "Madhubani painting of peacock with floral border", "P1", "Dense patterns — zoom detail test"),
        ("Indian Art", "Mandala", "Symmetric circular geometric patterns, spiritual motifs", "India/Tibet", "High", "Ultra", "Yes", "SVG/PNG", "DTG, Sublimation", "Good", "Low", "Circular symmetry precision, fine line at edges", "Intricate mandala with gold and turquoise details", "P0", "Extremely popular for POD — must handle detail"),
        ("Indian Art", "Kalamkari", "Hand-painted/block-printed, mythological themes, natural dyes", "India (AP/Telangana)", "Medium", "Complex", "No", "JPG/PNG", "DTG", "Good", "Low", "Organic pattern flow, earth tone reproduction", "Kalamkari tree of life with birds and flowers", "P2", "Earth tones may vary across print methods"),
        ("African Art", "Tribal / Geometric Patterns", "Bold geometric shapes, symbolic motifs, earth/bright palette", "Africa (various)", "High", "Moderate", "No", "SVG/PNG", "Screen Print, DTG, Sublimation", "Excellent", "None", "Pattern repeat alignment, bold contrast preservation", "African tribal geometric pattern in terracotta and black", "P1", "Test seamless repeat tiling"),
        ("African Art", "Kente / Ankara", "Woven/printed fabric-on-fabric patterns, cultural significance", "Ghana / West Africa", "Very High", "Complex", "No", "JPG/PNG", "Sublimation", "Good", "Low", "Color vibrancy of traditional palette, stripe alignment", "Kente cloth inspired allover pattern with gold and green", "P2", "Allover print test — edge-to-edge"),
        ("African Art", "Mud Cloth (Bògòlanfini)", "Hand-dyed with fermented mud, geometric earthy patterns", "Mali", "Low", "Simple", "No", "JPG/PNG", "Screen Print, DTG", "Excellent", "None", "Earth tone accuracy, organic texture rendering", "Mud cloth inspired pattern in brown and cream", "P2", "Testing organic imperfect pattern edges"),
        ("Middle Eastern Art", "Islamic Geometric / Arabesque", "Complex geometric tessellations, infinite patterns", "Middle East / Islamic World", "Medium", "Ultra", "No", "SVG/PNG", "Sublimation, DTG", "Excellent", "None", "Geometric precision, seamless tiling, fine detail", "Islamic geometric star pattern in blue and gold", "P1", "Perfect for allover print testing"),
        ("Latin American Art", "Mexican Folk Art / Alebrije", "Vibrant colors, fantastical creatures, Day of the Dead motifs", "Mexico", "Very High", "Complex", "Yes", "PNG", "DTG", "Good", "Low", "Extreme color vibrancy, detailed creature rendering", "Colorful alebrije dragon with intricate patterns", "P1", "Gamut check for extreme colors"),
        ("Aboriginal Art", "Dot Painting", "Patterns made from dots, earth tones, dreamtime stories", "Australia", "Medium", "Moderate", "No", "JPG/PNG", "DTG", "Good", "Low", "Dot pattern clarity, even dot size at print scale", "Aboriginal dot painting of kangaroo dreaming", "P2", "Small dots may merge at low resolution"),
        ("Korean Art", "Minhwa Folk Painting", "Bright folk painting, tigers, peonies, symbolic imagery", "Korea", "High", "Complex", "No", "JPG/PNG", "DTG", "Good", "Low", "Traditional color accuracy, folk style preservation", "Minhwa tiger painting with magpie and pine", "P2", "Growing market interest"),
        ("Thai Art", "Khon / Temple Art", "Gold leaf, intricate patterns, mythological figures", "Thailand", "Very High", "Ultra", "No", "JPG/PNG", "DTG, Sublimation", "Good", "Medium", "Gold metallic rendering, extreme ornamental detail", "Thai temple art Garuda with gold ornaments", "P3", "Gold metallic colors challenge for print"),
        ("European Art", "Renaissance Style", "Classical proportion, realistic anatomy, religious/mythological", "Europe (Italy)", "High", "Ultra", "No", "JPG/PNG", "DTG", "Good", "High", "Fine art reproduction, skin tone on AI tryon", "Renaissance style portrait in the manner of Raphael", "P2", "AI Try-on face detection vs painted face"),
        ("Celtic Art", "Knotwork / Interlace", "Interlocking patterns, no beginning/end, spirals", "Ireland / Scotland", "Low", "Complex", "Yes", "SVG/PNG", "Screen Print, DTG", "Excellent", "None", "Knot continuity, line consistency, scalability", "Celtic knotwork circle with trinity knot center", "P2", "Great for screen print testing"),
    ],

    "👕 POD-SPECIFIC / T-SHIRT DESIGN": [
        ("Typography Design", "Bold Slogan", "Large text, motivational/humorous, impactful fonts", "Global", "Low", "Simple", "Yes", "SVG/PNG", "Screen Print, HTV, DTG", "Excellent", "None", "Font rendering, text readability across sizes", "Bold typography 'JUST CREATE' in distressed font", "P0", "Core POD product — must be flawless"),
        ("Typography Design", "Retro Script", "Vintage script lettering, decorative flourishes", "USA", "Medium", "Moderate", "Yes", "SVG/PNG", "Screen Print, DTG", "Good", "None", "Script flow, thin stroke preservation", "Retro script lettering 'California Vibes' with sunset", "P1", "Thin strokes may disappear in small sizes"),
        ("Graphic Tee", "Vintage Badge / Emblem", "Circular/shield badge, distressed texture, retro", "USA", "Medium", "Moderate", "Yes", "SVG/PNG", "Screen Print, DTG", "Excellent", "None", "Badge symmetry, distress texture clarity", "Vintage camping badge with mountains and eagle", "P0", "Top-selling POD genre — critical priority"),
        ("Graphic Tee", "Streetwear / Urban", "Bold graphics, graffiti-influenced, edgy typography", "Global (Urban)", "High", "Complex", "Yes", "PNG", "DTG, Screen Print", "Good", "Low", "Edge detail, graffiti texture rendering", "Streetwear skull design with neon graffiti splash", "P1", "Popular for youth market"),
        ("Animal Art", "Realistic Animal Portrait", "Photorealistic animal depiction, wildlife focus", "Global", "High", "Ultra", "Yes", "PNG", "DTG", "Good", "Low", "Fur/feather detail, eye rendering, BG transparency", "Realistic wolf portrait with intense blue eyes", "P1", "Always-popular POD niche"),
        ("Animal Art", "Geometric / Polygonal Animal", "Animals built from geometric shapes, modern style", "Global", "Medium", "Moderate", "Yes", "SVG/PNG", "DTG, Screen Print", "Excellent", "None", "Shape precision, gradient fill within polygons", "Geometric low-poly bear head in jewel tones", "P1", "Clean geometric edges important"),
        ("Nature / Landscape", "Scenic Print", "Mountain, ocean, forest scenes for garment print", "Global", "High", "Complex", "No", "JPG/PNG", "DTG, Sublimation", "Good", "Low", "Natural color accuracy, horizon line placement", "Mountain landscape at golden hour with reflection lake", "P2", "Full-front or allover print format"),
        ("Pattern / Allover", "Floral Pattern", "Repeating flower/botanical patterns, seamless tile", "Global", "High", "Complex", "No", "JPG/PNG", "Sublimation", "Excellent", "Low", "Seamless repeat precision, color consistency across tile", "Seamless tropical floral pattern with hibiscus", "P1", "Sublimation allover — seam alignment test"),
        ("Pattern / Allover", "Geometric Repeat", "Repeating geometric shapes, tessellations", "Global", "Medium", "Moderate", "No", "SVG/PNG", "Sublimation", "Excellent", "None", "Tile alignment, pattern scale across sizes", "Seamless art deco geometric in gold and navy", "P1", "Test same pattern across S to 3XL"),
        ("Pattern / Allover", "Camo / Military", "Camouflage patterns in various color schemes", "Global", "Medium", "Moderate", "No", "JPG/PNG", "Sublimation", "Excellent", "Low", "Organic shape blending, color scheme accuracy", "Custom camouflage pattern in urban gray tones", "P2", "Test pattern disruption at seams"),
        ("Minimalist", "Single Icon / Symbol", "Very simple single-element design, clean lines", "Global", "Low", "Simple", "Yes", "SVG/PNG", "Screen Print, HTV", "Excellent", "None", "Icon clarity at small placement, crisp edges", "Minimalist line heart icon in center chest", "P0", "Simplest design — baseline performance test"),
        ("Pop Culture", "Parody / Mashup", "Pop culture references, mashup designs, humorous", "Global", "High", "Complex", "Yes", "PNG", "DTG", "Good", "Low", "Character detail, color accuracy for brand parody", "Pop culture mashup of space cats with pizza", "P2", "IP/Copyright concern — note for legal"),
        ("Sports / Team", "Jersey Number Style", "Athletic numbering, team-style graphics", "Global", "Low", "Simple", "Yes", "SVG/PNG", "Screen Print, HTV", "Excellent", "None", "Number alignment, front/back placement consistency", "Sports jersey number 23 with team name", "P2", "Test dual-side print registration"),
    ],

    "⚠️ TECHNICAL CHALLENGE CASES": [
        ("Edge Case", "Ultra-Wide Panoramic", "Very wide aspect ratio artwork (3:1 or wider)", "N/A", "High", "Complex", "No", "JPG/PNG", "DTG", "Challenging", "Medium", "Aspect ratio cropping, auto-fit behavior", "Panoramic mountain range 3:1 aspect ratio", "P0", "Tests platform auto-crop / fit logic"),
        ("Edge Case", "Ultra-Tall Vertical", "Very tall aspect ratio artwork (1:3 or taller)", "N/A", "Medium", "Moderate", "No", "JPG/PNG", "DTG", "Challenging", "Low", "Vertical auto-fit, garment zone alignment", "Tall vertical Japanese scroll painting", "P1", "Tests vertical alignment handling"),
        ("Edge Case", "Micro Detail", "Extremely fine lines and tiny details", "N/A", "Low", "Ultra", "Yes", "SVG/PNG", "DTG", "Challenging", "None", "Line visibility at print size, minimum line width", "Extremely detailed mandala with 0.5pt lines", "P0", "Tests minimum printable detail threshold"),
        ("Edge Case", "Neon / Out-of-Gamut Colors", "Colors outside typical CMYK gamut, electric neon", "N/A", "Very High", "Moderate", "No", "PNG", "DTG", "Challenging", "Low", "Color gamut clipping, neon rendering fallback", "Neon green and electric blue abstract art", "P0", "CRITICAL: Gamut mapping test for print accuracy"),
        ("Edge Case", "Full Black Artwork on Dark", "Black-only artwork placed on black garment", "N/A", "Low", "Simple", "Yes", "PNG", "DTG (white underbase)", "Challenging", "None", "Visibility check, white underbase activation", "Black tattoo-style art on black t-shirt", "P1", "Platform should warn or auto-adjust"),
        ("Edge Case", "Full White Artwork", "White-only design, requires colored garment context", "N/A", "Low", "Simple", "Yes", "PNG", "DTG (white ink), HTV", "Challenging", "None", "White ink handling, transparency vs white detection", "White silhouette logo on transparent background", "P0", "Tests white color detection logic"),
        ("Edge Case", "Very Low Resolution (72 DPI)", "Source image at web resolution, insufficient for print", "N/A", "Medium", "Simple", "No", "JPG", "DTG", "Challenging", "Low", "DPI warning system, upscale quality check", "Low-res 500x500px JPG artwork", "P0", "Platform should warn about poor quality"),
        ("Edge Case", "Ultra-High Resolution (600+ DPI)", "Extremely large file, high memory usage", "N/A", "High", "Ultra", "No", "PNG/TIFF", "DTG", "Excellent", "Low", "Upload performance, memory handling, processing time", "8000x8000px artwork at 600 DPI", "P0", "Performance / timeout / crash test"),
        ("Edge Case", "Animated GIF → Static", "User uploads animated GIF, platform must handle", "N/A", "Medium", "Simple", "Yes", "GIF", "DTG", "Challenging", "None", "First frame extraction, animation rejection handling", "Animated sparkle GIF uploaded as artwork", "P1", "Test error handling or frame selection"),
        ("Edge Case", "CMYK PDF Upload", "Print-ready PDF in CMYK color space", "N/A", "Medium", "Moderate", "No", "PDF", "DTG", "Good", "None", "CMYK to RGB conversion, color shift check", "CMYK PDF of brand logo", "P1", "Tests file format support breadth"),
        ("Edge Case", "Extremely Large Text (A)", "Large text covering entire garment front", "N/A", "Low", "Simple", "Yes", "SVG/PNG", "Screen Print, DTG", "Good", "None", "Text scale limits, garment boundary overflow", "Single letter 'A' filling entire print area", "P1", "Tests maximum scale boundary"),
        ("Edge Case", "Many Small Objects", "100+ small icons/objects in one artwork", "N/A", "High", "Ultra", "Yes", "PNG/SVG", "DTG", "Challenging", "Low", "Small object clarity, rendering performance", "Pattern of 100 different tiny emoji icons", "P2", "Rendering stress test"),
        ("Edge Case", "Photo with Human Face", "Real photograph with clearly visible human face", "N/A", "High", "Ultra", "No", "JPG", "DTG", "Good", "High", "AI Try-on face confusion with printed face", "High-quality portrait photo on t-shirt", "P0", "CRITICAL: AI try-on must distinguish print vs model face"),
        ("Edge Case", "Transparent Gradient Edge", "Artwork edges that fade to transparent gradually", "N/A", "Medium", "Moderate", "Yes", "PNG", "DTG", "Good", "Medium", "Gradient transparency rendering on fabric", "Smoke effect fading from opaque to transparent", "P1", "Tests alpha gradient handling on garment"),
    ],
}


def main():
    if not os.path.exists(SRC):
        print(f"❌ Source not found: {SRC}")
        return

    wb = openpyxl.load_workbook(SRC)

    # Remove existing sheet if exists
    if "ARTWORK TEST DATA" in wb.sheetnames:
        del wb["ARTWORK TEST DATA"]

    ws = wb.create_sheet("ARTWORK TEST DATA")

    # ─── HEADER ROW ───
    for ci, col_name in enumerate(COLUMNS, 1):
        c = ws.cell(1, ci, col_name)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci-1]

    ws.row_dimensions[1].height = 35
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"
    ws.freeze_panes = "A2"

    # ─── DATA ROWS ───
    row = 2
    idx = 1
    for category, items in ARTWORK_DATA.items():
        cat_fill = CATEGORY_COLORS.get(category, CAT_FILL)
        # Category separator row
        c = ws.cell(row, 1, category)
        c.font = CAT_FONT
        c.fill = cat_fill
        for ci in range(1, len(COLUMNS) + 1):
            ws.cell(row, ci).fill = cat_fill
            ws.cell(row, ci).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
        ws.row_dimensions[row].height = 28
        row += 1

        for item in items:
            (art_type, sub_style, desc, origin, color_complex, detail_lv,
             transparency, file_fmt, print_tech, print_qual,
             ai_tryon_challenge, test_focus, sample_prompt, priority, notes) = item

            ws.cell(row, 1, idx).font = BODY_FONT
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 2, category.split(' ', 1)[-1] if ' ' in category else category).font = BODY_FONT
            ws.cell(row, 3, art_type).font = Font(name=FF, size=SZ, bold=True)
            ws.cell(row, 4, sub_style).font = BODY_FONT
            ws.cell(row, 5, desc).font = BODY_FONT
            ws.cell(row, 6, origin).font = BODY_FONT
            ws.cell(row, 7, color_complex).font = BODY_FONT
            ws.cell(row, 8, detail_lv).font = BODY_FONT
            ws.cell(row, 9, transparency).font = BODY_FONT
            ws.cell(row, 10, file_fmt).font = BODY_FONT
            ws.cell(row, 11, print_tech).font = BODY_FONT
            ws.cell(row, 12, print_qual).font = BODY_FONT
            ws.cell(row, 13, ai_tryon_challenge).font = BODY_FONT
            ws.cell(row, 14, test_focus).font = BODY_FONT
            ws.cell(row, 15, sample_prompt).font = Font(name=FF, size=SZ, italic=True, color="555555")
            ws.cell(row, 16, priority).font = Font(name=FF, size=SZ, bold=True)
            ws.cell(row, 17, "Untested").font = BODY_FONT
            ws.cell(row, 18, notes).font = BODY_FONT

            # Priority coloring
            p_cell = ws.cell(row, 16)
            if priority == "P0":
                p_cell.fill = PatternFill("solid", fgColor="FF6B6B")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
            elif priority == "P1":
                p_cell.fill = PatternFill("solid", fgColor="FFA500")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
            elif priority == "P2":
                p_cell.fill = WARN_FILL
            elif priority == "P3":
                p_cell.fill = PatternFill("solid", fgColor="E2EFDA")

            # AI Try-on Challenge coloring
            ai_cell = ws.cell(row, 13)
            if ai_tryon_challenge == "High":
                ai_cell.fill = FAIL_FILL
                ai_cell.font = Font(name=FF, size=SZ, bold=True, color="9C0006")
            elif ai_tryon_challenge == "Medium":
                ai_cell.fill = WARN_FILL

            # Print Quality coloring
            pq_cell = ws.cell(row, 12)
            if print_qual == "Challenging":
                pq_cell.fill = FAIL_FILL

            for ci in range(1, len(COLUMNS) + 1):
                ws.cell(row, ci).alignment = WRAP
                ws.cell(row, ci).border = BORDER

            row += 1
            idx += 1

    # ─── SUMMARY STATS ───
    row += 1
    summary_start = row
    ws.cell(row, 1, "📊 SUMMARY STATISTICS").font = Font(name=FF, size=14, bold=True, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    stats = [
        ("Total Artwork Types", idx - 1),
        ("Categories", len(ARTWORK_DATA)),
        ("P0 (Critical)", sum(1 for items in ARTWORK_DATA.values() for i in items if i[13] == "P0")),
        ("P1 (High)", sum(1 for items in ARTWORK_DATA.values() for i in items if i[13] == "P1")),
        ("P2 (Medium)", sum(1 for items in ARTWORK_DATA.values() for i in items if i[13] == "P2")),
        ("P3 (Low)", sum(1 for items in ARTWORK_DATA.values() for i in items if i[13] == "P3")),
        ("AI Try-on High Challenge", sum(1 for items in ARTWORK_DATA.values() for i in items if i[10] == "High")),
        ("Print Quality: Challenging", sum(1 for items in ARTWORK_DATA.values() for i in items if i[9] == "Challenging")),
    ]
    for label, val in stats:
        ws.cell(row, 1, label).font = Font(name=FF, size=SZ, bold=True)
        ws.cell(row, 2, val).font = Font(name=FF, size=SZ, bold=True, color="1F4E79")
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2).border = BORDER
        row += 1

    wb.save(OUTPUT)
    total = idx - 1
    print(f"\n🎉 ARTWORK TEST DATA sheet added successfully!")
    print(f"   📁 Output: {os.path.basename(OUTPUT)}")
    print(f"   📊 Total artwork types: {total}")
    print(f"   📂 Categories: {len(ARTWORK_DATA)}")
    for cat, items in ARTWORK_DATA.items():
        print(f"     {cat}: {len(items)} types")
    print(f"\n   Priority breakdown:")
    for p in ["P0", "P1", "P2", "P3"]:
        count = sum(1 for items in ARTWORK_DATA.values() for i in items if i[13] == p)
        print(f"     {p}: {count}")


if __name__ == "__main__":
    main()
