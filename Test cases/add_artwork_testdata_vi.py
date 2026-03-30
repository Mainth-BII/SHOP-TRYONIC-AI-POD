"""
Thêm sheet 'DỮ LIỆU TEST ARTWORK' vào file v30 Excel (Tiếng Việt).
Sử dụng dữ liệu từ artwork_data_vi.py
"""
import openpyxl, os, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from artwork_data_vi import COLUMNS_VI, COL_WIDTHS, ARTWORK_DATA_VI

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
SRC = os.path.join(BASE, "TC_POD-TShirt-Platform_ExecutionSummary_v30_2026-03-26.xlsx")
today = datetime.date.today().strftime("%Y-%m-%d")
OUTPUT = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork_vi.xlsx")

FF = "Calibri"; SZ = 11
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
CAT_FONT = Font(name=FF, size=12, bold=True, color="1F4E79")
BODY_FONT = Font(name=FF, size=SZ)
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

CATEGORY_COLORS = {
    "🎨 NGHỆ THUẬT TRUYỀN THỐNG": PatternFill("solid", fgColor="F2DCDB"),
    "💻 NGHỆ THUẬT SỐ": PatternFill("solid", fgColor="DCE6F1"),
    "🤖 PHONG CÁCH AI TẠO SINH": PatternFill("solid", fgColor="E4DFEC"),
    "🌍 NGHỆ THUẬT VĂN HÓA / VÙNG MIỀN": PatternFill("solid", fgColor="EBF1DE"),
    "👕 THIẾT KẾ POD / ÁO THUN": PatternFill("solid", fgColor="FDE9D9"),
    "⚠️ CÁC TRƯỜNG HỢP THÁCH THỨC KỸ THUẬT": PatternFill("solid", fgColor="DAEEF3"),
}

WARN_FILL = PatternFill("solid", fgColor="FFEB9C")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")

# Group fills for 3 result groups
GRP_ARTWORK = PatternFill("solid", fgColor="D5E8D4")  # xanh lá nhạt
GRP_TRYON_NAM = PatternFill("solid", fgColor="DAE8FC")  # xanh dương nhạt
GRP_TRYON_NU = PatternFill("solid", fgColor="F8CECC")  # hồng nhạt
GRP_HDR_ARTWORK = PatternFill("solid", fgColor="82B366")  # xanh lá
GRP_HDR_NAM = PatternFill("solid", fgColor="6C8EBF")  # xanh dương
GRP_HDR_NU = PatternFill("solid", fgColor="B85450")  # đỏ nhạt


def main():
    if not os.path.exists(SRC):
        print(f"❌ Không tìm thấy file gốc: {SRC}")
        return

    wb = openpyxl.load_workbook(SRC)
    sheet_name = "DỮ LIỆU TEST ARTWORK"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # Header
    for ci, col_name in enumerate(COLUMNS_VI, 1):
        c = ws.cell(1, ci, col_name)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS[ci-1]

    ws.row_dimensions[1].height = 35
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS_VI))}1"
    ws.freeze_panes = "A2"
    
    # Color-code header row for 3 groups (shifted +3 due to restored metadata cols)
    for ci in [17, 18, 19]:
        ws.cell(1, ci).fill = GRP_HDR_ARTWORK
    for ci in [20, 21, 22]:
        ws.cell(1, ci).fill = GRP_HDR_NAM
    for ci in [23, 24, 25]:
        ws.cell(1, ci).fill = GRP_HDR_NU

    # 17. Result Data Validation
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_result = DataValidation(type="list", formula1='"Pass,Fail,N/A,Untested"', allow_blank=True)
    ws.add_data_validation(dv_result)
    
    row = 2
    idx = 1
    for category, items in ARTWORK_DATA_VI.items():
        cat_fill = CATEGORY_COLORS.get(category, PatternFill("solid", fgColor="D6E4F0"))
        c = ws.cell(row, 1, category)
        c.font = CAT_FONT
        c.fill = cat_fill
        for ci in range(1, len(COLUMNS_VI) + 1):
            ws.cell(row, ci).fill = cat_fill
            ws.cell(row, ci).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS_VI))
        ws.row_dimensions[row].height = 28
        row += 1

        for item in items:
            (art_type, sub_style, desc, origin, color_complex, detail_lv,
             transparency, file_fmt, print_tech, print_qual,
             ai_tryon_challenge, test_focus, sample_prompt, priority, notes) = item

            ws.cell(row, 1, idx).font = BODY_FONT
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')
            ws.cell(row, 2, category.split(' ', 1)[-1] if ' ' in category else category).font = BODY_FONT
            ws.cell(row, 3, art_type).font = Font(name=FF, size=SZ, bold=True)
            ws.cell(row, 4, sub_style).font = BODY_FONT
            ws.cell(row, 5, desc).font = BODY_FONT
            ws.cell(row, 6, origin).font = BODY_FONT
            ws.cell(row, 7, color_complex).font = BODY_FONT
            ws.cell(row, 8, detail_lv).font = BODY_FONT
            ws.cell(row, 9, transparency).font = BODY_FONT
            ws.cell(row, 10, file_fmt).font = BODY_FONT
            ws.cell(row, 11, print_tech).font = BODY_FONT
            ws.cell(row, 12, print_qual).font = BODY_FONT
            ws.cell(row, 13, ai_tryon_challenge).font = BODY_FONT
            ws.cell(row, 14, test_focus).font = BODY_FONT
            ws.cell(row, 15, sample_prompt).font = Font(name=FF, size=SZ, italic=True, color="555555")
            ws.cell(row, 16, priority).font = Font(name=FF, size=SZ, bold=True)
            
            # === GROUP 1: Tạo Artwork (col 17-19) === xanh lá
            res_art = ws.cell(row, 17, "")  # [Result]- Tạo Artwork
            res_art.font = BODY_FONT
            res_art.fill = GRP_ARTWORK
            dv_result.add(res_art)
            
            c18 = ws.cell(row, 18, "")  # Đánh giá kết quả Tạo Artwork
            c18.font = BODY_FONT
            c18.fill = GRP_ARTWORK
            
            c19 = ws.cell(row, 19, "")  # Link data kết quả Tạo Artwork
            c19.font = BODY_FONT
            c19.fill = GRP_ARTWORK
            
            # === GROUP 2: Try-on Nam (col 20-22) === xanh dương
            res_male = ws.cell(row, 20, "")  # [Result] Try-on Model AI (Nam)
            res_male.font = BODY_FONT
            res_male.fill = GRP_TRYON_NAM
            dv_result.add(res_male)
            
            c21 = ws.cell(row, 21, "")  # Đánh giá Model AI (Nam)
            c21.font = BODY_FONT
            c21.fill = GRP_TRYON_NAM
            
            c22 = ws.cell(row, 22, "")  # Link data kết quả Tryon (Nam)
            c22.font = BODY_FONT
            c22.fill = GRP_TRYON_NAM
            
            # === GROUP 3: Try-on Nữ (col 23-25) === hồng
            res_female = ws.cell(row, 23, "")  # [Result] Try-on Model AI (Nữ)
            res_female.font = BODY_FONT
            res_female.fill = GRP_TRYON_NU
            dv_result.add(res_female)
            
            c24 = ws.cell(row, 24, "")  # Đánh giá Model AI (Nữ)
            c24.font = BODY_FONT
            c24.fill = GRP_TRYON_NU
            
            c25 = ws.cell(row, 25, "")  # Link data kết quả Tryon (Nữ)
            c25.font = BODY_FONT
            c25.fill = GRP_TRYON_NU
            
            # === Col 26: Link Ảnh Sample ===
            tc_key = f"TC_{idx:03d}"
            safe_type = "".join(c if c.isalnum() else "_" for c in art_type).strip("_")
            safe_sub = "".join(c if c.isalnum() else "_" for c in sub_style).strip("_")
            filename = f"{tc_key}_{safe_type}_{safe_sub}.jpg"
            abs_path = os.path.join(BASE, "Artwork_Samples_Real", filename)
            
            link_cell = ws.cell(row, 26, f"📂 {tc_key}")
            link_cell.hyperlink = abs_path
            link_cell.font = Font(name=FF, size=11, color="0563C1", underline="single")
            link_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Priority coloring
            p_cell = ws.cell(row, 16)
            if priority == "P0":
                p_cell.fill = PatternFill("solid", fgColor="FF6B6B")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
            elif priority == "P1":
                p_cell.fill = PatternFill("solid", fgColor="FFA500")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
            elif priority == "P2":
                p_cell.fill = WARN_FILL
            elif priority == "P3":
                p_cell.fill = PatternFill("solid", fgColor="E2EFDA")

            # AI Try-on Challenge coloring
            ai_cell = ws.cell(row, 13)
            if ai_tryon_challenge == "Cao":
                ai_cell.fill = FAIL_FILL
                ai_cell.font = Font(name=FF, size=SZ, bold=True, color="9C0006")
            elif ai_tryon_challenge == "Trung bình":
                ai_cell.fill = WARN_FILL

            # Print Quality coloring
            pq_cell = ws.cell(row, 12)
            if print_qual == "Thách thức":
                pq_cell.fill = FAIL_FILL

            for ci in range(1, len(COLUMNS_VI) + 1):
                ws.cell(row, ci).alignment = WRAP
                ws.cell(row, ci).border = BORDER

            row += 1
            idx += 1

    # Summary
    row += 1
    ws.cell(row, 1, "📊 THỐNG KÊ TỔNG HỢP").font = Font(name=FF, size=14, bold=True, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    stats = [
        ("Tổng loại Artwork", idx - 1),
        ("Số danh mục", len(ARTWORK_DATA_VI)),
        ("P0 (Quan trọng)", sum(1 for items in ARTWORK_DATA_VI.values() for i in items if i[13] == "P0")),
        ("P1 (Cao)", sum(1 for items in ARTWORK_DATA_VI.values() for i in items if i[13] == "P1")),
        ("P2 (Trung bình)", sum(1 for items in ARTWORK_DATA_VI.values() for i in items if i[13] == "P2")),
        ("P3 (Thấp)", sum(1 for items in ARTWORK_DATA_VI.values() for i in items if i[13] == "P3")),
        ("AI Try-on thách thức cao", sum(1 for items in ARTWORK_DATA_VI.values() for i in items if i[10] == "Cao")),
        ("Chất lượng in: Thách thức", sum(1 for items in ARTWORK_DATA_VI.values() for i in items if i[9] == "Thách thức")),
    ]
    for label, val in stats:
        ws.cell(row, 1, label).font = Font(name=FF, size=SZ, bold=True)
        ws.cell(row, 2, val).font = Font(name=FF, size=SZ, bold=True, color="1F4E79")
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2).border = BORDER
        row += 1

    wb.save(OUTPUT)
    total = idx - 1
    print(f"\n🎉 Đã thêm sheet 'DỮ LIỆU TEST ARTWORK' thành công!")
    print(f"   📁 File: {os.path.basename(OUTPUT)}")
    print(f"   📊 Tổng loại artwork: {total}")
    print(f"   📂 Danh mục: {len(ARTWORK_DATA_VI)}")
    for cat, items in ARTWORK_DATA_VI.items():
        print(f"     {cat}: {len(items)} loại")
    print(f"\n   Phân bổ ưu tiên:")
    for p in ["P0", "P1", "P2", "P3"]:
        count = sum(1 for items in ARTWORK_DATA_VI.values() for i in items if i[13] == p)
        print(f"     {p}: {count}")


if __name__ == "__main__":
    main()
