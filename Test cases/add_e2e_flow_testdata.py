"""
Thêm sheet 'E2E FLOW — LUỒNG ĐẦY ĐỦ' vào file v30 Excel.
Mỗi row = 1 kịch bản full flow từ đầu đến cuối.
"""
import openpyxl, os, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from e2e_flow_data_vi import COLUMNS_E2E, COL_WIDTHS_E2E, E2E_FLOW_DATA

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")

ARTWORK_FILE = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork_vi.xlsx")
SRC_FILE = os.path.join(BASE, "TC_POD-TShirt-Platform_ExecutionSummary_v30_2026-03-26.xlsx")
SRC = ARTWORK_FILE if os.path.exists(ARTWORK_FILE) else SRC_FILE
OUTPUT = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork_vi.xlsx")

FF = "Calibri"; SZ = 11
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
CAT_FONT = Font(name=FF, size=12, bold=True, color="1F4E79")
BODY_FONT = Font(name=FF, size=SZ)
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

CATEGORY_COLORS = {
    "✅ HAPPY PATH — LUỒNG CHÍNH": PatternFill("solid", fgColor="D5E8D4"),
    "🔄 LUỒNG SỬA ARTWORK — LOOP DESIGN": PatternFill("solid", fgColor="E4DFEC"),
    "💳 LUỒNG CHECKOUT — ĐĂNG NHẬP": PatternFill("solid", fgColor="DAE8FC"),
    "⚙️ LUỒNG CHỈNH SỬA HÌNH ẢNH NÂNG CAO": PatternFill("solid", fgColor="DAEEF3"),
    "📱 LUỒNG MOBILE / RESPONSIVE": PatternFill("solid", fgColor="FDE9D9"),
    "⚠️ LUỒNG NEGATIVE / GIÁN ĐOẠN": PatternFill("solid", fgColor="F8CECC"),
    "🔥 LUỒNG ĐẶC BIỆT": PatternFill("solid", fgColor="FFF2CC"),
}

FLOW_TYPE_COLORS = {
    "Happy Path": PatternFill("solid", fgColor="D5E8D4"),
    "Sửa tiếp": PatternFill("solid", fgColor="E4DFEC"),
    "Sửa với AI": PatternFill("solid", fgColor="E1D5E7"),
    "Chọn mẫu": PatternFill("solid", fgColor="DAEEF3"),
    "Kết hợp": PatternFill("solid", fgColor="FFF2CC"),
    "Checkout Login": PatternFill("solid", fgColor="DAE8FC"),
    "Multi-item": PatternFill("solid", fgColor="FDE9D9"),
    "Chỉnh sửa": PatternFill("solid", fgColor="DAEEF3"),
    "Mobile": PatternFill("solid", fgColor="FDE9D9"),
    "Tablet": PatternFill("solid", fgColor="FDE9D9"),
    "Negative": PatternFill("solid", fgColor="F8CECC"),
    "Đặc biệt": PatternFill("solid", fgColor="FFF2CC"),
}

EXPECTED_HDR = PatternFill("solid", fgColor="82B366")
ACTUAL_HDR = PatternFill("solid", fgColor="6C8EBF")
EXPECTED_FILL = PatternFill("solid", fgColor="D5E8D4")
ACTUAL_FILL = PatternFill("solid", fgColor="DAE8FC")


def main():
    if not os.path.exists(SRC):
        print(f"❌ Không tìm thấy file gốc: {SRC}")
        return

    wb = openpyxl.load_workbook(SRC)
    sheet_name = "E2E FLOW — LUỒNG ĐẦY ĐỦ"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # Header
    for ci, col_name in enumerate(COLUMNS_E2E, 1):
        c = ws.cell(1, ci, col_name)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS_E2E[ci-1]

    # Color-code Result headers
    ws.cell(1, 9).fill = EXPECTED_HDR
    ws.cell(1, 10).fill = ACTUAL_HDR

    ws.row_dimensions[1].height = 35
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS_E2E))}1"
    ws.freeze_panes = "A2"

    # Data Validations
    dv_result = DataValidation(type="list", formula1='"Pass,Fail,N/A,Untested"', allow_blank=True)
    ws.add_data_validation(dv_result)

    row = 2
    idx = 1
    for category, items in E2E_FLOW_DATA.items():
        cat_fill = CATEGORY_COLORS.get(category, PatternFill("solid", fgColor="D6E4F0"))
        c = ws.cell(row, 1, category)
        c.font = CAT_FONT
        c.fill = cat_fill
        for ci in range(1, len(COLUMNS_E2E) + 1):
            ws.cell(row, ci).fill = cat_fill
            ws.cell(row, ci).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS_E2E))
        ws.row_dimensions[row].height = 28
        row += 1

        for item in items:
            (tc_id, flow_name, flow_type, steps, input_data, expected, priority, notes) = item

            # STT
            ws.cell(row, 1, idx).font = BODY_FONT
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')

            # TC_ID
            ws.cell(row, 2, tc_id).font = Font(name=FF, size=SZ, bold=True, color="1F4E79")

            # Tên luồng
            ws.cell(row, 3, flow_name).font = Font(name=FF, size=SZ, bold=True)

            # Loại luồng — color-coded
            ft_cell = ws.cell(row, 4, flow_type)
            ft_cell.font = Font(name=FF, size=SZ, bold=True)
            ft_cell.alignment = CENTER
            ft_fill = FLOW_TYPE_COLORS.get(flow_type, None)
            if ft_fill:
                ft_cell.fill = ft_fill

            # Các bước (multi-line)
            ws.cell(row, 5, steps).font = BODY_FONT

            # Dữ liệu đầu vào
            ws.cell(row, 6, input_data).font = Font(name=FF, size=SZ, italic=True, color="555555")

            # Kết quả mong đợi
            ws.cell(row, 7, expected).font = BODY_FONT

            # Độ ưu tiên
            p_cell = ws.cell(row, 8, priority)
            p_cell.font = Font(name=FF, size=SZ, bold=True)
            p_cell.alignment = CENTER
            if priority == "P0":
                p_cell.fill = PatternFill("solid", fgColor="FF6B6B")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
            elif priority == "P1":
                p_cell.fill = PatternFill("solid", fgColor="FFA500")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")

            # Expected_Result
            er_cell = ws.cell(row, 9, "")
            er_cell.font = BODY_FONT
            er_cell.fill = EXPECTED_FILL
            er_cell.alignment = CENTER
            dv_result.add(er_cell)

            # Actual_Result
            ar_cell = ws.cell(row, 10, "")
            ar_cell.font = BODY_FONT
            ar_cell.fill = ACTUAL_FILL
            ar_cell.alignment = CENTER
            dv_result.add(ar_cell)

            # Thời gian thực thi
            ws.cell(row, 11, "").font = BODY_FONT

            # Screenshot
            ws.cell(row, 12, "").font = BODY_FONT

            # Ghi chú
            ws.cell(row, 13, notes).font = BODY_FONT

            # Chiều cao dòng lớn hơn vì steps multi-line
            line_count = steps.count('\n') + 1
            ws.row_dimensions[row].height = max(15 * line_count, 60)

            for ci in range(1, len(COLUMNS_E2E) + 1):
                ws.cell(row, ci).alignment = Alignment(
                    horizontal=ws.cell(row, ci).alignment.horizontal or 'left',
                    vertical='top',
                    wrap_text=True
                )
                ws.cell(row, ci).border = BORDER

            row += 1
            idx += 1

    # Summary
    row += 1
    ws.cell(row, 1, "📊 THỐNG KÊ E2E FLOW").font = Font(name=FF, size=14, bold=True, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    total = idx - 1
    stats = [
        ("Tổng luồng E2E", total),
        ("Số nhóm", len(E2E_FLOW_DATA)),
        ("P0 (Critical)", sum(1 for items in E2E_FLOW_DATA.values() for i in items if i[6] == "P0")),
        ("P1 (High)", sum(1 for items in E2E_FLOW_DATA.values() for i in items if i[6] == "P1")),
    ]
    for label, val in stats:
        ws.cell(row, 1, label).font = Font(name=FF, size=SZ, bold=True)
        ws.cell(row, 2, val).font = Font(name=FF, size=SZ, bold=True, color="1F4E79")
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2).border = BORDER
        row += 1

    wb.save(OUTPUT)
    print(f"\n🎉 Đã thêm sheet '{sheet_name}' thành công!")
    print(f"   📁 File: {os.path.basename(OUTPUT)}")
    print(f"   📊 Tổng luồng E2E: {total}")
    print(f"   📂 Nhóm: {len(E2E_FLOW_DATA)}")
    for cat, items in E2E_FLOW_DATA.items():
        print(f"     {cat}: {len(items)} luồng")
    print(f"\n   Phân bổ ưu tiên:")
    for p in ["P0", "P1"]:
        count = sum(1 for items in E2E_FLOW_DATA.values() for i in items if i[6] == p)
        print(f"     {p}: {count}")


if __name__ == "__main__":
    main()
