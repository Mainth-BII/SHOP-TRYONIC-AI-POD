"""
Thêm sheet 'GỢI Ý SIZE - TEST DATA' vào file v30 Excel (Tiếng Việt).
"""
import openpyxl, os, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from size_suggestion_data_vi import COLUMNS_SIZE, COL_WIDTHS_SIZE, SIZE_TEST_DATA

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")

# Tìm file artwork_vi mới nhất hoặc file gốc
ARTWORK_FILE = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork_vi.xlsx")
SRC_FILE = os.path.join(BASE, "TC_POD-TShirt-Platform_ExecutionSummary_v30_2026-03-26.xlsx")
SRC = ARTWORK_FILE if os.path.exists(ARTWORK_FILE) else SRC_FILE
OUTPUT = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork_vi.xlsx")

FF = "Calibri"; SZ = 11
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
CAT_FONT = Font(name=FF, size=12, bold=True, color="1F4E79")
BODY_FONT = Font(name=FF, size=SZ)
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

CATEGORY_COLORS = {
    "📏 TRƯỜNG HỢP CHUẨN — NAM": PatternFill("solid", fgColor="DAE8FC"),
    "📏 TRƯỜNG HỢP CHUẨN — NỮ": PatternFill("solid", fgColor="F8CECC"),
    "⚖️ RANH GIỚI SIZE (Boundary)": PatternFill("solid", fgColor="FFF2CC"),
    "🏋️ CHIỀU CAO THẤP + CÂN NẶNG CAO (Overweight)": PatternFill("solid", fgColor="FFE6CC"),
    "🦒 CHIỀU CAO LỚN + CÂN NẶNG THẤP (Underweight)": PatternFill("solid", fgColor="D5E8D4"),
    "🚫 GIÁ TRỊ INVALID / EDGE CASE": PatternFill("solid", fgColor="F8CECC"),
    "🔄 CHUYỂN ĐỔI GIỚI TÍNH (Cùng chiều cao/cân nặng)": PatternFill("solid", fgColor="E1D5E7"),
    "📐 KIỂM TRA HIỂN THỊ UI GỢI Ý SIZE": PatternFill("solid", fgColor="DAEEF3"),
}

# Result column fills
RESULT_FILL = PatternFill("solid", fgColor="D5E8D4")  # xanh lá nhạt
ACTUAL_FILL = PatternFill("solid", fgColor="DAE8FC")   # xanh dương nhạt
RESULT_HDR = PatternFill("solid", fgColor="82B366")
ACTUAL_HDR = PatternFill("solid", fgColor="6C8EBF")


def main():
    if not os.path.exists(SRC):
        print(f"❌ Không tìm thấy file gốc: {SRC}")
        return

    wb = openpyxl.load_workbook(SRC)
    sheet_name = "GỢI Ý SIZE - TEST DATA"

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]

    ws = wb.create_sheet(sheet_name)

    # Header
    for ci, col_name in enumerate(COLUMNS_SIZE, 1):
        c = ws.cell(1, ci, col_name)
        c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS_SIZE[ci-1]

    # Color-code Result columns header
    ws.cell(1, 10).fill = RESULT_HDR  # Expected_Result
    ws.cell(1, 11).fill = ACTUAL_HDR  # Actual_Result

    ws.row_dimensions[1].height = 35
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS_SIZE))}1"
    ws.freeze_panes = "A2"

    # Data Validation for Result columns
    dv_result = DataValidation(type="list", formula1='"Pass,Fail,N/A,Untested"', allow_blank=True)
    ws.add_data_validation(dv_result)

    # Gender validation
    dv_gender = DataValidation(type="list", formula1='"Nam,Nữ"', allow_blank=True)
    ws.add_data_validation(dv_gender)

    row = 2
    idx = 1
    for category, items in SIZE_TEST_DATA.items():
        cat_fill = CATEGORY_COLORS.get(category, PatternFill("solid", fgColor="D6E4F0"))
        c = ws.cell(row, 1, category)
        c.font = CAT_FONT
        c.fill = cat_fill
        for ci in range(1, len(COLUMNS_SIZE) + 1):
            ws.cell(row, ci).fill = cat_fill
            ws.cell(row, ci).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS_SIZE))
        ws.row_dimensions[row].height = 28
        row += 1

        for item in items:
            (tc_id, gender, height, weight, expected_size, desc, priority, notes) = item

            ws.cell(row, 1, idx).font = BODY_FONT
            ws.cell(row, 1).alignment = Alignment(horizontal='center', vertical='top')

            ws.cell(row, 2, category.split(' ', 1)[-1] if ' ' in category else category).font = BODY_FONT

            ws.cell(row, 3, tc_id).font = Font(name=FF, size=SZ, bold=True, color="1F4E79")

            # Giới tính (with validation)
            g_cell = ws.cell(row, 4, gender)
            g_cell.font = Font(name=FF, size=SZ, bold=True)
            g_cell.alignment = CENTER
            if gender == "Nam":
                g_cell.fill = PatternFill("solid", fgColor="DAE8FC")
            elif gender == "Nữ":
                g_cell.fill = PatternFill("solid", fgColor="F8CECC")
            dv_gender.add(g_cell)

            # Chiều cao
            h_cell = ws.cell(row, 5, height)
            h_cell.font = BODY_FONT
            h_cell.alignment = CENTER

            # Cân nặng
            w_cell = ws.cell(row, 6, weight)
            w_cell.font = BODY_FONT
            w_cell.alignment = CENTER

            # Size kỳ vọng
            sz_cell = ws.cell(row, 7, expected_size)
            sz_cell.font = Font(name=FF, size=SZ, bold=True, color="1F4E79")
            sz_cell.alignment = CENTER

            # Mô tả
            ws.cell(row, 8, desc).font = BODY_FONT

            # Độ ưu tiên
            p_cell = ws.cell(row, 9, priority)
            p_cell.font = Font(name=FF, size=SZ, bold=True)
            p_cell.alignment = CENTER
            if priority == "P0":
                p_cell.fill = PatternFill("solid", fgColor="FF6B6B")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
            elif priority == "P1":
                p_cell.fill = PatternFill("solid", fgColor="FFA500")
                p_cell.font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")

            # Expected_Result
            er_cell = ws.cell(row, 10, "")
            er_cell.font = BODY_FONT
            er_cell.fill = RESULT_FILL
            er_cell.alignment = CENTER
            dv_result.add(er_cell)

            # Actual_Result
            ar_cell = ws.cell(row, 11, "")
            ar_cell.font = BODY_FONT
            ar_cell.fill = ACTUAL_FILL
            ar_cell.alignment = CENTER
            dv_result.add(ar_cell)

            # Ghi chú
            ws.cell(row, 12, notes).font = BODY_FONT

            for ci in range(1, len(COLUMNS_SIZE) + 1):
                ws.cell(row, ci).alignment = Alignment(
                    horizontal=ws.cell(row, ci).alignment.horizontal or 'left',
                    vertical='top',
                    wrap_text=True
                )
                ws.cell(row, ci).border = BORDER

            row += 1
            idx += 1

    # Summary
    row += 1
    ws.cell(row, 1, "📊 THỐNG KÊ").font = Font(name=FF, size=14, bold=True, color="1F4E79")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 1

    total = idx - 1
    stats = [
        ("Tổng test cases", total),
        ("Số danh mục", len(SIZE_TEST_DATA)),
        ("P0 (Quan trọng)", sum(1 for items in SIZE_TEST_DATA.values() for i in items if i[6] == "P0")),
        ("P1 (Cao)", sum(1 for items in SIZE_TEST_DATA.values() for i in items if i[6] == "P1")),
        ("Test cases Nam", sum(1 for items in SIZE_TEST_DATA.values() for i in items if i[1] == "Nam")),
        ("Test cases Nữ", sum(1 for items in SIZE_TEST_DATA.values() for i in items if i[1] == "Nữ")),
    ]
    for label, val in stats:
        ws.cell(row, 1, label).font = Font(name=FF, size=SZ, bold=True)
        ws.cell(row, 2, val).font = Font(name=FF, size=SZ, bold=True, color="1F4E79")
        ws.cell(row, 1).border = BORDER
        ws.cell(row, 2).border = BORDER
        row += 1

    wb.save(OUTPUT)
    print(f"\n🎉 Đã thêm sheet '{sheet_name}' thành công!")
    print(f"   📁 File: {os.path.basename(OUTPUT)}")
    print(f"   📊 Tổng test cases: {total}")
    print(f"   📂 Danh mục: {len(SIZE_TEST_DATA)}")
    for cat, items in SIZE_TEST_DATA.items():
        print(f"     {cat}: {len(items)} cases")
    print(f"\n   Phân bổ ưu tiên:")
    for p in ["P0", "P1"]:
        count = sum(1 for items in SIZE_TEST_DATA.values() for i in items if i[6] == p)
        print(f"     {p}: {count}")


if __name__ == "__main__":
    main()
