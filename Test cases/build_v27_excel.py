"""Build v27 Excel — FIXED: line breaks + Screen column."""
import openpyxl, os, re, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
MD = os.path.join(BASE, "test_cases_suite_v27.md")
today = datetime.date.today().strftime("%Y-%m-%d")
OUTPUT = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v27_{today}_v3.xlsx")
VER = 27

FF = "Calibri"; SZ = 11
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
CAT_FILL = PatternFill("solid", fgColor="D6E4F0")
CAT_FONT = Font(name=FF, size=12, bold=True, color="1F4E79")
BODY_FONT = Font(name=FF, size=SZ)
THIN = Side(style='thin', color='B0B0B0')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
REVIEW_FILL = PatternFill("solid", fgColor="FFC000")
SCREEN_FILL = PatternFill("solid", fgColor="2E75B6")

# ─── SCREEN MAPPING: Module keyword → Screen name ───
SCREEN_MAP_BY_FEATURE = {
    "HOME": "MH Trang chủ",
    "ORDER": "MH Đơn hàng",
    "E2E FLOW": "E2E",
}

def detect_screen(feature, module, tc_id, title):
    """Detect screen from feature + module + title context."""
    mod = module.lower()
    ttl = title.lower()

    # Feature-level defaults
    if feature in SCREEN_MAP_BY_FEATURE:
        return SCREEN_MAP_BY_FEATURE[feature]

    # LOGIN sheet
    if feature == "LOGIN":
        if any(k in mod for k in ['đăng ký', 'signup', 'registration']):
            return "MH Đăng ký"
        if any(k in mod for k in ['đăng nhập', 'login']):
            return "MH Đăng nhập"
        if any(k in mod for k in ['tài khoản', 'account']):
            return "MH Tài khoản"
        if any(k in mod for k in ['hồ sơ', 'profile']):
            return "MH Hồ sơ"
        if any(k in mod for k in ['guest']):
            return "MH Guest"
        if 'responsive' in mod:
            if 'login' in ttl or 'đăng nhập' in ttl:
                return "MH Đăng nhập"
            if 'signup' in ttl or 'đăng ký' in ttl or 'registration' in ttl:
                return "MH Đăng ký"
            return "MH Đăng nhập"
        return "MH Đăng nhập"

    # DESIGN STUDIO sheet
    if feature == "DESIGN STUDIO":
        if any(k in mod for k in ['ds sản phẩm', 'sản phẩm', 'chọn sản phẩm']):
            return "DS - Tab Sản phẩm"
        if any(k in mod for k in ['ds ảnh', 'upload', 'cài đặt hình ảnh']):
            return "DS - Tab Ảnh của bạn"
        if any(k in mod for k in ['ds thư viện', 'thư viện', 'tìm kiếm mẫu']):
            return "DS - Tab Thư viện"
        if any(k in mod for k in ['gợi ý size']):
            return "DS - Popup Gợi ý size"
        if any(k in mod for k in ['gallery']):
            return "DS - Gallery"
        if any(k in mod for k in ['smart fit']):
            return "DS - Smart Fit"
        if any(k in mod for k in ['editor', 'zoom']):
            return "DS - Editor/Canvas"
        if any(k in mod for k in ['canvas', 'toolbar']):
            return "DS - Canvas"
        if any(k in mod for k in ['bottom bar']):
            return "DS - Bottom Bar"
        if any(k in mod for k in ['sidebar']):
            return "DS - Sidebar"
        if 'responsive' in mod:
            return "DS - Responsive"
        return "DS - Chung"

    # AI GENERATE sheet
    if feature == "AI GENERATE":
        if any(k in mod for k in ['credit']):
            return "MH Credits"
        if any(k in mod for k in ['tạo ảnh ai', 'ai gen', 'ai generate']):
            return "DS - Tab Tạo ảnh AI"
        if 'responsive' in mod:
            return "DS - Tab Tạo ảnh AI"
        return "DS - Tab Tạo ảnh AI"

    # ĐẶT HÀNG sheet
    if feature == "ĐẶT HÀNG":
        if any(k in mod for k in ['giỏ hàng', 'cart']):
            return "MH Giỏ hàng"
        if any(k in mod for k in ['sản phẩm', 'product']):
            return "MH Chi tiết SP"
        if any(k in mod for k in ['đặt hàng', 'thanh toán']):
            return "DS - Tab Đặt hàng"
        if any(k in mod for k in ['header']):
            return "Header"
        if 'responsive' in mod:
            return "DS - Tab Đặt hàng"
        return "DS - Tab Đặt hàng"

    # THANH TOÁN sheet
    if feature == "THANH TOÁN":
        if any(k in mod for k in ['checkout']):
            return "MH Checkout"
        if any(k in mod for k in ['thanh toán', 'payment']):
            return "MH Thanh toán"
        if 'responsive' in mod:
            return "MH Checkout"
        return "MH Checkout"

    return feature

def fix_linebreaks(text):
    """Convert <br>, <br/>, <br /> to real newlines for Excel."""
    if not text:
        return text
    result = re.sub(r'<br\s*/?>', '\n', text)
    # Also add newline after period followed by space and a number (step numbering)
    result = re.sub(r'\.\s+(\d+\.)', r'.\n\1', result)
    return result

# ─── Parse MD ───
with open(MD, 'r', encoding='utf-8') as f:
    content = f.read()

features = {}
lines = content.split('\n')
cur_feat = None
cur_cat = None

for line in lines:
    s = line.strip()
    m = re.match(r'^## .+Feature:\s*(.+)$', s)
    if m:
        cur_feat = m.group(1).strip()
        cur_cat = None
        if cur_feat not in features:
            features[cur_feat] = {}
        continue
    m2 = re.match(r'^###\s+📌\s+(.+)$', s)
    if m2:
        cur_cat = m2.group(1).strip()
        if cur_feat and cur_cat not in features[cur_feat]:
            features[cur_feat][cur_cat] = []
        continue
    if s.startswith('| `TC_') and cur_feat and cur_cat:
        cols = [c.strip() for c in s.split('|')[1:-1]]
        if len(cols) >= 7:
            features[cur_feat][cur_cat].append(cols)

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ─── COVER PAGE ───
ws = wb.create_sheet("Cover Page")
ws.sheet_properties.showGridLines = False
cover_data = [
    ("Company:", "Tryonic AI Platform"),
    ("Project:", "POD T-Shirt Platform"),
    ("Document:", f"Test Case Execution Summary v{VER}"),
    ("Version:", f"v{VER}"),
    ("Date:", today),
    ("Structure:", "8 Feature Sheets (Consolidated)")
]
for i, (label, val) in enumerate(cover_data, 3):
    ws.cell(i, 2, label).font = Font(name=FF, size=14, bold=True, color="1F4E79")
    ws.cell(i, 3, val).font = Font(name=FF, size=14)
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 50

# ─── FEATURE SHEETS (with Screen column + line break fix) ───
TC_COLS = ["TC_ID", "Screen", "US_Mapping", "Module", "Title", "Type", "Priority", "Pre-condition", "Steps", "Expected Result",
           "Actual Result",
           "Action Type", "Create TCs Type", "Execution Type",
           "Result_R1", "Test Date_R1", "Tester_R1", "Bug ID_R1", "Bug Desc_R1",
           "Result_R2", "Test Date_R2", "Tester_R2", "Bug ID_R2", "Bug Desc_R2",
           "Evidence", "Notes", "Review_Manual (Feedback)"]
WIDTHS = [16, 22, 12, 22, 35, 14, 10, 30, 45, 40, 40, 12, 14, 14, 10, 12, 10, 14, 35, 10, 12, 10, 14, 35, 30, 15, 35]

dv_action = DataValidation(type="list", formula1='"Add new,Update,Delete"', allow_blank=True)
dv_create = DataValidation(type="list", formula1='"By AI,By Manual"', allow_blank=True)
dv_exec = DataValidation(type="list", formula1='"Auto,Manual"', allow_blank=True)
dv_result = DataValidation(type="list", formula1='"Untested,Pass,Fail,N/A"', allow_blank=True)

sheet_info = {}

for feat_name, categories in features.items():
    safe_name = feat_name[:31]
    ws = wb.create_sheet(safe_name)
    ws.add_data_validation(dv_action)
    ws.add_data_validation(dv_create)
    ws.add_data_validation(dv_exec)
    ws.add_data_validation(dv_result)
    
    # Header
    for ci, col_name in enumerate(TC_COLS, 1):
        c = ws.cell(1, ci, col_name)
        if col_name == "Screen":
            c.fill = SCREEN_FILL
        elif col_name.startswith("Review"):
            c.fill = REVIEW_FILL
        else:
            c.fill = HDR_FILL
        c.font = HDR_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS[ci-1]
    
    ws.row_dimensions[1].height = 30
    ws.auto_filter.ref = f"A1:{get_column_letter(len(TC_COLS))}1"
    ws.freeze_panes = "A2"
    
    row = 2
    first_tc_row = None
    
    cat_order = ["UI/UX", "Validation", "Functional", "Functional (Logic & Behavior)", "Security", "Performance"]
    all_cats = list(categories.keys())
    ordered = [c for c in cat_order if c in all_cats] + [c for c in all_cats if c not in cat_order]
    
    for cat in ordered:
        tcs = categories[cat]
        if not tcs:
            continue
        
        # Category header row
        c = ws.cell(row, 1, f"📌 {cat}")
        c.font = CAT_FONT
        c.fill = CAT_FILL
        for ci in range(1, len(TC_COLS)+1):
            ws.cell(row, ci).fill = CAT_FILL
            ws.cell(row, ci).border = BORDER
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(TC_COLS))
        row += 1
        
        for cols in tcs:
            if first_tc_row is None:
                first_tc_row = row
            
            # cols: [TC_ID, Mapping, Module, Title, Type, Priority, Steps, Expected]
            tc_id = cols[0].strip('` ')
            mapping = cols[1].strip('` ')
            module = cols[2]
            title = cols[3]
            tc_type = cols[4]
            priority = cols[5]
            steps_raw = cols[6] if len(cols) >= 7 else ''
            expected_raw = cols[7] if len(cols) >= 8 else ''

            # Fix line breaks: <br> → \n
            steps = fix_linebreaks(steps_raw)
            expected = fix_linebreaks(expected_raw)

            # Extract pre-condition from steps (lines starting with "Precondition:")
            precondition = ''
            if steps_raw.strip().lower().startswith('precondition'):
                parts = steps_raw.split('\n', 1)
                if len(parts) == 1:
                    # Try splitting by <br> first then by numbered step
                    parts2 = re.split(r'<br\s*/?>', steps_raw, maxsplit=1)
                    if len(parts2) >= 2:
                        precondition = fix_linebreaks(parts2[0].strip())
                        steps = fix_linebreaks(parts2[1].strip())
                    else:
                        precondition = fix_linebreaks(steps_raw)
                        steps = ''
                else:
                    precondition = fix_linebreaks(parts[0].strip())
                    steps = fix_linebreaks(parts[1].strip())

            # Detect screen
            screen = detect_screen(feat_name, module, tc_id, title)

            # Write to cells (col 8 = Pre-condition, col 9 = Steps, col 10 = Expected)
            ws.cell(row, 1, tc_id).font = BODY_FONT
            ws.cell(row, 2, screen).font = Font(name=FF, size=SZ, bold=True, color="2E75B6")
            ws.cell(row, 3, mapping).font = BODY_FONT
            ws.cell(row, 4, module).font = BODY_FONT
            ws.cell(row, 5, title).font = BODY_FONT
            ws.cell(row, 6, tc_type).font = BODY_FONT
            ws.cell(row, 7, priority).font = BODY_FONT
            ws.cell(row, 8, precondition).font = BODY_FONT
            ws.cell(row, 9, steps).font = BODY_FONT
            ws.cell(row, 10, expected).font = BODY_FONT

            for ci in range(1, 11):
                ws.cell(row, ci).alignment = WRAP
                ws.cell(row, ci).border = BORDER
            
            # col 11 = Actual Result (empty until test execution)
            ws.cell(row, 11, '').font = BODY_FONT
            # Default tracking values (shifted +1 for Actual Result)
            ws.cell(row, 12, "Add new").font = BODY_FONT
            ws.cell(row, 13, "By AI").font = BODY_FONT
            ws.cell(row, 14, "Manual").font = BODY_FONT
            ws.cell(row, 15, "Untested").font = BODY_FONT
            ws.cell(row, 20, "Untested").font = BODY_FONT
            
            dv_action.add(ws.cell(row, 12))
            dv_create.add(ws.cell(row, 13))
            dv_exec.add(ws.cell(row, 14))
            dv_result.add(ws.cell(row, 15))
            dv_result.add(ws.cell(row, 20))
            
            for ci in range(10, len(TC_COLS)+1):
                ws.cell(row, ci).border = BORDER
                ws.cell(row, ci).alignment = WRAP
                if not ws.cell(row, ci).font or ws.cell(row, ci).font == Font():
                    ws.cell(row, ci).font = BODY_FONT
            
            row += 1
    
    sheet_info[safe_name] = (first_tc_row or 3, row - 1, sum(len(t) for t in categories.values()))

# ─── EXECUTION SUMMARY ───
ws = wb.create_sheet("Execution Summary", 1)
ws.cell(1, 1, "Execution Summary — Test Progress Dashboard").font = Font(name=FF, size=16, bold=True, color="1F4E79")
ws.merge_cells("A1:H1")

headers = ["Feature", "Total TCs", "Pass", "Fail", "Untested", "N/A", "% Progress", "% Pass Rate"]
for ci, h in enumerate(headers, 1):
    c = ws.cell(3, ci, h)
    c.font = HDR_FONT
    c.fill = HDR_FILL
    c.alignment = Alignment(horizontal='center')
    c.border = BORDER

ws.column_dimensions['A'].width = 25
for col in 'BCDEFGH':
    ws.column_dimensions[col].width = 14

row = 4
for sname, (start_r, end_r, tc_count) in sheet_info.items():
    ws.cell(row, 1, sname).font = Font(name=FF, size=SZ, bold=True)
    ws.cell(row, 2, tc_count).font = BODY_FONT
    
    # Result_R1 is now column O (15) after adding Actual_Result
    col_letter = 'O'
    ws.cell(row, 3).value = f"=COUNTIF('{sname}'!{col_letter}{start_r}:{col_letter}{end_r},\"Pass\")"
    ws.cell(row, 4).value = f"=COUNTIF('{sname}'!{col_letter}{start_r}:{col_letter}{end_r},\"Fail\")"
    ws.cell(row, 5).value = f"=COUNTIF('{sname}'!{col_letter}{start_r}:{col_letter}{end_r},\"Untested\")"
    ws.cell(row, 6).value = f"=COUNTIF('{sname}'!{col_letter}{start_r}:{col_letter}{end_r},\"N/A\")"
    
    tested = f"(B{row}-E{row}-F{row})"
    ws.cell(row, 7).value = f"=IF(B{row}=0,0,{tested}/B{row})"
    ws.cell(row, 7).number_format = '0%'
    ws.cell(row, 8).value = f"=IF({tested}=0,0,C{row}/{tested})"
    ws.cell(row, 8).number_format = '0%'
    
    for ci in range(1, 9):
        ws.cell(row, ci).border = BORDER
        ws.cell(row, ci).font = BODY_FONT
        ws.cell(row, ci).alignment = Alignment(horizontal='center')
    
    row += 1

total_row = row
ws.cell(total_row, 1, "TOTAL").font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
ws.cell(total_row, 1).fill = PatternFill("solid", fgColor="1F4E79")
for ci in range(2, 7):
    ws.cell(total_row, ci).value = f"=SUM({get_column_letter(ci)}4:{get_column_letter(ci)}{total_row-1})"
    ws.cell(total_row, ci).font = Font(name=FF, size=SZ, bold=True, color="FFFFFF")
    ws.cell(total_row, ci).fill = PatternFill("solid", fgColor="1F4E79")
for ci in range(1, 9):
    ws.cell(total_row, ci).border = BORDER

wb.save(OUTPUT)
print(f"\n🎉 v{VER} Excel saved: {os.path.basename(OUTPUT)}")
print(f"   Sheets: Cover + Summary + {len(sheet_info)} features = {len(sheet_info)+2} total")
for s, (_, _, tc) in sheet_info.items():
    print(f"     📋 {s}: {tc} TCs")
