import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os
import re
import datetime
import glob

# Step 1: Find the latest CSV file
csv_pattern = "TC_POD-TShirt-Platform_v*_Full.csv"
csv_files = glob.glob(csv_pattern)

if not csv_files:
    # fallback to the hardcoded if pattern fails
    csv_path = "TC_POD-TShirt-Platform_v6_2026-03-16_Full.csv"
else:
    # Sort files to find the latest (assuming sort by name or time works)
    csv_files.sort(reverse=True)
    csv_path = csv_files[0]

if not os.path.exists(csv_path):
    print(f"File {csv_path} not found!")
    exit(1)

# Step 2: Determine the next version number for output Excel
# Find latest excel file to determine next version
excel_pattern = "TC_POD-TShirt-Platform_ExecutionSummary_v*_*.xlsx"
excel_files = glob.glob(excel_pattern)

next_version = 15 # default fallback
if excel_files:
    versions = []
    for f in excel_files:
        match = re.search(r'_v(\d+)_', f)
        if match:
            versions.append(int(match.group(1)))
    if versions:
        next_version = max(versions) + 1

today_str = datetime.datetime.now().strftime("%Y-%m-%d")
output_path = f"TC_POD-TShirt-Platform_ExecutionSummary_v{next_version}_{today_str}.xlsx"
print(f"Auto-incrementing version: Next version will be v{next_version}")


import datetime

df = pd.read_csv(csv_path)
df = df.fillna("")

def get_custom_category(row):
    module = str(row['Module']).lower()
    title = str(row['Title']).lower()
    tc_type = str(row['Type']).lower()

    if 'security' in module or 'lock' in title or 'xss' in title or 'injection' in title:
        return 'Security'
    elif 'network' in module or 'timeout' in title or 'rate limit' in title or 'performance' in module:
        return 'Performance'
    elif 'ui/ux' in tc_type or 'ui' in module:
        return 'UI/UX'
    elif tc_type in ['negative', 'boundary'] or 'validation' in title:
        return 'Validation'
    else:
        return 'Functional (Logic & Behavior)'

df['CustomCategory'] = df.apply(get_custom_category, axis=1)
categories_order = ['UI/UX', 'Validation', 'Functional (Logic & Behavior)', 'Security', 'Performance']

wb = openpyxl.Workbook()
default_sheet = wb.active

# --- GLOBAL STYLES ---
FONT_FAMILY = "Calibri"
STD_SIZE = 11

font_body = Font(name=FONT_FAMILY, size=STD_SIZE)
font_bold = Font(name=FONT_FAMILY, size=STD_SIZE, bold=True)
font_header = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")

# For cover page
font_cover_title = Font(name=FONT_FAMILY, size=26, bold=True, color="1F4E78")
font_cover_subtitle = Font(name=FONT_FAMILY, size=14, italic=True, color="595959")
font_cover_key = Font(name=FONT_FAMILY, size=12, bold=True, color="333333")
font_cover_val = Font(name=FONT_FAMILY, size=12, color="000000")

# Fills
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  
round_head_fill = PatternFill(start_color="3B73B9", end_color="3B73B9", fill_type="solid")  
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
summary_head_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid") 
total_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") 
cover_key_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# Borders
border_thin = Side(border_style="thin", color="BFBFBF")
full_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
cover_border = Border(bottom=Side(border_style="medium", color="1F4E78"))

# Alignments
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Data Validation definitions
dv_result = DataValidation(type="list", formula1='"Untested,Pass,Fail,N/A"', allow_blank=False)
dv_result.error = 'Your entry is not in the list'
dv_result.errorTitle = 'Invalid Entry'
dv_result.prompt = 'Please select from the list'
dv_result.promptTitle = 'Select Result'

dv_action = DataValidation(type="list", formula1='"Add new,Update,Delete"', allow_blank=False)
dv_auto = DataValidation(type="list", formula1='"By AI,By Manual"', allow_blank=False)

# ---------------------------------------------------------
# 1. GENERATE COVER PAGE
# ---------------------------------------------------------
cover_ws = wb.create_sheet(title="Cover Page", index=0)
cover_ws.sheet_view.showGridLines = False

# Decorator block
cover_ws.merge_cells("B2:E2")
cell_title = cover_ws.cell(row=2, column=2, value="TEST CASE")
cell_title.font = font_cover_title
cell_title.alignment = Alignment(horizontal="left", vertical="center")

cover_ws.merge_cells("B3:E3")
cell_sub = cover_ws.cell(row=3, column=2, value="Comprehensive Functional & UI/UX Test Suite")
cell_sub.font = font_cover_subtitle
cell_sub.alignment = Alignment(horizontal="left", vertical="center")
cell_sub.border = cover_border # Underline the header area

cover_data = [
    ("Company Name", "Tryonic Platform"),
    ("Project Name", "POD T-Shirt Platform"),
    ("Application", "Web Application / Mobile Web"),
    ("Module Scope", "Auth, Gallery, Editor, Checkout, CMS, AI, Credits"),
    ("Document Version", f"v{next_version}.0 (Excel formatted)"),
    ("Generated Date", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
    ("Author", "QA Automation System"),
]

start_row = 6
for r_idx, (k, v) in enumerate(cover_data, start=start_row):
    cover_ws.row_dimensions[r_idx].height = 25
    
    # Key cell
    c_key = cover_ws.cell(row=r_idx, column=2, value=k.upper())
    c_key.font = font_cover_key
    c_key.fill = cover_key_fill
    c_key.alignment = Alignment(horizontal="right", vertical="center")
    c_key.border = full_border
    
    # Val cell
    cover_ws.merge_cells(start_row=r_idx, start_column=3, end_row=r_idx, end_column=5)
    c_val = cover_ws.cell(row=r_idx, column=3, value=v)
    c_val.font = font_cover_val
    c_val.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    
    # Apply border to merged cells effectively
    for col in range(3, 6):
        cover_ws.cell(row=r_idx, column=col).border = full_border

cover_ws.column_dimensions['A'].width = 5
cover_ws.column_dimensions['B'].width = 25
cover_ws.column_dimensions['C'].width = 30
cover_ws.column_dimensions['D'].width = 20
cover_ws.column_dimensions['E'].width = 20


# ---------------------------------------------------------
# 2. GENERATE REFERENCE DOCUMENT
# ---------------------------------------------------------
ref_ws = wb.create_sheet(title="Reference Document", index=1)
ref_ws.sheet_view.showGridLines = False
ref_headers = ["No", "Document Name", "Link / Path", "Description"]
for col_idx, h in enumerate(ref_headers, 2):
    c = ref_ws.cell(row=2, column=col_idx, value=h)
    c.font = font_header
    c.fill = header_fill
    c.alignment = center_align
    c.border = full_border

ref_data = [
    (1, "SRS Confluence Specifications", "https://tryonic-ai.atlassian.net/wiki/spaces/TAS/pages/33947650/POD+T-Shirt+Platform+BA+Specifications", "Business rules and functional specs"),
    (2, "Figma UI/UX Design", "https://stitch.withgoogle.com/projects/6046886963733530705", "Wireframes and final UI designs"),
]
for r_idx, (no, doc, link, desc) in enumerate(ref_data, start=3):
    ref_ws.row_dimensions[r_idx].height = 30
    for col_idx, val in enumerate([no, doc, link, desc], 2):
        c = ref_ws.cell(row=r_idx, column=col_idx, value=val)
        c.border = full_border
        c.font = font_body
        c.alignment = wrap_align if col_idx > 2 else center_align
        if link == val:
            c.font = Font(name=FONT_FAMILY, size=STD_SIZE, color="0563C1", underline="single")
            c.hyperlink = val

ref_ws.column_dimensions['B'].width = 8
ref_ws.column_dimensions['C'].width = 35
ref_ws.column_dimensions['D'].width = 65
ref_ws.column_dimensions['E'].width = 50


# ---------------------------------------------------------
# 3. GENERATE CHANGE HISTORY
# ---------------------------------------------------------
hst_ws = wb.create_sheet(title="Change History", index=2)
hst_ws.sheet_view.showGridLines = False
hst_headers = ["Version", "Date", "Description", "Author"]
for col_idx, h in enumerate(hst_headers, 2):
    c = hst_ws.cell(row=2, column=col_idx, value=h)
    c.font = font_header
    c.fill = header_fill
    c.alignment = center_align
    c.border = full_border

hst_data = [
    ("v1.0", "2026-03-12", "Initial creation of test cases based on Confluence Spec", "QA Team"),
    ("v2.0", "2026-03-13", "Added comprehensive UI/UX validations and security cases against Stitch design", "QA Team"),
    ("v3.0", "2026-03-13", "Restructured matrix with multi-sheet execution tracking (Rounds 1 & 2)", "QA Team"),
    ("v4.0", "2026-03-13", "Standardized fonts (Calibri) and professional UI layout enhancements", "QA Team"),
    ("v5.0", "2026-03-13", "Added robust Test Automation tracking columns and specific Data Validations", "QA Team"),
]
for r_idx, (ver, date, desc, author) in enumerate(hst_data, start=3):
    hst_ws.row_dimensions[r_idx].height = 25
    for col_idx, val in enumerate([ver, date, desc, author], 2):
        c = hst_ws.cell(row=r_idx, column=col_idx, value=val)
        c.border = full_border
        c.font = font_body
        c.alignment = Alignment(horizontal="left" if col_idx == 4 else "center", vertical="center", wrap_text=True)

hst_ws.column_dimensions['B'].width = 15
hst_ws.column_dimensions['C'].width = 20
hst_ws.column_dimensions['D'].width = 80
hst_ws.column_dimensions['E'].width = 20


# ---------------------------------------------------------
# DATA PREPARATION (Features List)
# ---------------------------------------------------------
csv_headers = ["TC_ID", "US_Mapping", "Feature", "Module", "Title", "Type", "Priority", "Precondition", "Test_Data", "Steps", "Expected_Result"]
sub_headers = csv_headers + [
    "Action Type\nアクション",
    "Create TCs Type", # newly added
    "Execution Type\n実行タイプ", 
    "Result\n結果", "Test date\nテスト日", "Tester\nテスター", "ID Bug\nバグID",  
    "Result\n結果", "Test date\nテスト日", "Tester\nテスター", "ID Bug\nバグID",  
    "Evidence\n証拠", "Notes\n備考" 
]

grouped_features = df.groupby('Feature', sort=False)
summary_data = []

def safe_sheet_name(name):
    s = re.sub(r'[\\/*?:\[\]]', '', name)
    return s[:31]

# ---------------------------------------------------------
# 4. GENERATE FEATURE SHEETS
# ---------------------------------------------------------
for feature, feature_df in grouped_features:
    sheet_name = safe_sheet_name(feature)
    ws = wb.create_sheet(title=sheet_name)
    
    ws.add_data_validation(dv_result)
    ws.add_data_validation(dv_action)
    ws.add_data_validation(dv_auto)

    # Top grouping row
    ws.append([""] * 14 + ["Round 1 第1戦", "", "", "", "Round 2 第2戦", "", "", "", "", ""])
    ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=18)
    ws.merge_cells(start_row=1, start_column=19, end_row=1, end_column=22)
    
    for col in [15, 19]:
        c = ws.cell(row=1, column=col)
        c.fill = round_head_fill
        c.font = font_header
        c.alignment = center_align
        c.border = full_border
        for offset in range(1, 4):
            ws.cell(row=1, column=col+offset).border = full_border

    # Headers
    ws.append(sub_headers)
    ws.row_dimensions[2].height = 40
    for col_idx in range(1, len(sub_headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = full_border
        
    current_row = 3
    total_feature_tcs = 0
    
    for cat in categories_order:
        cat_df = feature_df[feature_df['CustomCategory'] == cat]
        if not cat_df.empty:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sub_headers))
            cell = ws.cell(row=current_row, column=1)
            cell.value = f"📌 {cat}"
            cell.font = Font(name=FONT_FAMILY, size=12, bold=True, color="1F4E78")
            cell.fill = category_fill
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for col_idx in range(1, len(sub_headers) + 1):
                ws.cell(row=current_row, column=col_idx).border = full_border
            current_row += 1
            
            for _, row in cat_df.iterrows():
                exec_type = "Manual" if "UI/UX" in cat else "Auto"
                # Row mapping exactly offsets to our headers
                row_data = [row[col] for col in csv_headers] + ["Add new", "By AI", exec_type, "Untested", "", "", "", "Untested", "", "", "", "", ""]
                ws.append(row_data)
                
                # Apply data validation ranges for this row
                dv_action.add(ws.cell(row=current_row, column=12)) # Action Type
                dv_auto.add(ws.cell(row=current_row, column=13))   # Automation Status
                dv_auto.add(ws.cell(row=current_row, column=14))   # Execution Type
                dv_result.add(ws.cell(row=current_row, column=15)) # Result 1
                dv_result.add(ws.cell(row=current_row, column=19)) # Result 2
                
                for col_idx in range(1, len(sub_headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = full_border
                    col_name = sub_headers[col_idx-1]
                    cell.font = font_body
                    
                    if col_name in ["Steps", "Expected_Result", "Precondition", "Title", "Test_Data", "Notes\n備考", "Evidence\n証拠"]:
                        cell.alignment = wrap_align
                    else:
                        cell.alignment = center_align
                        
                    if col_name == "Priority":
                        if row["Priority"] == "P0": cell.font = Font(name=FONT_FAMILY, size=STD_SIZE, color="C00000", bold=True)
                        elif row["Priority"] == "P1": cell.font = Font(name=FONT_FAMILY, size=STD_SIZE, color="ED7D31", bold=True)
                        elif row["Priority"] == "P2": cell.font = Font(name=FONT_FAMILY, size=STD_SIZE, color="70AD47", bold=True)
                            
                    if col_name == "Type":
                        if row["Type"] == "UI/UX": cell.font = Font(name=FONT_FAMILY, size=STD_SIZE, color="7030A0", bold=True)
                        elif row["Type"] == "Negative": cell.font = Font(name=FONT_FAMILY, size=STD_SIZE, color="C00000")
                        elif row["Type"] == "Positive": cell.font = Font(name=FONT_FAMILY, size=STD_SIZE, color="0070C0")

                current_row += 1
                total_feature_tcs += 1

    column_widths = {"TC_ID": 15, "US_Mapping": 11, "Feature": 12, "Module": 12, "Title": 30, "Type": 12, "Priority": 10, "Precondition": 20, "Test_Data": 15, "Steps": 45, "Expected_Result": 45, "Action Type\nアクション": 15, "Create TCs Type": 15, "Execution Type\n実行タイプ": 15, "Result\n結果": 12, "Test date\nテスト日": 15, "Tester\nテスター": 15, "ID Bug\nバグID": 15, "Evidence\n証拠": 20, "Notes\n備考": 20}
    for i, header in enumerate(sub_headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_widths.get(header, 15)

    summary_data.append({
        "Feature": feature,
        "SheetName": sheet_name,
        "Total": total_feature_tcs
    })


# ---------------------------------------------------------
# 5. GENERATE EXECUTION SUMMARY
# ---------------------------------------------------------
wb.remove(default_sheet)
summary_ws = wb.create_sheet(title="Execution Summary", index=3)
summary_ws.sheet_view.showGridLines = False

summary_headers = ["Sheet test\nテストシート", "Pass\n合格", "Fail\n不合格", "Untested\n未実行", "N/A\n対象外", "Total test case\nテスト数合計", "%Progress\n(Tested/Total)", "%Progress\n(Pass/Total)"]
summary_ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=len(summary_headers)+1)
summary_title = summary_ws.cell(row=2, column=2)
summary_title.value = "Test Execution Summary\nテスト実行の概要"
summary_title.font = Font(name=FONT_FAMILY, size=14, bold=True, color="FFFFFF")
summary_title.fill = summary_head_fill
summary_title.alignment = center_align
summary_title.border = full_border
for offset in range(1, len(summary_headers)):
    summary_ws.cell(row=2, column=2+offset).border = full_border

# Append exactly inside the box starting from Col 2 (B)
for col_idx, h in enumerate(summary_headers, 2):
    cell = summary_ws.cell(row=3, column=col_idx, value=h)
    cell.font = font_header
    cell.fill = summary_head_fill
    cell.alignment = center_align
    cell.border = full_border

# Round 1 Data
current_sum_row = 4
for i, data in enumerate(summary_data):
    sheet = data['SheetName']
    total = data['Total']
    
    f_pass = f'=COUNTIF(\'{sheet}\'!O:O, "Pass")'
    f_fail = f'=COUNTIF(\'{sheet}\'!O:O, "Fail")'
    f_na = f'=COUNTIF(\'{sheet}\'!O:O, "N/A")'
    f_untested = f'=G{current_sum_row} - C{current_sum_row} - D{current_sum_row} - F{current_sum_row}'
    f_prog_test = f'=IF(G{current_sum_row}>0, (C{current_sum_row}+D{current_sum_row})/G{current_sum_row}, 0)'
    f_prog_pass = f'=IF(G{current_sum_row}>0, C{current_sum_row}/G{current_sum_row}, 0)'
    
    row_data = [f"{i+1}. {data['Feature']}", f_pass, f_fail, f_untested, f_na, total, f_prog_test, f_prog_pass]
    for col_idx, val in enumerate(row_data, 2):
        cell = summary_ws.cell(row=current_sum_row, column=col_idx, value=val)
        cell.border = full_border
        cell.font = font_body
        cell.alignment = Alignment(horizontal="left", vertical="center") if col_idx == 2 else center_align
        if col_idx in [8, 9]: cell.number_format = '0%'
    
    current_sum_row += 1

r1_tot_row = current_sum_row
r1_tot_data = [
    "Round 1 第1戦", 
    f"=SUM(C4:C{r1_tot_row-1})", 
    f"=SUM(D4:D{r1_tot_row-1})", 
    f"=SUM(E4:E{r1_tot_row-1})", 
    f"=SUM(F4:F{r1_tot_row-1})", 
    f"=SUM(G4:G{r1_tot_row-1})",
    f"=IF(G{r1_tot_row}>0, (C{r1_tot_row}+D{r1_tot_row})/G{r1_tot_row}, 0)",
    f"=IF(G{r1_tot_row}>0, C{r1_tot_row}/G{r1_tot_row}, 0)"
]
for col_idx, val in enumerate(r1_tot_data, 2):
    cell = summary_ws.cell(row=r1_tot_row, column=col_idx, value=val)
    cell.border = full_border
    cell.font = font_bold
    cell.fill = total_fill
    cell.alignment = center_align
    if col_idx in [8, 9]: cell.number_format = '0%'

current_sum_row += 1

# Round 2 Data
r2_start = current_sum_row
for i, data in enumerate(summary_data):
    sheet = data['SheetName']
    total = data['Total']
    
    f_pass = f'=COUNTIF(\'{sheet}\'!S:S, "Pass")'
    f_fail = f'=COUNTIF(\'{sheet}\'!S:S, "Fail")'
    f_na = f'=COUNTIF(\'{sheet}\'!S:S, "N/A")'
    f_untested = f'=G{current_sum_row} - C{current_sum_row} - D{current_sum_row} - F{current_sum_row}'
    f_prog_test = f'=IF(G{current_sum_row}>0, (C{current_sum_row}+D{current_sum_row})/G{current_sum_row}, 0)'
    f_prog_pass = f'=IF(G{current_sum_row}>0, C{current_sum_row}/G{current_sum_row}, 0)'
    
    row_data = [f"{i+1}. {data['Feature']}", f_pass, f_fail, f_untested, f_na, total, f_prog_test, f_prog_pass]
    for col_idx, val in enumerate(row_data, 2):
        cell = summary_ws.cell(row=current_sum_row, column=col_idx, value=val)
        cell.border = full_border
        cell.font = font_body
        cell.alignment = Alignment(horizontal="left", vertical="center") if col_idx == 2 else center_align
        if col_idx in [8, 9]: cell.number_format = '0%'
    
    current_sum_row += 1

r2_tot_row = current_sum_row
r2_tot_data = [
    "Round 2 第2戦", 
    f"=SUM(C{r2_start}:C{r2_tot_row-1})", 
    f"=SUM(D{r2_start}:D{r2_tot_row-1})", 
    f"=SUM(E{r2_start}:E{r2_tot_row-1})", 
    f"=SUM(F{r2_start}:F{r2_tot_row-1})", 
    f"=SUM(G{r2_start}:G{r2_tot_row-1})",
    f"=IF(G{r2_tot_row}>0, (C{r2_tot_row}+D{r2_tot_row})/G{r2_tot_row}, 0)",
    f"=IF(G{r2_tot_row}>0, C{r2_tot_row}/G{r2_tot_row}, 0)"
]
for col_idx, val in enumerate(r2_tot_data, 2):
    cell = summary_ws.cell(row=r2_tot_row, column=col_idx, value=val)
    cell.border = full_border
    cell.font = font_bold
    cell.fill = total_fill
    cell.alignment = center_align
    if col_idx in [8, 9]: cell.number_format = '0%'

summary_ws.column_dimensions['A'].width = 5
for i, header in enumerate(summary_headers, 2):
    summary_ws.column_dimensions[get_column_letter(i)].width = 25 if i==2 else 15
summary_ws.row_dimensions[2].height = 40
summary_ws.row_dimensions[3].height = 30

wb.save(output_path)
print(f"🎉 Professional Multisheet Excel generated: {output_path}")
