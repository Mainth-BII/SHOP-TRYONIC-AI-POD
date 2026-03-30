import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

csv_path = "TC_POD-TShirt-Platform_v4_2026-03-13.csv"
output_path = "TC_POD-TShirt-Platform_Styled_2026-03-13.xlsx"

if not os.path.exists(csv_path):
    print(f"File {csv_path} not found!")
    exit(1)

df = pd.read_csv(csv_path)

# Ensure NaN is replaced with empty string
df = df.fillna("")

# Categorization logic
def get_custom_category(row):
    module = str(row['Module']).lower()
    title = str(row['Title']).lower()
    tc_type = str(row['Type']).lower()

    if 'security' in module or 'lock' in title or 'xss' in title or 'injection' in title:
        return 'Security'
    elif 'network' in module or 'timeout' in title or 'rate limit' in title or 'performance' in module:
        return 'Performance'
    elif 'ui/ux' in tc_type or 'ui' in module:
        return 'UI/UX'
    elif tc_type in ['negative', 'boundary'] or 'validation' in title:
        return 'Validation'
    else:
        return 'Functional (Logic & Behavior)'

df['CustomCategory'] = df.apply(get_custom_category, axis=1)

# Sort features properly to keep epic sequence logically
# But practically we can just group by Feature as they appear
# We will iterate feature by feature, then by category
categories_order = ['UI/UX', 'Validation', 'Functional (Logic & Behavior)', 'Security', 'Performance']

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Cases"

headers = ["TC_ID", "US_Mapping", "Feature", "Module", "Title", "Type", "Priority", "Precondition", "Test_Data", "Steps", "Expected_Result"]

# Styles
header_font = Font(bold=True, shadow=False, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")  # Dark Blue
feature_fill = PatternFill(start_color="4B0082", end_color="4B0082", fill_type="solid")  # Indigo
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")  # Light blueish gray
border = Side(border_style="thin", color="000000")
full_border = Border(left=border, right=border, top=border, bottom=border)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Write headers
ws.append(headers)
for col_idx in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col_idx)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = full_border

current_row = 2

# Group and write
grouped_features = df.groupby('Feature', sort=False)
for feature, feature_df in grouped_features:
    # Feature Title Row
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
    cell = ws.cell(row=current_row, column=1)
    cell.value = f"🚀 FEATURE: {feature.upper()}"
    cell.font = Font(bold=True, color="FFFFFF", size=14)
    cell.fill = feature_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=current_row, column=col_idx).border = full_border
    current_row += 1

    # Categories
    for cat in categories_order:
        cat_df = feature_df[feature_df['CustomCategory'] == cat]
        if not cat_df.empty:
            # Category Title Row
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(headers))
            cell = ws.cell(row=current_row, column=1)
            cell.value = f"📌 {cat}"
            cell.font = Font(bold=True, color="000000", size=12)
            cell.fill = category_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=current_row, column=col_idx).border = full_border
            current_row += 1

            # Test Cases
            for _, row in cat_df.iterrows():
                row_data = [row[col] for col in headers]
                ws.append(row_data)
                
                # Apply styles to normal row
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = full_border
                    
                    if headers[col_idx-1] in ["Steps", "Expected_Result", "Precondition", "Title", "Test_Data"]:
                        cell.alignment = wrap_align
                    else:
                        cell.alignment = center_align
                        
                    # Color formatting for Priority
                    if headers[col_idx-1] == "Priority":
                        if row["Priority"] == "P0":
                            cell.font = Font(color="FF0000", bold=True)
                        elif row["Priority"] == "P1":
                            cell.font = Font(color="FFA500", bold=True)
                        elif row["Priority"] == "P2":
                            cell.font = Font(color="008000", bold=True)
                            
                    # Color formatting for Type
                    if headers[col_idx-1] == "Type":
                        if row["Type"] == "UI/UX":
                            cell.font = Font(color="800080", bold=True)
                        elif row["Type"] == "Negative":
                            cell.font = Font(color="FF0000")
                        elif row["Type"] == "Positive":
                            cell.font = Font(color="0000FF")

                current_row += 1

# Adjust column widths
column_widths = {
    "TC_ID": 16,
    "US_Mapping": 12,
    "Feature": 15,
    "Module": 15,
    "Title": 35,
    "Type": 12,
    "Priority": 10,
    "Precondition": 20,
    "Test_Data": 15,
    "Steps": 45,
    "Expected_Result": 45
}

for i, header in enumerate(headers, 1):
    width = column_widths.get(header, 15)
    ws.column_dimensions[get_column_letter(i)].width = width

wb.save(output_path)
print(f"🎉 Excel file generated successfully: {output_path}")
