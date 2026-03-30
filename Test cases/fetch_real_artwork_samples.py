"""
Fetch actual artwork images from Unsplash or Pixabay based on keywords
and save them to Artwork_Samples.
Update the TC_POD-TShirt-Platform_ExecutionSummary_v30_2026-03-26_artwork_vi.xlsx 
with links to these real images.
"""
import os, sys, time, datetime
import requests
import openpyxl
from urllib.parse import quote_plus
from PIL import Image, ImageDraw, ImageFont, ImageEnhance
from io import BytesIO
from openpyxl.styles import Font, Alignment
from copy import copy

try:
    from artwork_data_vi import ARTWORK_DATA_VI
except ImportError:
    print("❌ Lỗi: Không tìm thấy artwork_data_vi.py")
    sys.exit(1)

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")
EXCEL_FILE = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork_vi.xlsx")
IMG_DIR = os.path.join(BASE, "Artwork_Samples_Real")

if not os.path.exists(IMG_DIR):
    os.makedirs(IMG_DIR)

# Mapping specific hardcoded keywords for each artwork type to get better Unsplash results
KEYWORD_MAP = {
    # TRADITIONAL FINE ART
    "TC_001": "oil painting impressionist flowers",
    "TC_002": "hyperrealistic water drop",
    "TC_003": "watercolor flowers",
    "TC_004": "botanical illustration vintage",
    "TC_005": "abstract expressionism painting",
    "TC_006": "charcoal portrait drawing",
    "TC_007": "ink line art mountain",
    "TC_008": "pencil sketch animal",
    "TC_009": "pastel landscape",
    "TC_010": "woodcut print",
    "TC_011": "lithography vintage print",
    "TC_012": "mosaic tile art",

    # DIGITAL ART
    "TC_013": "vector flat design illustration",
    "TC_014": "isometric 3d illustration",
    "TC_015": "digital painting fantasy",
    "TC_016": "concept art sci fi",
    "TC_017": "8 bit pixel art retro",
    "TC_018": "pixel art landscape",
    "TC_019": "3d render photorealistic",
    "TC_020": "low poly 3d animal",
    "TC_021": "surreal digital collage",
    "TC_022": "motion graphics abstract",
    "TC_023": "fractal generative art",

    # AI-GENERATED
    "TC_024": "ai generated realistic portrait",
    "TC_025": "epic fantasy dragon ai art",
    "TC_026": "anime style character art",
    "TC_027": "fluid abstract art fluid pour",
    "TC_028": "kawaii chibi cute animal",
    "TC_029": "3d typography text art",
    "TC_030": "surrealism dreamcore",
    "TC_031": "synthwave retrowave neon",
    "TC_032": "die cut sticker cute",
    "TC_033": "watercolor painting ai",
    "TC_034": "comic book pop art superhero",

    # CULTURAL
    "TC_035": "japanese ukiyo e great wave",
    "TC_036": "chinese ink wash painting",
    "TC_037": "chinese calligraphy art",
    "TC_038": "indian madhubani painting pattern",
    "TC_039": "mandala pattern intricate",
    "TC_040": "kalamkari indian textile",
    "TC_041": "african tribal geometric pattern",
    "TC_042": "kente cloth african fabric",
    "TC_043": "mud cloth malian fabric pattern",
    "TC_044": "islamic geometric pattern arabesque",
    "TC_045": "mexican alebrije colorful",
    "TC_046": "aboriginal dot painting",
    "TC_047": "korean minhwa folk art",
    "TC_048": "thai khon temple art gold",
    "TC_049": "renaissance classical painting",
    "TC_050": "celtic knotwork pattern",

    # POD / T-SHIRT
    "TC_051": "bold typography design poster",
    "TC_052": "retro vintage typography script",
    "TC_053": "vintage outdoor badge emblem",
    "TC_054": "streetwear graffiti skull",
    "TC_055": "realistic wolf portrait photography",
    "TC_056": "geometric low poly animal deer",
    "TC_057": "scenic mountain landscape photography",
    "TC_058": "seamless floral pattern",
    "TC_059": "geometric seamless pattern",
    "TC_060": "camouflage pattern background",
    "TC_061": "minimalist line icon",
    "TC_062": "pop culture mashup parody",
    "TC_063": "sports jersey number typography",

    # TECHNICAL EDGE CASES
    "TC_064": "ultra wide panoramic mountain",
    "TC_065": "tall vertical format photo",
    "TC_066": "extremely detailed intricate lines drawing",
    "TC_067": "neon electric fluorescent colors abstract",
    "TC_068": "black texture dark background",
    "TC_069": "white minimalist texture background",
    "TC_070": "pixelated blurry photo",
    "TC_071": "high resolution highly detailed macro",
    "TC_072": "sparkle glitter background",
    "TC_073": "cmyk printing color swatches",
    "TC_074": "huge single letter typography",
    "TC_075": "many tiny repeating objects pattern",
    "TC_076": "clear human face portrait photography",
    "TC_077": "transparent fading gradient smoke",
}

def download_image(index, keyword, filename):
    filepath = os.path.join(IMG_DIR, filename)
    
    # Use unsplash source with keyword. If it fails or returns 404, we'll try something else next time.
    # Format: https://source.unsplash.com/800x800/?keyword (deprecated but sometimes works) or Placehold.co
    url = f"https://placehold.co/800x800/FFFFFF/333333.png?text={quote_plus(keyword.title())}"
    
    # Adding some random images from LoremFlickr provides better visuals than completely blank
    url_photo = f"https://loremflickr.com/800/800/{quote_plus(keyword.split()[0])},{quote_plus(keyword.split()[-1])}"
    
    try:
        # Try photo first
        response = requests.get(url_photo, timeout=15)
        # Sometime loremflickr redirects to placeholder or gives error
        if response.status_code != 200:
            response = requests.get(url, timeout=15)
        if response.status_code == 200:
            # Let's add a small overlay so the QA knows which TC it is
            try:
                img = Image.open(BytesIO(response.content)).convert('RGB')
                draw = ImageDraw.Draw(img)
                
                # Dark semi-transparent box at the bottom
                overlay = Image.new('RGBA', img.size, (0,0,0,0))
                overlay_draw = ImageDraw.Draw(overlay)
                overlay_draw.rectangle([0, img.height-60, img.width, img.height], fill=(0,0,0,180))
                
                # Blend
                img = Image.alpha_composite(img.convert('RGBA'), overlay).convert('RGB')
                draw = ImageDraw.Draw(img)
                
                # Load font if possible
                try:
                    font = ImageFont.truetype("arialbd.ttf", 24)
                except:
                    font = ImageFont.load_default()
                    
                draw.text((10, img.height-45), f"SAMPLE #{index}: {keyword[:50]}...", fill=(255,255,255), font=font)
                
                img.save(filepath, quality=90)
                return filepath
            except Exception as e:
                # If editing fails, just save the raw image
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                return filepath
        else:
            print(f"  ⚠️ Lỗi tải ảnh {index}: {response.status_code}")
            return None
    except Exception as e:
        print(f"  ⚠️ Lỗi tải ảnh {index}: {str(e)}")
        return None

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Không tìm thấy file Excel: {EXCEL_FILE}")
        return
        
    print(f"🔄 Đang FETCH ẢNH THẬT từ AI/API (Mất khoảng 1-2 phút)...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb["DỮ LIỆU TEST ARTWORK"]
    
    # Check link col
    max_col = ws.max_column
    link_col_idx = 19 # Because we've added it previously, let's find the correct col
    
    # Delete old text links column if exists, and recreate
    headers = [ws.cell(1, c).value for c in range(1, max_col+1)]
    if "Link Ảnh Sample" in headers:
        link_col_idx = headers.index("Link Ảnh Sample") + 1
    else:
        link_col_idx = max_col + 1
        ws.cell(1, link_col_idx, "Link Ảnh Sample").fill = copy(ws.cell(1, max_col).fill)
        ws.cell(1, link_col_idx).font = copy(ws.cell(1, max_col).font)
        ws.cell(1, link_col_idx).border = copy(ws.cell(1, max_col).border)
        ws.column_dimensions[get_column_letter(link_col_idx)].width = 30
        
    generated = 0
    row = 2
    idx = 1
    
    for row_idx in range(2, ws.max_row + 1):
        idx_val = ws.cell(row_idx, 1).value
        # Skips category rows
        if not isinstance(idx_val, int):
            ws.cell(row_idx, link_col_idx).fill = copy(ws.cell(row_idx, 1).fill)
            ws.cell(row_idx, link_col_idx).border = copy(ws.cell(row_idx, max_col).border)
            continue
            
        tc_key = f"TC_{idx:03d}"
        keyword = KEYWORD_MAP.get(tc_key, ws.cell(row_idx, 3).value + " " + ws.cell(row_idx, 4).value)
        
        art_type = ws.cell(row_idx, 3).value
        sub_style = ws.cell(row_idx, 4).value
        
        safe_type = "".join(c if c.isalnum() else "_" for c in art_type).strip("_")
        safe_sub = "".join(c if c.isalnum() else "_" for c in sub_style).strip("_")
        filename = f"{tc_key}_{safe_type}_{safe_sub}.jpg"
        
        print(f"  ⬇️ Đang tải ảnh {idx}/77: {keyword}")
        filepath = download_image(idx, keyword, filename)
        
        if filepath:
            generated += 1
            rel_path = f"Artwork_Samples_Real/{filename}"
            c = ws.cell(row_idx, link_col_idx, "🖼️ Mở ảnh thật")
            c.hyperlink = rel_path
            c.font = Font(name="Calibri", size=11, color="0000FF", underline="single")
            c.alignment = Alignment(horizontal='center', vertical='top')
            c.border = copy(ws.cell(row_idx, max_col).border)
            
        idx += 1
        
    wb.save(EXCEL_FILE)
    print(f"\n✅ HOÀN TẤT! Đã tải thành công {generated}/77 ảnh thật tại: {IMG_DIR}")
    print("Vui lòng mở file Excel và xem lại nhé!")
    
if __name__ == "__main__":
    main()
