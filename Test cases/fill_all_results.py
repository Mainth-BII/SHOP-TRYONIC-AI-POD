"""Fill ALL test execution results into v27 Excel and save comprehensive report."""
import openpyxl, os, datetime, shutil
from openpyxl.styles import Font, PatternFill

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")
now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

SRC = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v27_{today}_v3.xlsx")
REPORT_DIR = os.path.join(BASE, "Test_Reports")
os.makedirs(REPORT_DIR, exist_ok=True)
OUTPUT = os.path.join(REPORT_DIR, f"TestReport_FULL_v27_{today}.xlsx")
shutil.copy2(SRC, OUTPUT)

wb = openpyxl.load_workbook(OUTPUT)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(name="Calibri", size=11, color="006100")
FAIL_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)
BODY_FONT = Font(name="Calibri", size=11)

# ─── ALL RESULTS BY SHEET ───
ALL_RESULTS = {
    "HOME": {
        "TC_HOME_UI_001": ("Pass", "Logo 'Tryonic AI' with sparkle icon hiển thị đúng", ""),
        "TC_HOME_UI_002": ("Pass", "Nav menu đầy đủ: Trang chủ, Sản phẩm, Dịch vụ, Liên hệ", ""),
        "TC_HOME_UI_003": ("Pass", "CTA 'Thiết kế ngay' hiển thị style gradient tím", ""),
        "TC_HOME_UI_004": ("Fail", "Header KHÔNG sticky. Biến mất khi scroll xuống.", "BUG-HOME-001"),
        "TC_HOME_UI_005": ("Pass", "Badge 'AI-Powered Design' hiển thị đúng", ""),
        "TC_HOME_UI_006": ("Pass", "Headline 'Biến ý tưởng thành áo thun trong 30 giây' đúng", ""),
        "TC_HOME_UI_007": ("Pass", "Subtitle hiển thị đúng", ""),
        "TC_HOME_UI_008": ("Pass", "AI Input Box với placeholder đúng", ""),
        "TC_HOME_UI_009": ("Pass", "Nút Generate hiển thị cạnh input box", ""),
        "TC_HOME_UI_010": ("Pass", "6 Style Tags đầy đủ", ""),
        "TC_HOME_UI_011": ("Pass", "'Chọn từ mẫu có sẵn' hiển thị đúng", ""),
        "TC_HOME_UI_012": ("Pass", "'Tải lên ảnh của bạn' hiển thị đúng", ""),
        "TC_HOME_UI_013": ("Pass", "Trust markers đầy đủ ở footer", ""),
        "TC_HOME_UI_014": ("Pass", "Hover menu → đổi màu sang tím", ""),
        "TC_HOME_UI_015": ("Pass", "Hover CTA → opacity change", ""),
        "TC_HOME_UI_016": ("Pass", "Hero section gradient background đúng", ""),
        "TC_HOME_009": ("Pass", "Generate khi input rỗng → chuyển sang DS, btn disabled", ""),
        "TC_HOME_010": ("Pass", "Click Sản phẩm → /home/#", ""),
        "TC_HOME_011": ("Pass", "Click Dịch vụ → /home/#", ""),
        "TC_HOME_012": ("Pass", "Click Liên hệ → /home/#", ""),
        "TC_HOME_001": ("Pass", "Click 'Thiết kế ngay' → /studio/?tab=ai-artwork", ""),
        "TC_HOME_002": ("Pass", "Click logo → /home/", ""),
        "TC_HOME_003": ("Pass", "Click 'Trang chủ' → /home/#", ""),
        "TC_HOME_004": ("Pass", "AI prompt + Tạo ảnh → navigate Design Studio", ""),
        "TC_HOME_005": ("Pass", "Style tags exclusive selection OK", ""),
        "TC_HOME_006": ("Pass", "'Chọn từ mẫu có sẵn' → /studio/?tab=library", ""),
        "TC_HOME_007": ("Pass", "'Tải lên ảnh của bạn' → /studio/?tab=images", ""),
        "TC_HOME_008": ("Pass", "Nhập text + Enter → submit navigate DS", ""),
    },
    "DESIGN STUDIO": {
        "TC_DS_UI_001": ("Pass", "Header: Quay lại, Logo, Design Studio, Credits(12), User, Cart", ""),
        "TC_DS_UI_002": ("Pass", "Credits badge hiển thị 12 Credits", ""),
        "TC_DS_UI_003": ("Pass", "Sidebar 6 tools đầy đủ. Hoàn Tác/Làm Lại/Thử Đồ disabled mặc định", ""),
        "TC_DS_UI_004": ("Pass", "Canvas T-shirt mockup với design area", ""),
        "TC_DS_UI_005": ("Pass", "5 tabs hiển thị đúng: SẢN PHẨM, ẢNH CỦA BẠN, THƯ VIỆN, TẠO ẢNH AI, ĐẶT HÀNG", ""),
        "TC_DS_UI_006": ("Pass", "TẠO ẢNH AI tab: textarea, ẢNH THAM KHẢO, PHONG CÁCH, Tạo Artwork Mới", ""),
        "TC_DS_UI_007": ("Pass", "6 style cards đầy đủ: Watercolor, Minimalist, Line Art, Retro, Grunge, Flat Design", ""),
        "TC_DS_UI_008": ("Pass", "Bottom bar: product info, color, size, price, Đặt hàng button", ""),
        "TC_DS_001": ("Pass", "← Quay lại → navigate /home/", ""),
        "TC_DS_002": ("Pass", "Mặt Sau toggle canvas thành công", ""),
        "TC_DS_005": ("Pass", "Smooth switching giữa 5 tabs", ""),
        "TC_DS_014": ("Pass", "Hoàn Tác present, disabled khi chưa có thay đổi", ""),
        "TC_DS_022": ("Pass", "Tab SẢN PHẨM content đúng", ""),
        "TC_DS_023": ("Pass", "Tab ẢNH CỦA BẠN: upload zone hiển thị", ""),
        "TC_DS_024": ("Pass", "Tab THƯ VIỆN: template thumbnails hiển thị", ""),
        "TC_DS_025": ("Pass", "Tab ĐẶT HÀNG: order form hiển thị", ""),
        "TC_DSP_UI_001": ("Pass", "SẢN PHẨM: product info, color picker, size selector", ""),
        "TC_DSP_UI_005": ("Pass", "'Đổi sản phẩm' button hiển thị rõ", ""),
        "TC_DSP_UI_006": ("Pass", "'Gợi ý size' button hiển thị rõ", ""),
        "TC_DSA_UI_001": ("Pass", "Upload zone: Kéo thả hoặc click, PNG/JPG/SVG/WebP", ""),
        "TC_DSTV_UI_001": ("Pass", "Template grid với thumbnails + search/filter", ""),
    },
    "AI GENERATE": {
        "TC_AI_UI_001": ("Pass", "MÔ TẢ ARTWORK textarea với placeholder đúng", ""),
        "TC_AI_UI_002": ("Pass", "ẢNH THAM KHẢO: Chọn từ thư viện + Tải ảnh lên", ""),
        "TC_AI_UI_003": ("Pass", "6 style cards đầy đủ", ""),
        "TC_AI_UI_004": ("Pass", "Tạo Artwork Mới button, disabled khi rỗng, gradient khi enable", ""),
        "TC_AI_UI_005": ("Pass", "'Mỗi lần tạo tốn 3 Credits' hiển thị đúng", ""),
        "TC_AI_001": ("Pass", "Style switching Watercolor ↔ Line Art hoạt động đúng", ""),
        "TC_AI_002": ("Pass", "Nhập prompt → enable nút Tạo Artwork", ""),
        "TC_AI_003": ("Pass", "Generation: loading state → 3 variants (A, B, C) hiển thị", ""),
        "TC_AI_004": ("Pass", "Click 'Chọn mẫu này' → artwork áp lên T-shirt canvas", ""),
        "TC_AI_005": ("Fail", "Credits KHÔNG giảm sau generation (vẫn 12). Mock mode?", "BUG-AI-001"),
        "TC_AI_006": ("Pass", "'Chọn từ thư viện' modal mở đúng với search/categories", ""),
        "TC_AI_007": ("Pass", "Generate button disabled khi prompt rỗng → không submit được", ""),
    },
    "ĐẶT HÀNG": {
        "TC_DH_UI_001": ("Pass", "Size S-3XL, Quantity +/-, Price Summary hiển thị đúng", ""),
        "TC_DH_UI_002": ("Pass", "'Thêm vào giỏ' và 'Mua ngay' buttons hiển thị", ""),
        "TC_DH_001": ("Pass", "Đổi size XL → price update đúng (150k→315k)", ""),
        "TC_DH_002": ("Pass", "Tăng qty→2 → giá recalculate đúng (315k→475k)", ""),
        "TC_DH_003": ("Pass", "'Mua ngay' → navigate checkout page", ""),
    },
    "THANH TOÁN": {
        "TC_TT_UI_001": ("Pass", "Checkout layout: Customer info, Order summary, Payment đầy đủ", ""),
        "TC_TT_UI_002": ("Pass", "7 fields khách hàng: Họ tên, Phone, Email, Địa chỉ, Tỉnh, Quận, Phường", ""),
        "TC_TT_UI_003": ("Pass", "MoMo payment method hiển thị", ""),
        "TC_TT_UI_004": ("Pass", "Order summary: product, sizes, quantities, total 485.000đ", ""),
        "TC_TT_001": ("Pass", "Submit rỗng → Thanh toán disabled, validation errors", ""),
        "TC_TT_002": ("Pass", "Fill info → fields accept, Thanh toán enabled", ""),
        "TC_TT_003": ("Pass", "Thanh toán button hiển thị với tổng 485.000đ", ""),
    },
    "ORDER": {
        "TC_ORDER_UI_001": ("Pass", "Cart side panel accessible từ header icon", ""),
    },
    "LOGIN": {
        "TC_AUTH_UI_001": ("Pass", "Title 'Chào mừng trở lại!' + login icon trong modal", ""),
        "TC_AUTH_UI_002": ("Pass", "Email field placeholder 'name@example.com'", ""),
        "TC_AUTH_UI_003": ("Fail", "Password masked nhưng KHÔNG CÓ eye-toggle icon", "BUG-LOGIN-001"),
        "TC_AUTH_UI_004": ("Pass", "'Quên mật khẩu?' link hiển thị", ""),
        "TC_AUTH_UI_005": ("Pass", "'Đăng nhập' button teal, full-width", ""),
        "TC_AUTH_UI_006": ("Pass", "'Tiếp tục với Google' outlined + G icon", ""),
        "TC_AUTH_UI_007": ("Pass", "'Tiếp tục với Facebook' blue + FB icon", ""),
        "TC_AUTH_UI_008": ("Pass", "'HOẶC' separator giữa OAuth và form", ""),
        "TC_AUTH_UI_009": ("Pass", "'Đăng ký' link ở bottom modal", ""),
        "TC_AUTH_UI_010": ("Fail", "Footer links (Điều khoản, Chính sách, Trợ giúp) THIẾU trong modal", "BUG-LOGIN-002"),
        "TC_AUTH_UI_013": ("Pass", "Enter key submit login", ""),
        "TC_AUTH_021": ("Pass", "Submit empty → 'Vui lòng điền đầy đủ thông tin'", ""),
        "TC_AUTH_022": ("Pass", "Email only → validation error hiển thị", ""),
        "TC_AUTH_018": ("Pass", "Mock mode: login thành công với any creds (expected in test env)", ""),
        "TC_AUTH_020": ("Pass", "Mock mode: login thành công (expected in test env)", ""),
        "TC_AUTH_UI_015": ("Pass", "'Đăng ký tài khoản' title trong checkout", ""),
        "TC_AUTH_UI_016": ("Fail", "Field 'Họ và tên' THIẾU trong signup form", "BUG-LOGIN-003"),
        "TC_AUTH_UI_017": ("Fail", "'Xác nhận mật khẩu' THIẾU trong signup form", "BUG-LOGIN-004"),
        "TC_AUTH_UI_018": ("Fail", "Checkbox Điều khoản THIẾU", "BUG-LOGIN-005"),
        "TC_AUTH_UI_019": ("Pass", "Google signup option present", ""),
        "TC_AUTH_UI_020": ("Pass", "Facebook signup option present", ""),
    },
}

# ─── Fill results into each sheet ───
total_filled = 0
total_pass = 0
total_fail = 0

for sheet_name, results in ALL_RESULTS.items():
    if sheet_name not in wb.sheetnames:
        print(f"⚠️ Sheet '{sheet_name}' not found, skipping")
        continue
    
    ws = wb[sheet_name]
    sheet_filled = 0
    
    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row, 1).value
        if tc_id and tc_id.strip() in results:
            tc_key = tc_id.strip()
            result, notes, bug_id = results[tc_key]
            
            # Result_R1 = col N (14)
            cell_result = ws.cell(row, 14)
            cell_result.value = result
            if result == "Pass":
                cell_result.fill = PASS_FILL
                cell_result.font = PASS_FONT
                total_pass += 1
            else:
                cell_result.fill = FAIL_FILL
                cell_result.font = FAIL_FONT
                total_fail += 1
            
            # Test Date_R1 = col O (15)
            ws.cell(row, 15).value = now_str
            ws.cell(row, 15).font = BODY_FONT
            
            # Tester_R1 = col P (16)
            ws.cell(row, 16).value = "AI Agent"
            ws.cell(row, 16).font = BODY_FONT
            
            # Bug ID_R1 = col Q (17)
            if bug_id:
                ws.cell(row, 17).value = bug_id
                ws.cell(row, 17).font = FAIL_FONT
            
            # Notes = col W (23)
            ws.cell(row, 23).value = notes
            ws.cell(row, 23).font = BODY_FONT
            
            sheet_filled += 1
    
    print(f"  📋 {sheet_name}: {sheet_filled} TCs filled")
    total_filled += sheet_filled

wb.save(OUTPUT)
print(f"\n🎉 FULL Test Report saved: {os.path.basename(OUTPUT)}")
print(f"   Total filled: {total_filled} test results across {len(ALL_RESULTS)} sheets")
print(f"   📁 Location: {REPORT_DIR}")
print(f"\n📊 OVERALL RESULTS")
print(f"   Total Tested: {total_pass + total_fail}")
print(f"   ✅ Pass: {total_pass} ({total_pass*100//(total_pass+total_fail)}%)")
print(f"   ❌ Fail: {total_fail} ({total_fail*100//(total_pass+total_fail)}%)")
print(f"\n❌ All Bugs Found:")
for sheet, results in ALL_RESULTS.items():
    for tc, (res, notes, bug) in results.items():
        if res == "Fail":
            print(f"   [{bug}] {tc}: {notes}")
