"""
Fill ALL test results into v27_v3 Excel with Actual_Result + Evidence columns.
New column structure (25 cols):
  1=TC_ID, 2=Screen, 3=US_Mapping, 4=Module, 5=Title, 6=Type, 7=Priority,
  8=Pre-condition, 9=Steps, 10=Expected Result,
  11=Actual Result ← NEW
  12=Action Type, 13=Create TCs Type, 14=Execution Type,
  15=Result_R1, 16=Test Date_R1, 17=Tester_R1, 18=Bug ID_R1,
  19=Result_R2, 20=Test Date_R2, 21=Tester_R2, 22=Bug ID_R2,
  23=Evidence, 24=Notes, 25=Review_Manual
"""
import openpyxl, os, datetime, shutil
from openpyxl.styles import Font, PatternFill, Alignment

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")
now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

REPORT_DIR = os.path.join(BASE, "Test_Reports")
os.makedirs(REPORT_DIR, exist_ok=True)

SRC = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v27_{today}_v3.xlsx")
OUTPUT = os.path.join(REPORT_DIR, f"TestReport_FULL_v27_{today}_v4.xlsx")
shutil.copy2(SRC, OUTPUT)

wb = openpyxl.load_workbook(OUTPUT)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(name="Calibri", size=11, color="006100")
FAIL_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)
BODY_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(vertical='top', wrap_text=True)

# Evidence base path
EVIDENCE_DIR = r"C:\Users\maiho\.gemini\antigravity\brain\137607bd-c663-49d8-aa44-945410ab72b3"

# ═══ ALL TEST RESULTS: (result, actual_result, evidence_file, notes, bug_id) ═══
ALL_RESULTS = {
    "HOME": {
        # ── UI/UX ──
        "TC_HOME_UI_001": ("Pass", "Logo 'Tryonic' với sparkle icon hiển thị góc trái header, click được", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_002": ("Pass", "Nav menu: Trang chủ, Sản phẩm, Dịch vụ, Liên hệ hiển thị đầy đủ", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_003": ("Pass", "Nút 'Thiết kế ngay' gradient tím→đỏ, text trắng, rounded-full", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_004": ("Fail", "Header KHÔNG sticky. Khi scroll xuống header biến mất hoàn toàn", "home_confirm_header_not_sticky_1774325609417.png", "", "BUG-HOME-001"),
        "TC_HOME_UI_005": ("Pass", "Badge '✨ AI-Powered Design' hiển thị trên Hero section", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_006": ("Pass", "Headline 'Biến ý tưởng thành áo thun trong 30 giây' hiển thị đúng", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_007": ("Pass", "Subtitle 'Chỉ cần mô tả — AI sẽ thiết kế cho bạn...' hiển thị đúng", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_008": ("Pass", "AI Input Box với placeholder 'Mô tả áo thun bạn muốn...' hiển thị", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_009": ("Pass", "Nút Generate hiển thị cạnh input", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_010": ("Pass", "6 Style Tags đầy đủ: Minimalist, Streetwear, Anime, Vintage, Y2K, Abstract Art", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_011": ("Pass", "'Chọn từ mẫu có sẵn' hiển thị đúng với grid icon", "home_page_scrolled_1774322028177.png", "", ""),
        "TC_HOME_UI_012": ("Pass", "'Tải lên ảnh của bạn' hiển thị đúng với upload icon", "home_page_scrolled_1774322028177.png", "", ""),
        "TC_HOME_UI_013": ("Pass", "Trust markers ở footer: Thanh toán an toàn, Giao hàng toàn quốc, Đổi trả 7 ngày", "home_scrolled_header_footer_1774325588963.png", "", ""),
        "TC_HOME_UI_014": ("Pass", "Hover menu → đổi màu, cursor pointer", "home_ui_tests_1774325532508.webp", "", ""),
        "TC_HOME_UI_015": ("Pass", "Hover CTA → opacity/shadow change", "home_ui_tests_1774325532508.webp", "", ""),
        "TC_HOME_UI_016": ("Pass", "Hero section gradient background lavender hiển thị đúng", "home_hero_full_viewport_1774325561450.png", "", ""),
        # Responsive
        "TC_HOME_UI_860": ("Pass", "Zoom 200% và 50% layout stable, text sắc nét", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_861": ("Pass", "Emulated iPhone 375x812: layout 1 cột, wrap tags OK. Nav links ẩn nhưng chưa có hamburger", "responsive_emulated_tests_1774338152517.webp", "Nav links hidden on mobile, no hamburger menu", ""),
        "TC_HOME_UI_862": ("Pass", "Emulated Android 360x740: responsive OK, UI co giãn tốt", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_863": ("Pass", "Emulated iPad 768x1024: centered layout, nút song song, nav hiển thị", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_864": ("Pass", "Emulated Tablet 800x1280: hiển thị tốt, icon text scale OK", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_865": ("Pass", "Landscape 812x375 layout adapt đúng", "responsive_emulated_tests_1774338152517.webp", "", ""),
        # ── Validation ──
        "TC_HOME_009": ("Pass", "Generate khi input rỗng → chuyển sang DS, btn disabled là đúng", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_010": ("Pass", "Click Sản phẩm → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_011": ("Pass", "Click Dịch vụ → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_012": ("Pass", "Click Liên hệ → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_013": ("Pass", "XSS payload sanitized, rendered as plain text, no script execution", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_018": ("Pass", "1-char 'A' accepted, generated 3 artworks successfully", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_020": ("Fail", "Spaces-only prompt navigate tới DS KHÔNG validate. Thiếu trim whitespace", "home_remaining_tests_1774336927869.webp", "", "BUG-HOME-002"),
        # ── Functional ──
        "TC_HOME_001": ("Pass", "Click 'Thiết kế ngay' → /studio/?tab=ai-artwork", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_002": ("Pass", "Click logo → /home/ redirect đúng", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_003": ("Pass", "Click 'Trang chủ' → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_004": ("Pass", "AI prompt + click → navigate DS + loading state", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_005": ("Pass", "Style tags exclusive selection — chỉ 1 tag active cùng lúc", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_006": ("Pass", "'Chọn từ mẫu có sẵn' → /studio/?tab=library", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_007": ("Pass", "'Tải lên ảnh của bạn' → /studio/?tab=images", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_008": ("Pass", "Nhập text + Enter → submit navigate DS", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_021": ("Pass", "Streetwear tag highlight đúng, auto-fill prompt tương ứng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_022": ("Pass", "Vintage/Retro tag highlight đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_023": ("Pass", "Y2K tag highlight đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_024": ("Pass", "Abstract Art/Nghệ thuật tag highlight đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_025": ("Pass", "Streetwear tag + custom prompt → generate thành công", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_026": ("Pass", "Badge AI-Powered Design KHÔNG phải link, đúng expected", "home_remaining_tests_1774336927869.webp", "", ""),
        # ── SEO ──
        "TC_HOME_014": ("Fail", "Title tab = 'POD Admin CMS' thay vì 'Tryonic AI'", "home_remaining_tests_1774336927869.webp", "SEO: Wrong brand title", "BUG-HOME-003"),
        "TC_HOME_015": ("Fail", "Meta description = 'Admin CMS for POD T-Shirt Platform'", "home_remaining_tests_1774336927869.webp", "SEO: Wrong meta description", "BUG-HOME-004"),
        "TC_HOME_016": ("Pass", "1 h1 chính, h2/h3 hierarchy đúng thứ tự", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_017": ("Pass", "Tab order hợp lý, focus ring rõ trên elements", "home_remaining_tests_1774336927869.webp", "", ""),
    },
    "DESIGN STUDIO": {
        "TC_DS_UI_001": ("Pass", "Header: Quay lại, Logo, Design Studio, 12 Credits, User, Cart", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_002": ("Pass", "Credits badge = 12 Credits + Chia sẻ", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_003": ("Pass", "Sidebar: 6 tools đầy đủ. Hoàn Tác/Làm Lại disabled mặc định", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_004": ("Pass", "Canvas T-shirt mockup trắng, vùng thiết kế nét đứt xanh cyan", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_005": ("Pass", "5 tabs: SẢN PHẨM, ẢNH CỦA BẠN, THƯ VIỆN, TẠO ẢNH AI, ĐẶT HÀNG", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_008": ("Pass", "Bottom bar: product info, color toggle, size L, 150.000đ, Đặt hàng", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_001": ("Pass", "← Quay lại → navigate /home/", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_002": ("Pass", "Mặt Sau toggle canvas thành công", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_005": ("Pass", "Smooth switching giữa 5 tabs", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_014": ("Pass", "Hoàn Tác present, disabled khi chưa có thay đổi", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_022": ("Pass", "Tab SẢN PHẨM content đúng", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DS_023": ("Pass", "Tab ẢNH CỦA BẠN: upload zone hiển thị", "ds_tab_anh_cua_ban_1774328778359.png", "", ""),
        "TC_DS_024": ("Pass", "Tab THƯ VIỆN: template thumbnails hiển thị", "ds_tab_thu_vien_1774328796600.png", "", ""),
        "TC_DS_025": ("Pass", "Tab ĐẶT HÀNG: order form hiển thị", "ds_tab_dat_hang_1774328831949.png", "", ""),
        "TC_DSP_UI_001": ("Pass", "SẢN PHẨM: product info, color picker, size selector", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DSP_UI_005": ("Pass", "'Đổi sản phẩm' button hiển thị rõ", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DSP_UI_006": ("Pass", "'Gợi ý size' button hiển thị rõ", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DSA_UI_001": ("Pass", "Upload zone: Kéo thả hoặc click, PNG/JPG/SVG/WebP", "ds_tab_anh_cua_ban_1774328778359.png", "", ""),
        "TC_DSTV_UI_001": ("Pass", "Template grid với thumbnails + search/filter", "ds_tab_thu_vien_1774328796600.png", "", ""),
    },
    "AI GENERATE": {
        "TC_AI_UI_001": ("Pass", "MÔ TẢ ARTWORK textarea với placeholder đúng", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_002": ("Pass", "ẢNH THAM KHẢO: Chọn từ thư viện + Tải ảnh lên", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_003": ("Pass", "6 style cards đầy đủ: Watercolor, Minimalist, Line Art, Retro, Grunge, Flat Design", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_004": ("Pass", "Tạo Artwork Mới btn: disabled khi rỗng, gradient khi enable", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_005": ("Pass", "'Mỗi lần tạo tốn 3 Credits' hiển thị rõ ràng", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_001": ("Pass", "Style switching Watercolor ↔ Line Art hoạt động đúng", "ai_generate_tests_1774329374136.webp", "", ""),
        "TC_AI_002": ("Pass", "Nhập prompt → nút Tạo Artwork enabled", "ai_generate_tests_1774329374136.webp", "", ""),
        "TC_AI_003": ("Pass", "Loading state → hiển thị 3 variants (A, B, C)", "ai_generation_results_1774329491079.png", "", ""),
        "TC_AI_004": ("Pass", "Click 'Chọn mẫu này' → artwork áp lên T-shirt canvas", "ai_artwork_applied_to_shirt_1774329508643.png", "", ""),
        "TC_AI_005": ("Fail", "Credits = 12 KHÔNG giảm sau generation (expected -3)", "ai_generation_results_1774329491079.png", "Mock mode: credits không bị trừ", "BUG-AI-001"),
        "TC_AI_006": ("Pass", "'Chọn từ thư viện' modal mở đúng với search + categories", "ai_generate_tests_1774329374136.webp", "", ""),
        "TC_AI_007": ("Pass", "Generate button disabled khi prompt rỗng", "ai_generate_tests_1774329374136.webp", "", ""),
    },
    "ĐẶT HÀNG": {
        "TC_DH_UI_001": ("Pass", "Size S-3XL, Quantity +/-, Price Summary hiển thị đúng", "ds_tab_dat_hang_1774328831949.png", "", ""),
        "TC_DH_UI_002": ("Pass", "'Thêm vào giỏ' và 'Mua ngay' buttons present", "ds_tab_dat_hang_1774328831949.png", "", ""),
        "TC_DH_001": ("Pass", "Đổi size XL → price update (150k→315k)", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_DH_002": ("Pass", "Tăng qty=2 → giá recalculate (315k→475k)", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_DH_003": ("Pass", "'Mua ngay' → navigate checkout page", "order_checkout_tests_1774329568589.webp", "", ""),
    },
    "THANH TOÁN": {
        "TC_TT_UI_001": ("Pass", "Checkout layout: Customer info, Order summary, Payment sections đầy đủ", "checkout_page_auth_check_1774329274021.png", "", ""),
        "TC_TT_UI_002": ("Pass", "7 fields: Họ tên, Phone, Email, Địa chỉ, Tỉnh, Quận, Phường", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_TT_UI_003": ("Pass", "MoMo payment method hiển thị", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
        "TC_TT_UI_004": ("Pass", "Order summary: products, sizes, quantities, total = 485.000đ", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
        "TC_TT_001": ("Pass", "Submit rỗng → Thanh toán disabled, validation errors", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_TT_002": ("Pass", "Fill info → fields accept input, Thanh toán enabled", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
        "TC_TT_003": ("Pass", "Thanh toán button hiển thị với tổng 485.000đ", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
    },
    "ORDER": {
        "TC_ORDER_UI_001": ("Pass", "Cart side panel accessible từ header cart icon", "cart_side_panel_accessible_1774329808737.png", "", ""),
    },
    "LOGIN": {
        "TC_AUTH_UI_001": ("Pass", "Title 'Chào mừng trở lại!' + login icon modal", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_002": ("Pass", "Email field placeholder 'name@example.com'", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_003": ("Fail", "Password masked nhưng KHÔNG CÓ eye-toggle icon", "login_modal_ui_1774329065496.png", "Missing eye-toggle", "BUG-LOGIN-001"),
        "TC_AUTH_UI_004": ("Pass", "'Quên mật khẩu?' link hiển thị", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_005": ("Pass", "'Đăng nhập' button teal full-width", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_006": ("Pass", "'Tiếp tục với Google' outlined + G icon", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_007": ("Pass", "'Tiếp tục với Facebook' blue + FB icon", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_008": ("Pass", "'HOẶC' separator hiển thị giữa OAuth và form", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_009": ("Pass", "'Đăng ký' link ở bottom modal", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_010": ("Fail", "Footer links (Điều khoản, Chính sách, Trợ giúp) THIẾU trong modal", "login_modal_ui_1774329065496.png", "Missing footer links", "BUG-LOGIN-002"),
        "TC_AUTH_UI_013": ("Pass", "Enter key submit login", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_021": ("Pass", "Submit empty → 'Vui lòng điền đầy đủ thông tin'", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_022": ("Pass", "Email only → validation error hiển thị", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_018": ("Pass", "Mock mode: login thành công với any creds", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_020": ("Pass", "Mock mode: login thành công", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_UI_015": ("Pass", "'Đăng ký tài khoản' title trong checkout", "checkout_page_auth_check_1774329274021.png", "", ""),
        "TC_AUTH_UI_016": ("Fail", "Field 'Họ và tên' THIẾU trong signup form", "checkout_page_auth_check_1774329274021.png", "Signup missing name", "BUG-LOGIN-003"),
        "TC_AUTH_UI_017": ("Fail", "'Xác nhận mật khẩu' THIẾU trong signup form", "checkout_page_auth_check_1774329274021.png", "Signup missing confirm pw", "BUG-LOGIN-004"),
        "TC_AUTH_UI_018": ("Fail", "Checkbox Điều khoản THIẾU", "checkout_page_auth_check_1774329274021.png", "Signup missing terms", "BUG-LOGIN-005"),
        "TC_AUTH_UI_019": ("Pass", "Google signup option present", "checkout_page_auth_check_1774329274021.png", "", ""),
        "TC_AUTH_UI_020": ("Pass", "Facebook signup option present", "checkout_page_auth_check_1774329274021.png", "", ""),
    },
}

# ─── Fill results into each sheet ───
total_filled = 0
total_pass = 0
total_fail = 0

for sheet_name, results in ALL_RESULTS.items():
    if sheet_name not in wb.sheetnames:
        print(f"⚠️ Sheet '{sheet_name}' not found")
        continue
    
    ws = wb[sheet_name]
    sheet_filled = 0
    
    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row, 1).value
        if tc_id and tc_id.strip() in results:
            tc_key = tc_id.strip()
            result, actual_result, evidence_file, notes, bug_id = results[tc_key]
            
            # Col 11 = Actual Result
            ws.cell(row, 11).value = actual_result
            ws.cell(row, 11).font = BODY_FONT
            ws.cell(row, 11).alignment = WRAP
            
            # Col 15 = Result_R1
            cell_result = ws.cell(row, 15)
            cell_result.value = result
            if result == "Pass":
                cell_result.fill = PASS_FILL
                cell_result.font = PASS_FONT
                total_pass += 1
            else:
                cell_result.fill = FAIL_FILL
                cell_result.font = FAIL_FONT
                total_fail += 1
            
            # Col 16 = Test Date_R1
            ws.cell(row, 16).value = now_str
            ws.cell(row, 16).font = BODY_FONT
            
            # Col 17 = Tester_R1
            ws.cell(row, 17).value = "AI Agent"
            ws.cell(row, 17).font = BODY_FONT
            
            # Col 18 = Bug ID_R1
            if bug_id:
                ws.cell(row, 18).value = bug_id
                ws.cell(row, 18).font = FAIL_FONT
            
            # Col 23 = Evidence (screenshot file path)
            if evidence_file:
                evidence_path = os.path.join(EVIDENCE_DIR, evidence_file)
                ws.cell(row, 23).value = evidence_file
                ws.cell(row, 23).font = Font(name="Calibri", size=10, color="0563C1", underline="single")
                ws.cell(row, 23).alignment = WRAP
            
            # Col 24 = Notes 
            if notes:
                ws.cell(row, 24).value = notes
                ws.cell(row, 24).font = BODY_FONT
            
            sheet_filled += 1
    
    print(f"  📋 {sheet_name}: {sheet_filled} TCs filled")
    total_filled += sheet_filled

wb.save(OUTPUT)
print(f"\n🎉 FULL Report saved: {os.path.basename(OUTPUT)}")
print(f"   Total filled: {total_filled} TCs with Actual_Result + Evidence")
print(f"   📁 Location: {REPORT_DIR}")
print(f"\n📊 OVERALL RESULTS")
print(f"   Total Tested: {total_pass + total_fail}")
print(f"   ✅ Pass: {total_pass} ({total_pass*100//(total_pass+total_fail) if total_pass+total_fail else 0}%)")
print(f"   ❌ Fail: {total_fail} ({total_fail*100//(total_pass+total_fail) if total_pass+total_fail else 0}%)")
print(f"\n🐛 ALL BUGS:")
bugs = []
for sheet, results in ALL_RESULTS.items():
    for tc, (res, actual, _, notes, bug) in results.items():
        if res == "Fail":
            bugs.append((bug, tc, actual))
            print(f"   [{bug}] {tc}: {actual}")

print(f"\n📎 Total bugs for Jira: {len(bugs)}")
