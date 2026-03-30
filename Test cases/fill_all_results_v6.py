"""
Fill ALL test results into v27_v3 Excel with Bug Description column.
New column structure (27 cols):
  15=Result_R1, 16=Test Date_R1, 17=Tester_R1, 18=Bug ID_R1, 19=Bug Desc_R1,
  20=Result_R2, ...22=Bug ID_R2, 23=Bug Desc_R2,
  25=Evidence, 26=Notes
"""
import openpyxl, os, datetime, shutil
from openpyxl.styles import Font, PatternFill, Alignment

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")
now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

REPORT_DIR = os.path.join(BASE, "Test_Reports")
EVIDENCE_SRC = r"C:\Users\maiho\.gemini\antigravity\brain\137607bd-c663-49d8-aa44-945410ab72b3"
EVIDENCE_DST = os.path.join(REPORT_DIR, "Evidence")
os.makedirs(EVIDENCE_DST, exist_ok=True)

SRC = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v27_{today}_v3.xlsx")
OUTPUT = os.path.join(REPORT_DIR, f"TestReport_FULL_v27_{today}_v6.xlsx")
shutil.copy2(SRC, OUTPUT)

wb = openpyxl.load_workbook(OUTPUT)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(name="Calibri", size=11, color="006100")
FAIL_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)
BODY_FONT = Font(name="Calibri", size=11)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
WRAP = Alignment(vertical='top', wrap_text=True)

# Bug descriptions
BUG_DESCRIPTIONS = {
    "BUG-HOME-001": "Header không sticky — biến mất hoàn toàn khi scroll xuống. Expected: Header cố định (sticky) ở trên cùng.",
    "BUG-HOME-002": "AI Input chấp nhận prompt chỉ có spaces, không trim whitespace trước khi validate. Navigate tới Design Studio thay vì hiển thị lỗi.",
    "BUG-HOME-003": "Title tag hiển thị 'POD Admin CMS' thay vì tên thương hiệu 'Tryonic AI'. Ảnh hưởng SEO nghiêm trọng.",
    "BUG-HOME-004": "Meta description = 'Admin CMS for POD T-Shirt Platform'. Không phù hợp cho trang khách hàng. Cần cập nhật SEO-friendly.",
    "BUG-AI-001": "Credits không giảm sau khi AI generate artwork. Tốn 3 credits nhưng số dư vẫn giữ nguyên 12. Có thể do mock mode.",
    "BUG-LOGIN-001": "Password field thiếu icon eye-toggle để show/hide password. UX standard cần có nút hiện/ẩn mật khẩu.",
    "BUG-LOGIN-002": "Login modal thiếu footer links: Điều khoản sử dụng, Chính sách bảo mật, Trợ giúp. Cần bổ sung để đúng spec.",
    "BUG-LOGIN-003": "Form đăng ký thiếu field 'Họ và tên'. Chỉ có Email + Password. Cần bổ sung Full Name field theo spec.",
    "BUG-LOGIN-004": "Form đăng ký thiếu field 'Xác nhận mật khẩu'. Người dùng không thể verify password khi đăng ký.",
    "BUG-LOGIN-005": "Form đăng ký thiếu checkbox đồng ý Điều khoản sử dụng. Cần bổ sung theo quy định pháp lý.",
}

# ═══ ALL RESULTS: (result, actual_result, evidence_file, notes, bug_id) ═══
ALL_RESULTS = {
    "HOME": {
        "TC_HOME_UI_001": ("Pass", "Logo 'Tryonic' với sparkle icon hiển thị góc trái header, click được", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_002": ("Pass", "Nav menu: Trang chủ, Sản phẩm, Dịch vụ, Liên hệ hiển thị đầy đủ", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_003": ("Pass", "Nút 'Thiết kế ngay' gradient tím→đỏ, text trắng, rounded-full", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_004": ("Fail", "Header KHÔNG sticky. Khi scroll xuống header biến mất hoàn toàn", "home_confirm_header_not_sticky_1774325609417.png", "", "BUG-HOME-001"),
        "TC_HOME_UI_005": ("Pass", "Badge '✨ AI-Powered Design' hiển thị", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_006": ("Pass", "Headline 'Biến ý tưởng thành áo thun trong 30 giây' hiển thị đúng", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_007": ("Pass", "Subtitle hiển thị đúng nội dung", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_008": ("Pass", "AI Input Box với placeholder đúng", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_009": ("Pass", "Nút Generate hiển thị cạnh input", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_010": ("Pass", "6 Style Tags đầy đủ", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_011": ("Pass", "'Chọn từ mẫu có sẵn' hiển thị đúng", "home_page_scrolled_1774322028177.png", "", ""),
        "TC_HOME_UI_012": ("Pass", "'Tải lên ảnh của bạn' hiển thị đúng", "home_page_scrolled_1774322028177.png", "", ""),
        "TC_HOME_UI_013": ("Pass", "Trust markers ở footer đầy đủ", "home_scrolled_header_footer_1774325588963.png", "", ""),
        "TC_HOME_UI_014": ("Pass", "Hover menu → đổi màu, cursor pointer", "home_ui_tests_1774325532508.webp", "", ""),
        "TC_HOME_UI_015": ("Pass", "Hover CTA → opacity change", "home_ui_tests_1774325532508.webp", "", ""),
        "TC_HOME_UI_016": ("Pass", "Hero gradient background lavender đúng", "home_hero_full_viewport_1774325561450.png", "", ""),
        "TC_HOME_UI_860": ("Pass", "Zoom 200%/50% layout stable", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_861": ("Pass", "iPhone 375x812: 1 cột, tags wrap OK", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_862": ("Pass", "Android 360x740: responsive OK", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_863": ("Pass", "iPad 768x1024: centered layout OK", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_864": ("Pass", "Tablet 800x1280: scale OK", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_UI_865": ("Pass", "Landscape 812x375 adapt OK", "responsive_emulated_tests_1774338152517.webp", "", ""),
        "TC_HOME_009": ("Pass", "Generate rỗng → DS, btn disabled đúng", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_010": ("Pass", "Click Sản phẩm → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_011": ("Pass", "Click Dịch vụ → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_012": ("Pass", "Click Liên hệ → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_013": ("Pass", "XSS sanitized, no script execution", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_018": ("Pass", "1-char prompt accepted", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_020": ("Fail", "Spaces-only navigate DS, no validate", "home_remaining_tests_1774336927869.webp", "", "BUG-HOME-002"),
        "TC_HOME_001": ("Pass", "'Thiết kế ngay' → /studio/", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_002": ("Pass", "Logo → /home/", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_003": ("Pass", "'Trang chủ' → /home/#", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_004": ("Pass", "AI prompt + click → DS loading OK", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_005": ("Pass", "Style tags exclusive selection OK", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_006": ("Pass", "'Chọn mẫu' → /studio/?tab=library", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_007": ("Pass", "'Tải ảnh' → /studio/?tab=images", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_008": ("Pass", "Enter → submit DS", "home_functional_tests_1774325655120.webp", "", ""),
        "TC_HOME_021": ("Pass", "Streetwear tag highlight + auto-fill", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_022": ("Pass", "Vintage tag highlight đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_023": ("Pass", "Y2K tag highlight đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_024": ("Pass", "Abstract Art tag highlight đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_025": ("Pass", "Streetwear + prompt → generate OK", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_026": ("Pass", "Badge non-interactive, đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_014": ("Fail", "Title = 'POD Admin CMS'", "home_remaining_tests_1774336927869.webp", "", "BUG-HOME-003"),
        "TC_HOME_015": ("Fail", "Meta desc = 'Admin CMS...'", "home_remaining_tests_1774336927869.webp", "", "BUG-HOME-004"),
        "TC_HOME_016": ("Pass", "1 h1, hierarchy đúng", "home_remaining_tests_1774336927869.webp", "", ""),
        "TC_HOME_017": ("Pass", "Tab order OK, focus ring rõ", "home_remaining_tests_1774336927869.webp", "", ""),
    },
    "DESIGN STUDIO": {
        "TC_DS_UI_001": ("Pass", "Header đầy đủ elements", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_002": ("Pass", "12 Credits badge OK", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_003": ("Pass", "Sidebar 6 tools OK", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_004": ("Pass", "Canvas mockup đúng", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_005": ("Pass", "5 tabs OK", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_UI_008": ("Pass", "Bottom bar OK", "ds_initial_layout_1774328741540.png", "", ""),
        "TC_DS_001": ("Pass", "← Quay lại → /home/", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_002": ("Pass", "Mặt Sau toggle OK", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_005": ("Pass", "Tab switching smooth", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_014": ("Pass", "Hoàn Tác disabled OK", "design_studio_tests_1774328706450.webp", "", ""),
        "TC_DS_022": ("Pass", "SẢN PHẨM content OK", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DS_023": ("Pass", "ẢNH CỦA BẠN upload OK", "ds_tab_anh_cua_ban_1774328778359.png", "", ""),
        "TC_DS_024": ("Pass", "THƯ VIỆN templates OK", "ds_tab_thu_vien_1774328796600.png", "", ""),
        "TC_DS_025": ("Pass", "ĐẶT HÀNG form OK", "ds_tab_dat_hang_1774328831949.png", "", ""),
        "TC_DSP_UI_001": ("Pass", "SẢN PHẨM panel OK", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DSP_UI_005": ("Pass", "'Đổi sản phẩm' button OK", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DSP_UI_006": ("Pass", "'Gợi ý size' button OK", "ds_tab_san_pham_1774328760405.png", "", ""),
        "TC_DSA_UI_001": ("Pass", "Upload zone OK", "ds_tab_anh_cua_ban_1774328778359.png", "", ""),
        "TC_DSTV_UI_001": ("Pass", "Template grid OK", "ds_tab_thu_vien_1774328796600.png", "", ""),
    },
    "AI GENERATE": {
        "TC_AI_UI_001": ("Pass", "MÔ TẢ ARTWORK textarea OK", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_002": ("Pass", "ẢNH THAM KHẢO buttons OK", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_003": ("Pass", "6 style cards OK", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_004": ("Pass", "Tạo Artwork btn state OK", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_UI_005": ("Pass", "3 Credits info OK", "ai_generate_initial_view_1774329436903.png", "", ""),
        "TC_AI_001": ("Pass", "Style switching OK", "ai_generate_tests_1774329374136.webp", "", ""),
        "TC_AI_002": ("Pass", "Prompt → enable btn OK", "ai_generate_tests_1774329374136.webp", "", ""),
        "TC_AI_003": ("Pass", "Loading → 3 variants OK", "ai_generation_results_1774329491079.png", "", ""),
        "TC_AI_004": ("Pass", "Apply artwork to canvas OK", "ai_artwork_applied_to_shirt_1774329508643.png", "", ""),
        "TC_AI_005": ("Fail", "Credits = 12 không giảm", "ai_generation_results_1774329491079.png", "", "BUG-AI-001"),
        "TC_AI_006": ("Pass", "Library modal OK", "ai_generate_tests_1774329374136.webp", "", ""),
        "TC_AI_007": ("Pass", "Empty prompt → disabled OK", "ai_generate_tests_1774329374136.webp", "", ""),
    },
    "ĐẶT HÀNG": {
        "TC_DH_UI_001": ("Pass", "Size, Qty, Price OK", "ds_tab_dat_hang_1774328831949.png", "", ""),
        "TC_DH_UI_002": ("Pass", "Thêm giỏ + Mua ngay OK", "ds_tab_dat_hang_1774328831949.png", "", ""),
        "TC_DH_001": ("Pass", "Size XL → price update OK", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_DH_002": ("Pass", "Qty=2 → recalculate OK", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_DH_003": ("Pass", "Mua ngay → checkout OK", "order_checkout_tests_1774329568589.webp", "", ""),
    },
    "THANH TOÁN": {
        "TC_TT_UI_001": ("Pass", "Checkout layout đầy đủ", "checkout_page_auth_check_1774329274021.png", "", ""),
        "TC_TT_UI_002": ("Pass", "7 customer fields OK", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_TT_UI_003": ("Pass", "MoMo payment OK", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
        "TC_TT_UI_004": ("Pass", "Order summary 485.000đ", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
        "TC_TT_001": ("Pass", "Empty → disabled + errors", "order_checkout_tests_1774329568589.webp", "", ""),
        "TC_TT_002": ("Pass", "Fill → enabled OK", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
        "TC_TT_003": ("Pass", "Thanh toán btn 485.000đ", "checkout_form_filled_enabled_button_1774329755041.png", "", ""),
    },
    "ORDER": {
        "TC_ORDER_UI_001": ("Pass", "Cart panel accessible OK", "cart_side_panel_accessible_1774329808737.png", "", ""),
    },
    "LOGIN": {
        "TC_AUTH_UI_001": ("Pass", "'Chào mừng trở lại!' modal OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_002": ("Pass", "Email placeholder OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_003": ("Fail", "KHÔNG CÓ eye-toggle icon", "login_modal_ui_1774329065496.png", "", "BUG-LOGIN-001"),
        "TC_AUTH_UI_004": ("Pass", "'Quên mật khẩu?' OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_005": ("Pass", "Đăng nhập btn teal OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_006": ("Pass", "Google OAuth OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_007": ("Pass", "Facebook OAuth OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_008": ("Pass", "'HOẶC' separator OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_009": ("Pass", "'Đăng ký' link OK", "login_modal_ui_1774329065496.png", "", ""),
        "TC_AUTH_UI_010": ("Fail", "Footer links THIẾU", "login_modal_ui_1774329065496.png", "", "BUG-LOGIN-002"),
        "TC_AUTH_UI_013": ("Pass", "Enter submit OK", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_021": ("Pass", "Empty submit → error OK", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_022": ("Pass", "Email-only → error OK", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_018": ("Pass", "Mock login OK", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_020": ("Pass", "Mock login OK", "login_page_tests_1774328929094.webp", "", ""),
        "TC_AUTH_UI_015": ("Pass", "Signup title OK", "checkout_page_auth_check_1774329274021.png", "", ""),
        "TC_AUTH_UI_016": ("Fail", "Thiếu 'Họ và tên'", "checkout_page_auth_check_1774329274021.png", "", "BUG-LOGIN-003"),
        "TC_AUTH_UI_017": ("Fail", "Thiếu 'Xác nhận MK'", "checkout_page_auth_check_1774329274021.png", "", "BUG-LOGIN-004"),
        "TC_AUTH_UI_018": ("Fail", "Thiếu checkbox ĐK", "checkout_page_auth_check_1774329274021.png", "", "BUG-LOGIN-005"),
        "TC_AUTH_UI_019": ("Pass", "Google signup OK", "checkout_page_auth_check_1774329274021.png", "", ""),
        "TC_AUTH_UI_020": ("Pass", "Facebook signup OK", "checkout_page_auth_check_1774329274021.png", "", ""),
    },
}

# Copy evidence files
for f in os.listdir(EVIDENCE_SRC):
    if f.endswith(('.png', '.webp', '.jpg')):
        src = os.path.join(EVIDENCE_SRC, f)
        dst = os.path.join(EVIDENCE_DST, f)
        if not os.path.exists(dst):
            shutil.copy2(src, dst)

# Fill results
total_filled = 0
total_pass = total_fail = 0

for sheet_name, results in ALL_RESULTS.items():
    if sheet_name not in wb.sheetnames:
        continue
    ws = wb[sheet_name]
    filled = 0
    for row in range(2, ws.max_row + 1):
        tc_id = ws.cell(row, 1).value
        if tc_id and tc_id.strip() in results:
            tc_key = tc_id.strip()
            result, actual, evidence_file, notes, bug_id = results[tc_key]
            
            # Col 11 = Actual Result
            ws.cell(row, 11).value = actual
            ws.cell(row, 11).font = BODY_FONT
            ws.cell(row, 11).alignment = WRAP
            
            # Col 15 = Result_R1
            c = ws.cell(row, 15)
            c.value = result
            if result == "Pass":
                c.fill = PASS_FILL; c.font = PASS_FONT; total_pass += 1
            else:
                c.fill = FAIL_FILL; c.font = FAIL_FONT; total_fail += 1
            
            # Col 16 = Date, Col 17 = Tester
            ws.cell(row, 16).value = now_str; ws.cell(row, 16).font = BODY_FONT
            ws.cell(row, 17).value = "AI Agent"; ws.cell(row, 17).font = BODY_FONT
            
            # Col 18 = Bug ID_R1, Col 19 = Bug Desc_R1
            if bug_id:
                ws.cell(row, 18).value = bug_id
                ws.cell(row, 18).font = FAIL_FONT
                ws.cell(row, 19).value = BUG_DESCRIPTIONS.get(bug_id, "")
                ws.cell(row, 19).font = BODY_FONT
                ws.cell(row, 19).alignment = WRAP
            
            # Col 25 = Evidence (hyperlink)
            if evidence_file:
                dst_path = os.path.join(EVIDENCE_DST, evidence_file)
                if os.path.exists(dst_path):
                    abs_path = os.path.abspath(dst_path)
                    ws.cell(row, 25).hyperlink = abs_path
                    ws.cell(row, 25).value = evidence_file
                    ws.cell(row, 25).font = LINK_FONT
                    ws.cell(row, 25).alignment = WRAP
            
            # Col 26 = Notes
            if notes:
                ws.cell(row, 26).value = notes
                ws.cell(row, 26).font = BODY_FONT
            
            filled += 1
    
    print(f"  📋 {sheet_name}: {filled} TCs")
    total_filled += filled

wb.save(OUTPUT)
print(f"\n🎉 Report v6 saved: {os.path.basename(OUTPUT)}")
print(f"   {total_filled} TCs | ✅ {total_pass} Pass | ❌ {total_fail} Fail")
print(f"\n🐛 Bugs with descriptions:")
for bug_id, desc in BUG_DESCRIPTIONS.items():
    print(f"   [{bug_id}] {desc[:80]}...")
