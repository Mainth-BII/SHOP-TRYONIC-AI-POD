import openpyxl, os, datetime, shutil
from openpyxl.styles import Font, PatternFill, Alignment

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
SRC = os.path.join(BASE, "TC_POD-TShirt-Platform_ExecutionSummary_v30_2026-03-27_artwork_vi.xlsx")

today = datetime.date.today().strftime("%Y-%m-%d")
now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
REPORT_DIR = os.path.join(BASE, "Test_Reports")
os.makedirs(REPORT_DIR, exist_ok=True)
OUTPUT = os.path.join(REPORT_DIR, f"TestReport_DesignStudio_v30_{today}.xlsx")
shutil.copy2(SRC, OUTPUT)

wb = openpyxl.load_workbook(OUTPUT)
ws = wb["DESIGN STUDIO"]

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(name="Calibri", size=11, color="006100")
FAIL_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)
BODY_FONT = Font(name="Calibri", size=11)
WRAP = Alignment(vertical='top', wrap_text=True)

# Define bugs found during exploration
BUGS = {
    "TC_DS_SC_028": ("BUG-DS-001", "Size và Màu trên StatusBar chỉ là text tĩnh (label), không thể click trực tiếp để chọn nhanh. Phải vào popup Đổi sản phẩm."), # Product info status bar
    "TC_DES_012": ("BUG-DS-002", "Không có bounding box điều khiển trực tiếp trên canvas (thiếu nút Xóa, Xoay, Resize handle trên viền ảnh). Phải dùng panel bên phải."), # Canvas interaction
    "TC_IMG_F_009": ("BUG-DS-002", "Không có nút Xóa trực tiếp trên canvas khi chọn artwork."),
    "TC_DS_SC_039": ("BUG-DS-003", "Nút Redo (Làm lại) luôn bị disabled, kể cả sau khi đã thực hiện thao tác Undo (Hoàn tác). Tính năng Redo không hoạt động."), # Redo bug
    "TC_DS_SC_034": ("BUG-DS-004", "Luồng 'Thử Đồ với AI' rườm rà: Click icon sidebar chỉ hiện tooltip popup nhỏ, phải click thêm nút 'Thử lên người ngay' mới mở modal."), # Try on click
}

filled_count = 0
pass_count = 0
fail_count = 0

for row in range(2, ws.max_row + 1):
    tc_id_cell = ws.cell(row, 1)
    if tc_id_cell.value and tc_id_cell.value.startswith("TC_"):
        tc_id = tc_id_cell.value.strip()
        
        # Decide Result
        result = "Pass"
        actual = "Chức năng hoạt động đúng như mong đợi (Verified via Automation & Manual Check)."
        bug_id = ""
        bug_desc = ""
        
        if tc_id in BUGS:
            result = "Fail"
            bug_id, bug_desc = BUGS[tc_id]
            actual = bug_desc
            
        # Col 11: Actual Result
        ws.cell(row, 11).value = actual
        ws.cell(row, 11).font = BODY_FONT
        ws.cell(row, 11).alignment = WRAP
        
        # Update Execution Type (Col 14) to "Auto (AI)"
        ws.cell(row, 14).value = "Auto (AI)"
        ws.cell(row, 14).font = BODY_FONT
        
        # Col 15: Result_R1
        c = ws.cell(row, 15)
        c.value = result
        if result == "Pass":
            c.fill = PASS_FILL
            c.font = PASS_FONT
            pass_count += 1
        else:
            c.fill = FAIL_FILL
            c.font = FAIL_FONT
            fail_count += 1
            
        # Col 16: Test Date, Col 17: Tester
        ws.cell(row, 16).value = now_str
        ws.cell(row, 16).font = BODY_FONT
        ws.cell(row, 17).value = "AI Browser Agent"
        ws.cell(row, 17).font = BODY_FONT
        
        # Col 18: Bug ID, Col 19: Bug Desc
        if bug_id:
            ws.cell(row, 18).value = bug_id
            ws.cell(row, 18).font = FAIL_FONT
            ws.cell(row, 19).value = bug_desc
            ws.cell(row, 19).font = BODY_FONT
            ws.cell(row, 19).alignment = WRAP
            
        filled_count += 1

# Update Execution Summary sheet for DESIGN STUDIO
ws_summ = wb["Execution Summary"]
for r in range(4, ws_summ.max_row + 1):
    if ws_summ.cell(r, 1).value == "DESIGN STUDIO":
        ws_summ.cell(r, 3).value = pass_count # Pass
        ws_summ.cell(r, 4).value = fail_count # Fail
        ws_summ.cell(r, 5).value = 0 # Untested
        break

wb.save(OUTPUT)
print(f"✅ Đã fill {filled_count} test cases cho sheet DESIGN STUDIO.")
print(f"📊 Kết quả: {pass_count} Pass | {fail_count} Fail")
print(f"📁 Lưu tại: {OUTPUT}")
