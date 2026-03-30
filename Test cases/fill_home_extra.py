"""Fill additional HOME test results from the remaining test batch."""
import openpyxl, os, datetime
from openpyxl.styles import Font, PatternFill

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")
now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

REPORT_DIR = os.path.join(BASE, "Test_Reports")
OUTPUT_SRC = os.path.join(REPORT_DIR, f"TestReport_FULL_v27_{today}.xlsx")
OUTPUT = os.path.join(REPORT_DIR, f"TestReport_FULL_v27_{today}_v2.xlsx")
import shutil
shutil.copy2(OUTPUT_SRC, OUTPUT)

wb = openpyxl.load_workbook(OUTPUT)

PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(name="Calibri", size=11, color="006100")
FAIL_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)
BODY_FONT = Font(name="Calibri", size=11)

# Additional HOME results
EXTRA_RESULTS = {
    "TC_HOME_013": ("Pass", "XSS payload sanitized, rendered as plain text. No script execution.", ""),
    "TC_HOME_018": ("Pass", "1-char prompt 'A' accepted, generated 3 artworks. No min-length validation.", ""),
    "TC_HOME_019": ("N/A", "5000+ char test skipped - browser input limitation. Need manual test.", ""),
    "TC_HOME_020": ("Fail", "Spaces-only prompt navigate tới DS KHÔNG validate. Thiếu trim whitespace.", "BUG-HOME-002"),
    "TC_HOME_021": ("Pass", "Streetwear tag highlight đúng, auto-fill prompt tương ứng", ""),
    "TC_HOME_022": ("Pass", "Vintage/Retro tag highlight đúng", ""),
    "TC_HOME_023": ("Pass", "Y2K tag highlight đúng", ""),
    "TC_HOME_024": ("Pass", "Abstract Art/Nghệ thuật tag highlight đúng", ""),
    "TC_HOME_025": ("Pass", "Streetwear tag + custom prompt → navigate DS + generate thành công", ""),
    "TC_HOME_026": ("Pass", "Badge AI-Powered Design KHÔNG phải link, đúng expected", ""),
    "TC_HOME_014": ("Fail", "Title tab = 'POD Admin CMS' thay vì 'Tryonic AI'. Sai branding SEO.", "BUG-HOME-003"),
    "TC_HOME_015": ("Fail", "Meta description = 'Admin CMS for POD T-Shirt Platform'. Sai SEO.", "BUG-HOME-004"),
    "TC_HOME_016": ("Pass", "Heading hierarchy đúng: 1 h1 chính, h2/h3 đúng thứ tự", ""),
    "TC_HOME_017": ("Pass", "Tab order hợp lý, focus ring hiển thị rõ trên interactive elements", ""),
    "TC_HOME_UI_860": ("Pass", "Zoom 200% và 50% layout vẫn intact", ""),
    "TC_HOME_UI_861": ("Fail", "iPhone 375px: Header Logo + nút 'Thiết kế ngay' bị overlap", "BUG-HOME-005"),
    "TC_HOME_UI_862": ("N/A", "Cần test trên thiết bị Android thật", ""),
    "TC_HOME_UI_863": ("N/A", "Cần test trên iPad thật", ""),
    "TC_HOME_UI_864": ("N/A", "Cần test trên Tablet Android thật", ""),
    "TC_HOME_UI_865": ("Pass", "Landscape 812x375 layout adapt đúng", ""),
}

ws = wb["HOME"]
filled = 0
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, 1).value
    if tc_id and tc_id.strip() in EXTRA_RESULTS:
        tc_key = tc_id.strip()
        result, notes, bug_id = EXTRA_RESULTS[tc_key]
        
        # Skip if already filled
        existing = ws.cell(row, 14).value
        if existing and existing in ('Pass', 'Fail'):
            continue
        
        cell_result = ws.cell(row, 14)
        cell_result.value = result
        if result == "Pass":
            cell_result.fill = PASS_FILL
            cell_result.font = PASS_FONT
        elif result == "Fail":
            cell_result.fill = FAIL_FILL
            cell_result.font = FAIL_FONT
        else:
            cell_result.font = BODY_FONT
        
        ws.cell(row, 15).value = now_str
        ws.cell(row, 15).font = BODY_FONT
        ws.cell(row, 16).value = "AI Agent"
        ws.cell(row, 16).font = BODY_FONT
        
        if bug_id:
            ws.cell(row, 17).value = bug_id
            ws.cell(row, 17).font = FAIL_FONT
        
        ws.cell(row, 23).value = notes
        ws.cell(row, 23).font = BODY_FONT
        filled += 1

wb.save(OUTPUT)
print(f"✅ Filled {filled} additional HOME TCs into report")

# Count totals
pass_c = fail_c = na_c = untested_c = 0
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, 1).value
    result = ws.cell(row, 14).value
    if tc_id:
        if result == "Pass": pass_c += 1
        elif result == "Fail": fail_c += 1
        elif result == "N/A": na_c += 1
        else: untested_c += 1

print(f"\n📊 HOME SHEET COVERAGE")
print(f"   ✅ Pass: {pass_c}")
print(f"   ❌ Fail: {fail_c}")
print(f"   ⚠️  N/A: {na_c}")
print(f"   ❓ Untested: {untested_c}")
print(f"   Total: {pass_c+fail_c+na_c+untested_c}")
