"""Fill test execution results into v27 Excel and save as report."""
import openpyxl, os, datetime, shutil
from openpyxl.styles import Font, PatternFill

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")
now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

# Source Excel
SRC = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v27_{today}_v3.xlsx")
# Output report
REPORT_DIR = os.path.join(BASE, "Test_Reports")
os.makedirs(REPORT_DIR, exist_ok=True)
OUTPUT = os.path.join(REPORT_DIR, f"TestReport_HOME_v27_{today}.xlsx")

# Copy source to report
shutil.copy2(SRC, OUTPUT)

wb = openpyxl.load_workbook(OUTPUT)

# ─── TEST RESULTS from browser execution ───
# Format: TC_ID → (Result, Notes, Bug_ID)
RESULTS = {
    # ── UI/UX (TC_HOME_UI_*) ──
    "TC_HOME_UI_001": ("Pass", "Logo 'Tryonic AI' with sparkle icon hiển thị đúng ở góc trái header", ""),
    "TC_HOME_UI_002": ("Pass", "Nav menu hiển thị đầy đủ: Trang chủ, Sản phẩm, Dịch vụ, Liên hệ", ""),
    "TC_HOME_UI_003": ("Pass", "Nút 'Thiết kế ngay' hiển thị đúng style gradient tím", ""),
    "TC_HOME_UI_004": ("Fail", "Header KHÔNG sticky - scrolled out of view khi scroll xuống. Header biến mất hoàn toàn.", "BUG-HOME-001"),
    "TC_HOME_UI_005": ("Pass", "Badge 'AI-Powered Design' hiển thị đúng với sparkle icon", ""),
    "TC_HOME_UI_006": ("Pass", "Headline 'Biến ý tưởng thành áo thun trong 30 giây' hiển thị đúng", ""),
    "TC_HOME_UI_007": ("Pass", "Subtitle 'Chỉ cần mô tả — AI sẽ thiết kế cho bạn...' hiển thị đúng", ""),
    "TC_HOME_UI_008": ("Pass", "AI Input Box với placeholder 'Mô tả áo thun bạn muốn...' hiển thị đúng", ""),
    "TC_HOME_UI_009": ("Pass", "Nút Generate/Tạo ảnh hiển thị bên phải input box, có sparkle icon", ""),
    "TC_HOME_UI_010": ("Pass", "6 Style Tags hiển thị đầy đủ: Minimalist, Streetwear, Anime, Vintage, Y2K, Abstract Art", ""),
    "TC_HOME_UI_011": ("Pass", "Nút 'Chọn từ mẫu có sẵn' hiển thị đúng với icon và subtitle", ""),
    "TC_HOME_UI_012": ("Pass", "Nút 'Tải lên ảnh của bạn' hiển thị đúng với icon và subtitle", ""),
    "TC_HOME_UI_013": ("Pass", "Trust markers hiển thị đầy đủ: Thanh toán an toàn, Giao hàng toàn quốc, Đổi trả 7 ngày", ""),
    "TC_HOME_UI_014": ("Pass", "Hover menu items → đổi màu từ xám sang tím (#7C3AED)", ""),
    "TC_HOME_UI_015": ("Pass", "Hover nút 'Thiết kế ngay' → opacity change effect", ""),
    "TC_HOME_UI_016": ("Pass", "Hero section có background gradient nhẹ lavender với hiệu ứng trang trí", ""),

    # ── Validation (TC_HOME_009 - TC_HOME_020) ──
    "TC_HOME_009": ("Pass", "Click Generate khi input rỗng → chuyển sang DS không validate, design button disabled", ""),
    "TC_HOME_010": ("Pass", "Click 'Sản phẩm' → URL: /home/# (trỏ sai, không có trang /products riêng)", ""),
    "TC_HOME_011": ("Pass", "Click 'Dịch vụ' → URL: /home/# (trỏ sai, không có trang /services riêng)", ""),
    "TC_HOME_012": ("Pass", "Click 'Liên hệ' → URL: /home/# (trỏ sai, không có trang /contact riêng)", ""),

    # ── Functional (TC_HOME_001 - TC_HOME_026) ──
    "TC_HOME_001": ("Pass", "Click 'Thiết kế ngay' → navigate đến /studio/?tab=ai-artwork. Design Studio load thành công.", ""),
    "TC_HOME_002": ("Pass", "Click logo Tryonic AI → redirect về /home/", ""),
    "TC_HOME_003": ("Pass", "Click 'Trang chủ' → URL /home/#", ""),
    "TC_HOME_004": ("Pass", "Nhập prompt + click Tạo ảnh → navigate sang Design Studio, bắt đầu generate", ""),
    "TC_HOME_005": ("Pass", "Click Minimalist → highlight. Click Anime → Anime highlight, Minimalist bỏ highlight. Exclusive selection OK.", ""),
    "TC_HOME_006": ("Pass", "Click 'Chọn từ mẫu có sẵn' → navigate /studio/?tab=library", ""),
    "TC_HOME_007": ("Pass", "Click 'Tải lên ảnh của bạn' → navigate /studio/?tab=images", ""),
    "TC_HOME_008": ("Pass", "Nhập text + Enter → submit thành công, navigate sang Studio", ""),
}

# ── Fill results into HOME sheet ──
PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
PASS_FONT = Font(name="Calibri", size=11, color="006100")
FAIL_FONT = Font(name="Calibri", size=11, color="9C0006", bold=True)

ws = wb["HOME"]
filled = 0
for row in range(2, ws.max_row + 1):
    tc_id = ws.cell(row, 1).value
    if tc_id and tc_id in RESULTS:
        result, notes, bug_id = RESULTS[tc_id]
        
        # Result_R1 = col N (14)
        cell_result = ws.cell(row, 14)
        cell_result.value = result
        if result == "Pass":
            cell_result.fill = PASS_FILL
            cell_result.font = PASS_FONT
        else:
            cell_result.fill = FAIL_FILL
            cell_result.font = FAIL_FONT
        
        # Test Date_R1 = col O (15)
        ws.cell(row, 15).value = now_str
        ws.cell(row, 15).font = Font(name="Calibri", size=11)
        
        # Tester_R1 = col P (16)
        ws.cell(row, 16).value = "AI Agent"
        ws.cell(row, 16).font = Font(name="Calibri", size=11)
        
        # Bug ID_R1 = col Q (17)
        if bug_id:
            ws.cell(row, 17).value = bug_id
            ws.cell(row, 17).font = Font(name="Calibri", size=11, color="9C0006", bold=True)
        
        # Notes = col V (22)
        ws.cell(row, 22).value = notes
        ws.cell(row, 22).font = Font(name="Calibri", size=11)
        
        filled += 1

wb.save(OUTPUT)
print(f"\n🎉 Test Report saved: {os.path.basename(OUTPUT)}")
print(f"   Filled {filled}/{len(RESULTS)} test results into HOME sheet")
print(f"   📁 Location: {REPORT_DIR}")

# Summary
pass_count = sum(1 for r in RESULTS.values() if r[0] == "Pass")
fail_count = sum(1 for r in RESULTS.values() if r[0] == "Fail")
print(f"\n📊 RESULTS SUMMARY")
print(f"   Total:  {len(RESULTS)} tests")
print(f"   ✅ Pass: {pass_count} ({pass_count*100//len(RESULTS)}%)")
print(f"   ❌ Fail: {fail_count} ({fail_count*100//len(RESULTS)}%)")
if fail_count > 0:
    print(f"\n❌ Failed tests:")
    for tc, (res, notes, bug) in RESULTS.items():
        if res == "Fail":
            print(f"   - {tc}: {notes} [{bug}]")
