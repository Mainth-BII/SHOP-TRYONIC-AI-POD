"""
Copy evidence screenshots to Test_Reports/Evidence/ folder 
and update Excel Evidence column with clickable hyperlinks.
"""
import openpyxl, os, datetime, shutil
from openpyxl.styles import Font, Alignment

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")

REPORT_DIR = os.path.join(BASE, "Test_Reports")
EVIDENCE_SRC = r"C:\Users\maiho\.gemini\antigravity\brain\137607bd-c663-49d8-aa44-945410ab72b3"
EVIDENCE_DST = os.path.join(REPORT_DIR, "Evidence")
os.makedirs(EVIDENCE_DST, exist_ok=True)

# Copy all evidence files
copied = set()
for f in os.listdir(EVIDENCE_SRC):
    if f.endswith(('.png', '.webp', '.jpg', '.jpeg')):
        src = os.path.join(EVIDENCE_SRC, f)
        dst = os.path.join(EVIDENCE_DST, f)
        if not os.path.exists(dst):
            shutil.copy2(src, dst)
        copied.add(f)

print(f"📁 Copied {len(copied)} evidence files to Evidence/")

# Open report and update Evidence column with hyperlinks
SRC_XL = os.path.join(REPORT_DIR, f"TestReport_FULL_v27_{today}_v4.xlsx")
OUTPUT = os.path.join(REPORT_DIR, f"TestReport_FULL_v27_{today}_v5.xlsx")
shutil.copy2(SRC_XL, OUTPUT)

wb = openpyxl.load_workbook(OUTPUT)
LINK_FONT = Font(name="Calibri", size=10, color="0563C1", underline="single")
WRAP = Alignment(vertical='top', wrap_text=True)

updated = 0
for ws in wb.worksheets:
    if ws.title in ("Cover Page", "Execution Summary"):
        continue
    
    for row in range(2, ws.max_row + 1):
        evidence_val = ws.cell(row, 23).value  # col W = Evidence
        if evidence_val and isinstance(evidence_val, str) and evidence_val.strip():
            filename = evidence_val.strip()
            evidence_path = os.path.join(EVIDENCE_DST, filename)
            
            if os.path.exists(evidence_path):
                # Create clickable hyperlink using absolute path
                abs_path = os.path.abspath(evidence_path)
                ws.cell(row, 23).hyperlink = abs_path
                ws.cell(row, 23).value = filename
                ws.cell(row, 23).font = LINK_FONT
                ws.cell(row, 23).alignment = WRAP
                updated += 1

wb.save(OUTPUT)
print(f"✅ Updated {updated} Evidence cells with clickable hyperlinks")
print(f"📁 Report: {os.path.basename(OUTPUT)}")
print(f"📁 Evidence folder: {EVIDENCE_DST}")
print(f"   Total evidence files: {len(os.listdir(EVIDENCE_DST))}")
