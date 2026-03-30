"""
Tạo 77 hình ảnh sample cho Artwork Test Data và cập nhật vào Excel.
File này sẽ sinh ra thư mục Artwork_Samples chứa 77 file ảnh.
Mỗi ảnh sẽ có tên file và text trên ảnh thể hiện rõ loại artwork.
Cuối cùng update file Excel với cột 'Link Ảnh Sample'.
"""
import os, sys, datetime
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image, ImageDraw, ImageFont
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

BASE = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
today = datetime.date.today().strftime("%Y-%m-%d")
EXCEL_FILE = os.path.join(BASE, f"TC_POD-TShirt-Platform_ExecutionSummary_v30_{today}_artwork_vi.xlsx")
IMG_DIR = os.path.join(BASE, "Artwork_Samples")

if not os.path.exists(IMG_DIR):
    os.makedirs(IMG_DIR)

# Lấy dữ liệu từ file python đã tạo
try:
    from artwork_data_vi import ARTWORK_DATA_VI
except ImportError:
    print("❌ Lỗi: Không tìm thấy artwork_data_vi.py")
    sys.exit(1)

# Color palettes by category
CAT_COLORS = {
    "🎨 NGHỆ THUẬT TRUYỀN THỐNG": ((242, 220, 219), (105, 59, 58)),  # Light pink, dark red text
    "💻 NGHỆ THUẬT SỐ": ((220, 230, 241), (54, 96, 146)),          # Light blue, dark blue text
    "🤖 PHONG CÁCH AI TẠO SINH": ((228, 223, 236), (95, 73, 122)),  # Light purple, dark purple text
    "🌍 NGHỆ THUẬT VĂN HÓA / VÙNG MIỀN": ((235, 241, 222), (118, 147, 60)), # Light green, solid green text
    "👕 THIẾT KẾ POD / ÁO THUN": ((253, 233, 217), (226, 107, 10)), # Light orange, dark orange text
    "⚠️ CÁC TRƯỜNG HỢP THÁCH THỨC KỸ THUẬT": ((218, 238, 243), (49, 134, 155)) # Light cyan, dark cyan text
}

def generate_sample_image(filename, category, art_type, sub_style, idx):
    """Tạo ảnh placeholder chất lượng cao chứa thông tin type"""
    W, H = 800, 800
    bg_col, txt_col = CAT_COLORS.get(category, ((240, 240, 240), (50, 50, 50)))
    
    img = Image.new('RGB', (W, H), color=bg_col)
    draw = ImageDraw.Draw(img)
    
    # Kẻ khung viền
    border_w = 20
    draw.rectangle([border_w, border_w, W-border_w, H-border_w], outline=txt_col, width=5)
    
    # Cố gắng load font chuẩn, nếu không có xài default
    title_font = ImageFont.load_default()
    sub_font = ImageFont.load_default()
    try:
        # Trên Windows thường có arial.ttf
        title_font = ImageFont.truetype("arialbd.ttf", 48)
        sub_font = ImageFont.truetype("arial.ttf", 32)
        small_font = ImageFont.truetype("ariali.ttf", 24)
        large_font = ImageFont.truetype("arialbd.ttf", 64)
    except:
        small_font = title_font
        large_font = title_font
    
    # Căn giữa text cơ bản bằng cách ước lượng (tránh lỗi getbox trên Python cũ)
    center_y = H // 2
    
    draw.text((W//2, 100), f"SAMPLE #{idx}", fill=txt_col, font=large_font, anchor="ms")
    draw.text((W//2, center_y - 60), art_type.upper(), fill=txt_col, font=title_font, anchor="ms")
    draw.text((W//2, center_y + 20), sub_style, fill=txt_col, font=sub_font, anchor="ms")
    
    # Print technique info
    cat_short = category.split(" ", 1)[-1] if " " in category else category
    draw.text((W//2, H - 120), f"Category: {cat_short}", fill=txt_col, font=small_font, anchor="ms")
    draw.text((W//2, H - 70), "For Print-On-Demand QA Testing", fill=(100,100,100), font=small_font, anchor="ms")
    
    filepath = os.path.join(IMG_DIR, filename)
    img.save(filepath, quality=95)
    return filepath

def main():
    if not os.path.exists(EXCEL_FILE):
        print(f"❌ Không tìm thấy file Excel: {EXCEL_FILE}")
        print("Hãy chạy add_artwork_testdata_vi.py trước.")
        return
        
    print(f"🔄 Đang tạo ảnh sample vào thư mục: {IMG_DIR}...")
    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "DỮ LIỆU TEST ARTWORK" not in wb.sheetnames:
        print("❌ Không tìm thấy sheet 'DỮ LIỆU TEST ARTWORK'")
        return
        
    ws = wb["DỮ LIỆU TEST ARTWORK"]
    
    # Tìm cột cuối cùng để thêm link tải ảnh
    max_col = ws.max_column
    link_col_idx = max_col + 1
    
    # Lấy title row (row 1)
    from copy import copy
    ws.cell(1, link_col_idx, "Link Ảnh Sample").fill = copy(ws.cell(1, max_col).fill)
    ws.cell(1, link_col_idx).font = copy(ws.cell(1, max_col).font)
    ws.cell(1, link_col_idx).border = copy(ws.cell(1, max_col).border)
    ws.column_dimensions[get_column_letter(link_col_idx)].width = 30
    
    row = 2
    idx = 1
    generated_count = 0
    
    # Lặp qua data (cần mapping qua structure của Excel sheet)
    for row_idx in range(2, ws.max_row + 1):
        idx_val = ws.cell(row_idx, 1).value
        # Nếu dòng đó là danh mục (không có số STT)
        if not isinstance(idx_val, int):
            ws.cell(row_idx, link_col_idx).fill = copy(ws.cell(row_idx, 1).fill)
            ws.cell(row_idx, link_col_idx).border = copy(ws.cell(row_idx, max_col).border)
            continue
            
        cat_name = ws.cell(row_idx, 2).value
        art_type = ws.cell(row_idx, 3).value
        sub_style = ws.cell(row_idx, 4).value
        
        # Determine full category name from dictionary keys
        full_cat = ""
        for c in CAT_COLORS.keys():
            if cat_name in c:
                full_cat = c
                break
                
        # Generate safe filename
        safe_type = "".join(c if c.isalnum() else "_" for c in art_type).strip("_")
        safe_sub = "".join(c if c.isalnum() else "_" for c in sub_style).strip("_")
        filename = f"TC_{idx:03d}_{safe_type}_{safe_sub}.jpg"
        
        # Create image
        filepath = generate_sample_image(filename, full_cat, art_type, sub_style, idx)
        generated_count += 1
        
        # Cập nhật excel: tạo hyperlink local
        rel_path = f"Artwork_Samples/{filename}"
        c = ws.cell(row_idx, link_col_idx, "📁 Mở ảnh mẫu")
        c.hyperlink = rel_path
        c.font = Font(name="Calibri", size=11, color="0000FF", underline="single")
        c.alignment = Alignment(horizontal='center', vertical='top')
        c.border = copy(ws.cell(row_idx, max_col).border)
        
        idx += 1
        
    wb.save(EXCEL_FILE)
    print(f"\n✅ Đã tạo thành công {generated_count} ảnh sample tại: {IMG_DIR}")
    print(f"✅ Đã cập nhật file Excel (thêm cột 'Link Ảnh Sample')")
    
if __name__ == "__main__":
    main()
