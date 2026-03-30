"""
Rebuild HOME PAGE and DESIGN STUDIO sheets with proper 5-block categories.
Categories per skill 13.2: UI/UX, Functional, Validation, Security, SEO & Accessibility
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os, re, datetime, shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
MD_FILE = os.path.join(BASE_DIR, "test_cases_suite.md")
today = datetime.datetime.now().strftime("%Y-%m-%d")

SOURCE = None
for f in sorted(os.listdir(BASE_DIR), reverse=True):
    if f.startswith("TC_POD") and f.endswith(".xlsx") and "v23" in f:
        SOURCE = os.path.join(BASE_DIR, f)
        break

OUTPUT = os.path.join(BASE_DIR, f"TC_POD-TShirt-Platform_ExecutionSummary_v25_{today}.xlsx")
VER = 25

# ─── Styles ───
FF = "Calibri"; SZ = 11
fb = Font(name=FF, size=SZ)
fh = Font(name=FF, size=11, bold=True, color="FFFFFF")
hf = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
hf2 = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
bt = Side(border_style="thin", color="BFBFBF")
bdr = Border(left=bt, right=bt, top=bt, bottom=bt)
wa = Alignment(wrap_text=True, vertical="top")
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
fcat = Font(name=FF, size=12, bold=True, color="1F4E78")
cf = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
fp0 = Font(name=FF, size=SZ, color="C00000", bold=True)
fp1 = Font(name=FF, size=SZ, color="ED7D31", bold=True)
fp2 = Font(name=FF, size=SZ, color="70AD47", bold=True)
fui = Font(name=FF, size=SZ, color="7030A0", bold=True)
fneg = Font(name=FF, size=SZ, color="C00000")
fpos = Font(name=FF, size=SZ, color="0070C0")

HEADERS = [
    "TC_ID","US_Mapping","Feature","Module","Title","Type","Priority",
    "Precondition","Test_Data","Steps","Expected_Result",
    "Action Type\n\u30a2\u30af\u30b7\u30e7\u30f3","Create TCs Type",
    "Execution Type\n\u5b9f\u884c\u30bf\u30a4\u30d7",
    "Result\n\u7d50\u679c (R1)","Test date\n(R1)","Tester\n(R1)","ID Bug\n(R1)",
    "Result\n\u7d50\u679c (R2)","Test date\n(R2)","Tester\n(R2)","ID Bug\n(R2)",
    "Evidence\n\u8a3c\u62e0","Notes\n\u5099\u8003",
    "Review_Manual\n(Feedback)"
]

# Review_Manual column style
frv = Font(name=FF, size=11, bold=True, color="000000")
hf_review = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Yellow/Gold

COL_WIDTHS = {
    1:15, 2:12, 3:14, 4:14, 5:40, 6:12, 7:10,
    8:22, 9:15, 10:50, 11:50, 12:12, 13:10, 14:12,
    15:10, 16:12, 17:10, 18:10, 19:10, 20:12, 21:10, 22:10, 23:15, 24:20,
    25:35  # Review_Manual - wide for feedback text
}

def pfont(p):
    if "P0" in str(p): return fp0
    if "P1" in str(p): return fp1
    return fp2

def tfont(t):
    if "UI/UX" in str(t): return fui
    if "Negative" in str(t): return fneg
    return fpos

def clean(s):
    s = s.strip()
    s = s.replace('`','')
    s = re.sub(r'\*\*(.*?)\*\*', r'\1', s)
    s = s.replace('<br>','\n')
    return s

def extract_priority(s):
    m = re.search(r'(P[0-2])', s)
    return m.group(1) if m else "P1"

def extract_type_label(s):
    if "Negative" in s: return "Negative"
    if "UI/UX" in s: return "UI/UX"
    if "Boundary" in s: return "Boundary"
    if "Edge" in s: return "Edge Case"
    return "Positive"

def parse_md_feature(lines, feature_name):
    """Parse a feature section from test_cases_suite.md. Returns list of (category, row_dict)."""
    results = []
    current_category = "Uncategorized"
    in_table = False
    
    for raw_line in lines:
        line = raw_line.rstrip('\r\n').strip()
        
        # Detect category header: ### 📌 SomeName
        if "###" in line and "\U0001f4cc" in line:
            current_category = line.split("\U0001f4cc")[-1].strip()
            in_table = False
            print(f"      Category found: '{current_category}'")
            continue
        
        # Table separator (skip)
        if line.startswith("|:---") or line.startswith("| :---"):
            continue
        
        # Table header row (skip but mark we're in a table)
        if line.startswith("| TC_ID"):
            in_table = True
            continue
        
        # Data row
        if in_table and line.startswith("|") and "TC_" in line:
            cols = line.split("|")
            cols = [c.strip() for c in cols]
            cols = [c for c in cols if c]
            
            if len(cols) >= 8:
                tc_id = clean(cols[0])
                us_map = clean(cols[1])
                module = clean(cols[2])
                title = clean(cols[3])
                tc_type = extract_type_label(cols[4])
                priority = extract_priority(cols[5])
                steps = clean(cols[6])
                expected = clean(cols[7])
                
                # Check for precondition in steps
                precondition = ""
                if "Precondition:" in steps:
                    parts = steps.split("\n", 1)
                    precondition = parts[0].replace("Precondition:","").strip()
                    steps = parts[1] if len(parts) > 1 else steps
                
                results.append({
                    "category": current_category,
                    "tc_id": tc_id,
                    "us_map": us_map,
                    "feature": feature_name,
                    "module": module,
                    "title": title,
                    "type": tc_type,
                    "priority": priority,
                    "precondition": precondition,
                    "test_data": "",
                    "steps": steps,
                    "expected": expected,
                })
        
        # Non-table, non-header line after table started = table ended
        if in_table and not line.startswith("|") and line and not line.startswith("###") and not line.startswith("##"):
            in_table = False
    
    return results

def create_sheet(wb, sheet_name, feature_name, rows):
    """Create formatted feature sheet with proper category blocks."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    # Row 1: Feature title bar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    c1 = ws.cell(row=1, column=1, value=f"\U0001f680 Feature: {feature_name}")
    c1.font = Font(name=FF, size=14, bold=True, color="FFFFFF")
    c1.fill = hf; c1.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, len(HEADERS)+1):
        ws.cell(row=1, column=c).fill = hf

    # Row 2: Column headers
    for ci, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=2, column=ci, value=h)
        cell.border = bdr; cell.alignment = ca
        if "Review_Manual" in h:
            cell.font = frv; cell.fill = hf_review  # Yellow for Review_Manual
        else:
            cell.font = fh; cell.fill = hf2

    # Column widths
    for ci, w in COL_WIDTHS.items():
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Write rows grouped by category
    current_cat = None
    r = 3
    cat_set = set()

    for row_data in rows:
        cat = row_data["category"]
        if cat != current_cat:
            current_cat = cat
            cat_set.add(cat)
            # Category header row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(HEADERS))
            cell = ws.cell(row=r, column=1, value=f"\U0001f4cc {cat}")
            cell.font = fcat; cell.fill = cf
            cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            for c in range(1, len(HEADERS)+1):
                ws.cell(row=r, column=c).border = bdr
                ws.cell(row=r, column=c).fill = cf
            r += 1

        # Data row
        exec_type = "Manual" if row_data["type"] == "UI/UX" else "Auto"
        values = [
            row_data["tc_id"], row_data["us_map"], row_data["feature"],
            row_data["module"], row_data["title"], row_data["type"],
            row_data["priority"], row_data["precondition"], row_data["test_data"],
            row_data["steps"], row_data["expected"],
            "Add new", "By AI", exec_type,
            "Untested", "", "", "",
            "Untested", "", "", "",
            "", "",
            ""  # Review_Manual - empty, user fills in
        ]

        for ci, val in enumerate(values, 1):
            cell = ws.cell(row=r, column=ci, value=val)
            cell.border = bdr; cell.font = fb
            h_name = HEADERS[ci-1]
            if any(k in h_name for k in ["Steps","Expected","Precondition","Title","Test_Data","Notes","Evidence"]):
                cell.alignment = wa
            else:
                cell.alignment = ca
            if "Priority" in h_name: cell.font = pfont(row_data["priority"])
            if h_name == "Type": cell.font = tfont(row_data["type"])
        r += 1

    # Data Validations
    max_r = r - 1
    for dv_range, opts in [
        (f"O3:O{max_r}", "Untested,Pass,Fail,N/A"),
        (f"S3:S{max_r}", "Untested,Pass,Fail,N/A"),
        (f"L3:L{max_r}", "Add new,Update,Delete"),
        (f"M3:M{max_r}", "By AI,By Manual"),
        (f"N3:N{max_r}", "Auto,Manual"),
    ]:
        dv = DataValidation(type="list", formula1=f'"{opts}"', allow_blank=True)
        dv.add(dv_range)
        ws.add_data_validation(dv)

    ws.freeze_panes = "A3"
    print(f"   \u2705 {sheet_name}: {len(rows)} TCs | {len(cat_set)} categories: {list(cat_set)}")
    return len(rows)


def main():
    if not SOURCE:
        print("\u274c No v23 source Excel found!"); return
    print(f"\U0001f4d6 Source: {os.path.basename(SOURCE)}")

    # Delete old v24 outputs if needed
    for f in os.listdir(BASE_DIR):
        if "v24" in f and f.endswith(".xlsx") and f != os.path.basename(OUTPUT):
            try:
                os.remove(os.path.join(BASE_DIR, f))
                print(f"   Removed old: {f}")
            except: pass

    shutil.copy2(SOURCE, OUTPUT)
    wb = openpyxl.load_workbook(OUTPUT)

    # Read markdown
    with open(MD_FILE, 'r', encoding='utf-8') as f:
        content = f.read()

    # Split sections
    home_start = content.find("## \U0001f680 Feature: HOME PAGE")
    ds_start = content.find("## \U0001f680 Feature: DESIGN STUDIO")
    if home_start < 0 or ds_start < 0:
        print("\u274c Cannot find feature sections!"); return

    home_lines = content[home_start:ds_start].split('\n')
    ds_lines = content[ds_start:].split('\n')

    print("\n\U0001f50d Parsing HOME PAGE...")
    home_rows = parse_md_feature(home_lines, "HOME PAGE")
    print(f"   Total: {len(home_rows)} TCs")

    print("\n\U0001f50d Parsing DESIGN STUDIO...")
    ds_rows = parse_md_feature(ds_lines, "DESIGN STUDIO")
    print(f"   Total: {len(ds_rows)} TCs")

    print("\n\U0001f3d7\ufe0f  Building sheets...")
    h_count = create_sheet(wb, "HOME PAGE", "HOME PAGE", home_rows)
    d_count = create_sheet(wb, "DESIGN STUDIO", "DESIGN STUDIO", ds_rows)

    # Move sheets to end
    for sn in ["HOME PAGE", "DESIGN STUDIO"]:
        if sn in wb.sheetnames:
            sheets = wb.sheetnames
            wb.move_sheet(sn, offset=len(sheets)-sheets.index(sn)-1)

    # Update Cover Page
    if "Cover Page" in wb.sheetnames:
        cover = wb["Cover Page"]
        for row in cover.iter_rows(min_row=1, max_row=20, min_col=2, max_col=5):
            for cell in row:
                if cell.value and "DOCUMENT VERSION" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = f"v{VER}.0"
                if cell.value and "GENERATED DATE" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Update Change History
    if "Change History" in wb.sheetnames:
        hst = wb["Change History"]
        r = hst.max_row + 1
        hist_data = [f"v{VER}.0", today, 
            f"Full rebuild with categories: HOME({h_count} TCs), DS({d_count} TCs). +17 click logic TCs. Proper 📌 category blocks.", 
            "QA Team"]
        for ci, val in enumerate(hist_data, 2):
            c = hst.cell(row=r, column=ci, value=val)
            c.border = bdr; c.font = fb

    wb.save(OUTPUT)
    print(f"\n\U0001f389 v{VER} saved: {os.path.basename(OUTPUT)}")
    print(f"   HOME PAGE: {h_count} TCs | DESIGN STUDIO: {d_count} TCs | Total: {h_count + d_count}")

if __name__ == "__main__":
    main()
