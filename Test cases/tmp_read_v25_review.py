"""Read Review_Manual (Feedback) column from v25 Excel - write to JSON for reliable output."""
import openpyxl, os, json, sys

sys.stdout.reconfigure(encoding='utf-8')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
f = os.path.join(BASE_DIR, 'TC_POD-TShirt-Platform_ExecutionSummary_v25_2026-03-20.xlsx')
wb = openpyxl.load_workbook(f, read_only=True, data_only=True)

all_feedback = {}

for sn in wb.sheetnames:
    ws = wb[sn]
    
    review_col = None
    header_row = None
    for i, row in enumerate(ws.iter_rows(max_row=5, values_only=False)):
        for cell in row:
            val = str(cell.value or '')
            if 'Review_Manual' in val or ('Review' in val and 'Feedback' in val):
                review_col = cell.column
                header_row = cell.row
                break
        if review_col:
            break
    
    if not review_col:
        continue
    
    sheet_feedback = []
    for row in ws.iter_rows(min_row=header_row+1, max_row=ws.max_row):
        tc_id_cell = row[0]
        review_cell = None
        for cell in row:
            if cell.column == review_col:
                review_cell = cell
                break
        
        tc_id = str(tc_id_cell.value or '').strip()
        review_val = str(review_cell.value or '').strip() if review_cell else ''
        
        if review_val and review_val != 'None' and review_val != '':
            title = ''
            steps = ''
            expected = ''
            module = ''
            feature = ''
            tc_type = ''
            priority = ''
            for cell in row:
                if cell.column == 3: feature = str(cell.value or '')
                if cell.column == 4: module = str(cell.value or '')
                if cell.column == 5: title = str(cell.value or '')
                if cell.column == 6: tc_type = str(cell.value or '')
                if cell.column == 7: priority = str(cell.value or '')
                if cell.column == 10: steps = str(cell.value or '')
                if cell.column == 11: expected = str(cell.value or '')
            
            sheet_feedback.append({
                'tc_id': tc_id,
                'title': title,
                'module': module,
                'feature': feature,
                'type': tc_type,
                'priority': priority,
                'steps': steps,
                'expected': expected,
                'review': review_val
            })
    
    if sheet_feedback:
        all_feedback[sn] = sheet_feedback

# Write to JSON file
out_path = os.path.join(BASE_DIR, 'tmp_v25_feedback.json')
with open(out_path, 'w', encoding='utf-8') as fp:
    json.dump(all_feedback, fp, ensure_ascii=False, indent=2)

# Summary
for sn, items in all_feedback.items():
    print(f'{sn}: {len(items)} feedbacks')
    for item in items:
        status = 'FIX' if '[Fix]' in item['review'] or '[fix]' in item['review'].lower() else 'OK'
        print(f'  {status:4} {item["tc_id"]}: {item["review"][:80]}')

wb.close()
print(f'\nSaved to: {out_path}')
