"""Read Review_Manual feedback from v26_v3 Excel file."""
import openpyxl, os, json

BASE_DIR = r"e:\BII\QA-NEW\Tool\antigravity-tryonic-main\Test cases"
EXCEL = None
for f in sorted(os.listdir(BASE_DIR), reverse=True):
    if f.startswith("TC_POD") and f.endswith(".xlsx") and "v26" in f and "v3" in f and not f.startswith("~$"):
        EXCEL = os.path.join(BASE_DIR, f)
        break

if not EXCEL:
    print("No v26_v3 file found!"); exit()

print(f"Reading: {os.path.basename(EXCEL)}")
wb = openpyxl.load_workbook(EXCEL, data_only=True)

feedback = {}
for sn in wb.sheetnames:
    ws = wb[sn]
    review_col = None
    tcid_col = None
    title_col = None
    steps_col = None
    expected_col = None
    
    for row in ws.iter_rows(min_row=1, max_row=3, max_col=30):
        for cell in row:
            val = str(cell.value or "")
            if "Review_Manual" in val:
                review_col = cell.column
            if val.strip() == "TC_ID":
                tcid_col = cell.column
            if val.strip() == "Title":
                title_col = cell.column
            if val.strip() == "Steps":
                steps_col = cell.column
            if "Expected" in val:
                expected_col = cell.column
    
    if not review_col or not tcid_col:
        continue
    
    sheet_feedback = []
    max_col = max(review_col, tcid_col, title_col or 1, steps_col or 1, expected_col or 1)
    for row in ws.iter_rows(min_row=3, max_col=max_col):
        tc_id = None
        review = None
        title = None
        steps = None
        expected = None
        for cell in row:
            if cell.column == tcid_col:
                tc_id = str(cell.value or "").strip()
            if cell.column == review_col:
                review = str(cell.value or "").strip()
            if title_col and cell.column == title_col:
                title = str(cell.value or "").strip()
            if steps_col and cell.column == steps_col:
                steps = str(cell.value or "").strip()
            if expected_col and cell.column == expected_col:
                expected = str(cell.value or "").strip()
        
        if review and review != "None" and review != "" and tc_id and "TC_" in tc_id:
            sheet_feedback.append({
                "tc_id": tc_id,
                "title": title,
                "steps": steps,
                "expected": expected,
                "review": review
            })
    
    if sheet_feedback:
        feedback[sn] = sheet_feedback
        print(f"  {sn}: {len(sheet_feedback)} feedbacks")

output = os.path.join(BASE_DIR, "tmp_v26v3_feedback.json")
with open(output, 'w', encoding='utf-8') as f:
    json.dump(feedback, f, ensure_ascii=False, indent=2)

print(f"\nTotal: {sum(len(v) for v in feedback.values())} feedbacks across {len(feedback)} sheets")

for sheet, items in feedback.items():
    print(f"\n{'='*60}")
    print(f"Sheet: {sheet}")
    print(f"{'='*60}")
    for item in items:
        print(f"\n  TC: {item['tc_id']}")
        print(f"  Title: {item['title']}")
        print(f"  Review: {item['review']}")
