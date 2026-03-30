"""
Update Excel v21 → v22: Add HOME PAGE + DESIGN STUDIO test cases
Based on export_tc_multisheet.py format and conventions.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os
import copy
import datetime
import shutil

# --- CONFIG ---
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE_FILE = os.path.join(BASE_DIR, "TC_POD-TShirt-Platform_ExecutionSummary_v21_2026-03-16.xlsx")
today_str = datetime.datetime.now().strftime("%Y-%m-%d")
OUTPUT_FILE = os.path.join(BASE_DIR, f"TC_POD-TShirt-Platform_ExecutionSummary_v22_{today_str}.xlsx")
VERSION = 22

# --- STYLES (same as export_tc_multisheet.py) ---
FONT_FAMILY = "Calibri"
STD_SIZE = 11
font_body = Font(name=FONT_FAMILY, size=STD_SIZE)
font_bold = Font(name=FONT_FAMILY, size=STD_SIZE, bold=True)
font_header = Font(name=FONT_FAMILY, size=11, bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
round_head_fill = PatternFill(start_color="3B73B9", end_color="3B73B9", fill_type="solid")
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
summary_head_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
total_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
border_thin = Side(border_style="thin", color="BFBFBF")
full_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Priority fonts
font_p0 = Font(name=FONT_FAMILY, size=STD_SIZE, color="C00000", bold=True)
font_p1 = Font(name=FONT_FAMILY, size=STD_SIZE, color="ED7D31", bold=True)
font_p2 = Font(name=FONT_FAMILY, size=STD_SIZE, color="70AD47", bold=True)
font_uiux = Font(name=FONT_FAMILY, size=STD_SIZE, color="7030A0", bold=True)
font_negative = Font(name=FONT_FAMILY, size=STD_SIZE, color="C00000")
font_positive = Font(name=FONT_FAMILY, size=STD_SIZE, color="0070C0")
font_cat = Font(name=FONT_FAMILY, size=12, bold=True, color="1F4E78")

# Sub headers in exact order
sub_headers = [
    "TC_ID", "US_Mapping", "Feature", "Module", "Title", "Type", "Priority",
    "Precondition", "Test_Data", "Steps", "Expected_Result",
    "Action Type\nアクション", "Create TCs Type", "Execution Type\n実行タイプ",
    "Result\n結果", "Test date\nテスト日", "Tester\nテスター", "ID Bug\nバグID",
    "Result\n結果", "Test date\nテスト日", "Tester\nテスター", "ID Bug\nバグID",
    "Evidence\n証拠", "Notes\n備考"
]

column_widths = {
    "TC_ID": 15, "US_Mapping": 11, "Feature": 12, "Module": 12,
    "Title": 30, "Type": 12, "Priority": 10, "Precondition": 20,
    "Test_Data": 15, "Steps": 45, "Expected_Result": 45,
    "Action Type\nアクション": 15, "Create TCs Type": 15,
    "Execution Type\n実行タイプ": 15, "Result\n結果": 12,
    "Test date\nテスト日": 15, "Tester\nテスター": 15, "ID Bug\nバグID": 15,
    "Evidence\n証拠": 20, "Notes\n備考": 20
}

# ===================================================================
# NEW TEST CASES DATA
# ===================================================================

HOME_PAGE_TCS = {
    "UI/UX": [
        ("TC_HOME_UI_001", "US-HP-01", "HOME PAGE", "Header", "Home page: Logo Tryonic AI hiển thị đúng", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Mở /home/\n3. Quan sát góc trái header", "Logo hiển thị: icon tím gradient + text 'Tryonic AI'. Logo có thể click và trỏ về trang chủ"),
        ("TC_HOME_UI_002", "US-HP-01", "HOME PAGE", "Header", "Home page: Navigation menu hiển thị đầy đủ", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát thanh navigation giữa header", "Hiển thị đầy đủ 4 menu items: 'Trang chủ', 'Sản phẩm', 'Dịch vụ', 'Liên hệ'"),
        ("TC_HOME_UI_003", "US-HP-01", "HOME PAGE", "Header", "Home page: Nút 'Thiết kế ngay' style đúng", "UI/UX", "P0", "", "", "1. Truy cập vào trang \n2. Quan sát nút CTA góc phải header", "Nút 'Thiết kế ngay': background gradient tím→đỏ, text trắng, bo tròn, font semi-bold"),
        ("TC_HOME_UI_004", "US-HP-01", "HOME PAGE", "Header", "Home page: Header sticky khi scroll", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Scroll xuống cuối trang\n3. Quan sát header", "Header cố định (sticky) ở trên cùng khi scroll. Nền trắng, không bị mờ hoặc biến mất"),
        ("TC_HOME_UI_005", "US-HP-02", "HOME PAGE", "Hero", "Home page: Badge 'AI-Powered Design' hiển thị", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Quan sát phần trên cùng Hero section", "Badge hiển thị: '✨ AI-Powered Design' với background tím nhạt, icon sparkle"),
        ("TC_HOME_UI_006", "US-HP-02", "HOME PAGE", "Hero", "Home page: Headline text đúng nội dung", "UI/UX", "P0", "", "", "1. Truy cập vào trang \n2. Quan sát tiêu đề chính Hero section", "Headline: 'Biến ý tưởng thành áo thun trong 30 giây'. Phần 'áo thun trong 30 giây' màu tím. Font ~48-56px, bold"),
        ("TC_HOME_UI_007", "US-HP-02", "HOME PAGE", "Hero", "Home page: Subtitle text đúng nội dung", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Quan sát dòng mô tả bên dưới Headline", "Subtitle: 'Chỉ cần mô tả — AI sẽ thiết kế cho bạn. Chất liệu premium, giao tận nơi.' Text xám, italic"),
        ("TC_HOME_UI_008", "US-HP-03", "HOME PAGE", "AI Input", "Home page: AI Input Box placeholder đúng", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát ô nhập liệu AI ở giữa trang", "Placeholder: 'Mô tả áo thun bạn muốn... VD: Áo minimalist hoa sak...'. Icon ảnh bên trái, bo tròn, shadow nhẹ"),
        ("TC_HOME_UI_009", "US-HP-03", "HOME PAGE", "AI Input", "Home page: Nút Generate style đúng", "UI/UX", "P0", "", "", "1. Truy cập vào trang \n2. Quan sát nút Generate bên phải input box", "Nút: text 'Generate' + icon sparkle ✨. Background tím gradient, text trắng, bo tròn"),
        ("TC_HOME_UI_010", "US-HP-03", "HOME PAGE", "AI Input", "Home page: 6 Style Tags hiển thị đầy đủ", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát các tag phía dưới AI Input Box", "6 tags: 'Minimalist', 'Streetwear', 'Anime', 'Vintage', 'Y2K', 'Abstract Art'. Mỗi tag có icon, bo tròn"),
        ("TC_HOME_UI_011", "US-HP-03", "HOME PAGE", "AI Input", "Home page: Nút 'Chọn từ mẫu có sẵn' hiển thị", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát 2 nút phía dưới Style Tags", "Nút 'Chọn từ mẫu có sẵn': Grid icon tím, subtitle 'Khám phá thư viện mẫu'"),
        ("TC_HOME_UI_012", "US-HP-03", "HOME PAGE", "AI Input", "Home page: Nút 'Tải lên ảnh của bạn' hiển thị", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát nút thứ hai phía dưới Style Tags", "Nút 'Tải lên ảnh của bạn': Upload icon tím, subtitle 'Sử dụng file thiết kế riêng'"),
        ("TC_HOME_UI_013", "US-HP-04", "HOME PAGE", "Footer", "Home page: Trust Markers hiển thị đầy đủ", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Scroll xuống cuối trang\n3. Quan sát footer strip", "3 markers: '✅ Thanh toán an toàn', '✅ Giao hàng toàn quốc', '✅ Đổi trả 7 ngày'"),
        ("TC_HOME_UI_014", "US-HP-01", "HOME PAGE", "Header", "Home page: Hover effect trên menu items", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Hover chuột lần lượt qua menu items", "Mỗi menu item có hiệu ứng hover. Cursor chuyển sang pointer"),
        ("TC_HOME_UI_015", "US-HP-01", "HOME PAGE", "Header", "Home page: Hover effect trên nút 'Thiết kế ngay'", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Hover chuột qua nút 'Thiết kế ngay'", "Nút có hiệu ứng hover (shadow đậm, scale nhẹ). Cursor pointer"),
        ("TC_HOME_UI_016", "US-HP-02", "HOME PAGE", "Hero", "Home page: Background gradient Hero section", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Quan sát nền Hero section", "Nền gradient nhẹ: trắng lavender (#F8F7FF) với hiệu ứng trang trí 2 bên"),
        ("TC_HOME_UI_860", "Global", "HOME PAGE", "Responsive & Zoom", "Màn hình Home: Browser Zoom In/Out (50%-200%)", "UI/UX", "P1", "", "", "1. Truy cập trang Home.\n2. Nhấn Ctrl + [+] để Zoom In 200%.\n3. Nhấn Ctrl + [-] để Zoom Out 50%.", "Text/SVG sắc nét, layout thu phóng đúng tỷ lệ, không che lấp button"),
        ("TC_HOME_UI_861", "Global", "HOME PAGE", "Responsive (iPhone)", "Màn hình Home: Responsive iPhone (Portrait)", "UI/UX", "P0", "", "", "1. Truy cập trang Home trên iPhone Portrait.\n2. Cuộn lên/xuống.\n3. Tap thử các nút.", "Không cuộn ngang thừa. Style Tags wrap xuống hàng mới. Touch Targets đủ lớn"),
        ("TC_HOME_UI_862", "Global", "HOME PAGE", "Responsive (Android)", "Màn hình Home: Responsive Android Phone (Portrait)", "UI/UX", "P0", "", "", "1. Truy cập trang Home trên Android Portrait.\n2. Focus vào input AI.", "Responsive tương tự iPhone. Bàn phím ảo không che nút Generate"),
        ("TC_HOME_UI_863", "Global", "HOME PAGE", "Responsive (iPad)", "Màn hình Home: Responsive iPad (Portrait)", "UI/UX", "P1", "", "", "1. Truy cập trang Home trên iPad Portrait.\n2. Quan sát bố cục.", "AI Input Box không trải rộng quá mức. Style Tags trên 1 hàng"),
        ("TC_HOME_UI_864", "Global", "HOME PAGE", "Responsive (Android Tablet)", "Màn hình Home: Responsive Android Tablet (Portrait)", "UI/UX", "P1", "", "", "1. Truy cập trang Home trên Tablet Android Portrait.", "Tương tự iPad, icon/text scale đúng tỉ lệ"),
        ("TC_HOME_UI_865", "Global", "HOME PAGE", "Responsive (Landscape)", "Màn hình Home: Responsive Landscape", "UI/UX", "P0", "", "", "1. Xoay ngang thiết bị.\n2. Quan sát chuyển đổi giao diện.", "UI sắp xếp lại phù hợp. Header navigation không bị tràn"),
    ],
    "Functional (Logic & Behavior)": [
        ("TC_HOME_001", "US-HP-01", "HOME PAGE", "Navigation", "Click nút 'Thiết kế ngay' mở Design Studio", "Positive", "P0", "", "", "1. Truy cập vào trang \n2. Mở /home/\n3. Click nút 'Thiết kế ngay'", "Chuyển sang Design Studio. Canvas mockup hiển thị"),
        ("TC_HOME_002", "US-HP-01", "HOME PAGE", "Navigation", "Click logo Tryonic AI trỏ về trang chủ", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click logo 'Tryonic AI'", "Redirect về /home/. Trang chủ load đầy đủ"),
        ("TC_HOME_003", "US-HP-01", "HOME PAGE", "Navigation", "Click menu 'Trang chủ'", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click 'Trang chủ' trên menu", "URL trỏ đến /home/. Trang load thành công"),
        ("TC_HOME_004", "US-HP-03", "HOME PAGE", "AI Input", "Nhập prompt và click Generate", "Positive", "P0", "", "", "1. Truy cập vào trang \n2. Nhập: 'Áo minimalist hoa sakura'\n3. Click 'Generate'", "Chuyển sang Design Studio hoặc hiển thị kết quả AI. Loading state hiển thị"),
        ("TC_HOME_005", "US-HP-03", "HOME PAGE", "AI Input", "Click Style Tag thay đổi trạng thái", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click tag 'Minimalist'\n3. Click tag 'Anime'", "Tag được chọn có highlight. Chỉ 1 tag active hoặc multi-select"),
        ("TC_HOME_006", "US-HP-03", "HOME PAGE", "AI Input", "Click 'Chọn từ mẫu có sẵn'", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click nút 'Chọn từ mẫu có sẵn'", "Chuyển sang Gallery/Templates hoặc mở modal thư viện mẫu"),
        ("TC_HOME_007", "US-HP-03", "HOME PAGE", "AI Input", "Click 'Tải lên ảnh của bạn'", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click nút 'Tải lên ảnh của bạn'", "Mở file picker hoặc chuyển sang Design Studio upload"),
        ("TC_HOME_008", "US-HP-03", "HOME PAGE", "AI Input", "Submit bằng phím Enter", "Positive", "P2", "", "", "1. Truy cập vào trang \n2. Nhập: 'Áo vintage rock band'\n3. Nhấn Enter", "Form submit tương đương click Generate"),
    ],
    "Validation": [
        ("TC_HOME_009", "US-HP-03", "HOME PAGE", "AI Input", "Click Generate khi input rỗng", "Negative", "P0", "", "", "1. Truy cập vào trang \n2. Không nhập gì\n3. Click 'Generate'", "Validation: 'Vui lòng nhập mô tả' hoặc nút disabled"),
        ("TC_HOME_010", "US-HP-01", "HOME PAGE", "Navigation", "Menu 'Sản phẩm' trỏ đúng route", "Negative", "P1", "", "", "1. Truy cập vào trang \n2. Click menu 'Sản phẩm'\n3. Quan sát URL", "URL → /products. Không trỏ về /home/ (BUG hiện tại)"),
        ("TC_HOME_011", "US-HP-01", "HOME PAGE", "Navigation", "Menu 'Dịch vụ' trỏ đúng route", "Negative", "P1", "", "", "1. Truy cập vào trang \n2. Click menu 'Dịch vụ'\n3. Quan sát URL", "URL → /services. Không trỏ về /home/ (BUG hiện tại)"),
        ("TC_HOME_012", "US-HP-01", "HOME PAGE", "Navigation", "Menu 'Liên hệ' trỏ đúng route", "Negative", "P1", "", "", "1. Truy cập vào trang \n2. Click menu 'Liên hệ'\n3. Quan sát URL", "URL → /contact. Không trỏ về /home/ (BUG hiện tại)"),
        ("TC_HOME_013", "US-HP-03", "HOME PAGE", "AI Input", "Nhập XSS vào prompt", "Negative", "P1", "", "", "1. Truy cập vào trang \n2. Nhập '<script>alert(1)</script>'\n3. Click Generate", "Sanitize input. Không execute script"),
        ("TC_HOME_014", "US-HP-05", "HOME PAGE", "SEO", "Title tag hiển thị đúng thương hiệu", "Negative", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát tab title", "Title = 'Tryonic AI' không phải 'POD Admin CMS' (BUG)"),
        ("TC_HOME_015", "US-HP-05", "HOME PAGE", "SEO", "Meta description tối ưu", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Kiểm tra source code <meta>", "Meta description phù hợp, không rỗng"),
        ("TC_HOME_016", "US-HP-05", "HOME PAGE", "Accessibility", "Heading hierarchy đúng chuẩn", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Kiểm tra h1, h2, h3", "Chỉ 1 h1. Heading đúng thứ tự"),
        ("TC_HOME_017", "US-HP-05", "HOME PAGE", "Accessibility", "Keyboard navigation (Tab order)", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Nhấn Tab liên tục", "Tab order hợp lý. Focus ring hiển thị rõ"),
    ],
}

DESIGN_STUDIO_TCS = {
    "UI/UX": [
        ("TC_DS_UI_001", "US-DS-01", "DESIGN STUDIO", "Header", "Design Studio: Header hiển thị đầy đủ", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Mở Design Studio\n3. Quan sát header", "Header: ← 'Quay lại', Logo 'Tryonic', 'Design Studio', Credits + User + Cart icons"),
        ("TC_DS_UI_002", "US-DS-01", "DESIGN STUDIO", "Header", "Design Studio: Credits badge hiển thị", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Mở Design Studio\n3. Quan sát góc phải", "'12 Credits' + nút 'Nạp'. Icon coin/circle"),
        ("TC_DS_UI_003", "US-DS-02", "DESIGN STUDIO", "Sidebar", "Design Studio: Sidebar trái đầy đủ công cụ", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Mở Design Studio\n3. Quan sát sidebar trái", "6 công cụ: Hoàn Tác, Làm Lại, Mặt Sau, Thu Phóng, Thử Đồ AI, Chia Sẻ"),
        ("TC_DS_UI_004", "US-DS-03", "DESIGN STUDIO", "Canvas", "Design Studio: Canvas mockup mặc định", "UI/UX", "P0", "", "", "1. Truy cập vào trang \n2. Mở Design Studio\n3. Quan sát canvas", "Mockup áo trắng trên nền xám. Vùng thiết kế nét đứt cyan. Hướng dẫn text hiển thị"),
        ("TC_DS_UI_005", "US-DS-04", "DESIGN STUDIO", "Panel", "Design Studio: 5 tabs panel phải", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Mở Design Studio\n3. Quan sát panel phải", "5 tabs: SẢN PHẨM, ẢNH CỦA BẠN, THƯ VIỆN, TẠO ẢNH AI, ĐẶT HÀNG. 'TẠO ẢNH AI' active"),
        ("TC_DS_UI_006", "US-DS-04", "DESIGN STUDIO", "Panel", "Design Studio: Tab TẠO ẢNH AI nội dung", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Click tab 'TẠO ẢNH AI'", "MÔ TẢ ARTWORK textarea, ẢNH THAM KHẢO 2 nút, PHONG CÁCH 6 cards, '3 Credits', nút Tạo Artwork"),
        ("TC_DS_UI_007", "US-DS-04", "DESIGN STUDIO", "Panel", "Design Studio: 6 phong cách AI", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát phần PHONG CÁCH", "6 cards: Watercolor, Minimalist, Line Art, Retro, Grunge, Flat Design. Mỗi card có ảnh"),
        ("TC_DS_UI_008", "US-DS-05", "DESIGN STUDIO", "Bottom Bar", "Design Studio: Bottom bar thông tin SP", "UI/UX", "P1", "", "", "1. Truy cập vào trang \n2. Quan sát thanh đáy", "'Áo Thun Cotton Gildan 5000', Trắng, Size L, 150.000đ, 'Giá chưa gồm phí in', nút 'Đặt hàng'"),
        ("TC_DS_UI_009", "US-DS-04", "DESIGN STUDIO", "Panel", "Design Studio: MÔ TẢ ARTWORK placeholder", "UI/UX", "P2", "", "", "1. Truy cập vào trang \n2. Quan sát textarea", "Placeholder: 'VD: Rồng Việt Nam phong cách watercolor, màu xanh lam và vàng...'"),
    ],
    "Functional (Logic & Behavior)": [
        ("TC_DS_001", "US-DS-01", "DESIGN STUDIO", "Header", "Click 'Quay lại' về trang chủ", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Mở Design Studio\n3. Click '← Quay lại'", "Chuyển về /home/. Design Studio đóng"),
        ("TC_DS_002", "US-DS-02", "DESIGN STUDIO", "Sidebar", "Click 'Mặt Sau' chuyển view", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click 'Mặt Sau'", "Canvas hiển thị mặt sau áo thun"),
        ("TC_DS_003", "US-DS-02", "DESIGN STUDIO", "Sidebar", "Click 'Thử Đồ với AI'", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click 'Thử Đồ với AI'", "Mở tính năng AI Try-on. Form upload hoặc modal thử đồ"),
        ("TC_DS_004", "US-DS-02", "DESIGN STUDIO", "Sidebar", "Click 'Chia Sẻ'", "Positive", "P2", "", "", "1. Truy cập vào trang \n2. Click 'Chia Sẻ'", "Mở popup share hoặc copy link"),
        ("TC_DS_005", "US-DS-04", "DESIGN STUDIO", "Panel", "Switch giữa các tabs", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click lần lượt SẢN PHẨM → THƯ VIỆN → TẠO ẢNH AI → ĐẶT HÀNG", "Mỗi tab hiển thị nội dung tương ứng. Tab active underline/highlight"),
        ("TC_DS_006", "US-DS-04", "DESIGN STUDIO", "Panel", "Click style card chọn phong cách", "Positive", "P1", "", "", "1. Truy cập vào trang \n2. Click card 'Watercolor'\n3. Click card 'Retro'", "Card chọn highlight. Card trước bỏ highlight"),
        ("TC_DS_007", "US-DS-05", "DESIGN STUDIO", "Bottom Bar", "Click nút 'Đặt hàng'", "Positive", "P0", "", "", "1. Truy cập vào trang \n2. Click nút 'Đặt hàng' (hồng)", "Chuyển sang checkout. Thông tin SP được truyền đúng"),
        ("TC_DS_008", "US-DS-01", "DESIGN STUDIO", "Header", "Click icon User", "Positive", "P2", "", "", "1. Truy cập vào trang \n2. Click icon User", "Mở profile/account hoặc popup login"),
        ("TC_DS_009", "US-DS-01", "DESIGN STUDIO", "Header", "Click icon Giỏ hàng", "Positive", "P2", "", "", "1. Truy cập vào trang \n2. Click icon Giỏ hàng", "Mở trang giỏ hàng hoặc sidebar cart"),
        ("TC_DS_010", "US-DS-01", "DESIGN STUDIO", "Header", "Click nút 'Nạp' credits", "Positive", "P2", "", "", "1. Truy cập vào trang \n2. Click nút 'Nạp'", "Mở trang/modal nạp credits"),
    ],
    "Validation": [
        ("TC_DS_011", "US-DS-04", "DESIGN STUDIO", "Panel", "Tạo Artwork khi MÔ TẢ rỗng", "Negative", "P0", "", "", "1. Truy cập vào trang \n2. Không nhập MÔ TẢ\n3. Click 'Tạo Artwork Mới'", "Validation: 'Vui lòng nhập mô tả'. Không trừ credits"),
        ("TC_DS_012", "US-DS-04", "DESIGN STUDIO", "Panel", "Tạo Artwork khi hết credits", "Negative", "P0", "", "", "1. Truy cập vào trang \n2. Tài khoản 0 credits\n3. Click 'Tạo Artwork Mới'", "'Hết credits'. Gợi ý nạp thêm"),
    ],
}

categories_order = ['UI/UX', 'Validation', 'Functional (Logic & Behavior)', 'Security', 'Performance']

def get_priority_font(priority):
    if priority == "P0": return font_p0
    elif priority == "P1": return font_p1
    elif priority == "P2": return font_p2
    return font_body

def get_type_font(tc_type):
    if tc_type == "UI/UX": return font_uiux
    elif tc_type == "Negative": return font_negative
    elif tc_type == "Positive": return font_positive
    return font_body

def create_feature_sheet(wb, feature_name, tc_data):
    """Create a new feature sheet with the standard format."""
    sheet_name = feature_name[:31]
    ws = wb.create_sheet(title=sheet_name)

    dv_result = DataValidation(type="list", formula1='"Untested,Pass,Fail,N/A"', allow_blank=False)
    dv_action = DataValidation(type="list", formula1='"Add new,Update,Delete"', allow_blank=False)
    dv_auto = DataValidation(type="list", formula1='"By AI,By Manual"', allow_blank=False)
    ws.add_data_validation(dv_result)
    ws.add_data_validation(dv_action)
    ws.add_data_validation(dv_auto)

    # Row 1: Round grouping
    ws.append([""] * 14 + ["Round 1 第1戦", "", "", "", "Round 2 第2戦", "", "", "", "", ""])
    ws.merge_cells(start_row=1, start_column=15, end_row=1, end_column=18)
    ws.merge_cells(start_row=1, start_column=19, end_row=1, end_column=22)
    for col in [15, 19]:
        c = ws.cell(row=1, column=col)
        c.fill = round_head_fill
        c.font = font_header
        c.alignment = center_align
        c.border = full_border
        for offset in range(1, 4):
            ws.cell(row=1, column=col+offset).border = full_border

    # Row 2: Headers
    ws.append(sub_headers)
    ws.row_dimensions[2].height = 40
    for col_idx in range(1, len(sub_headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = font_header
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = full_border

    current_row = 3
    total = 0

    for cat in categories_order:
        if cat not in tc_data or not tc_data[cat]:
            continue

        # Category header row
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=len(sub_headers))
        cell = ws.cell(row=current_row, column=1)
        cell.value = f"📌 {cat}"
        cell.font = font_cat
        cell.fill = category_fill
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for col_idx in range(1, len(sub_headers) + 1):
            ws.cell(row=current_row, column=col_idx).border = full_border
        current_row += 1

        for tc in tc_data[cat]:
            tc_id, us_map, feature, module, title, tc_type, priority, precond, test_data, steps, expected = tc
            exec_type = "Manual" if "UI/UX" in cat else "Auto"
            row_data = [tc_id, us_map, feature, module, title, tc_type, priority, precond, test_data, steps, expected,
                        "Add new", "By AI", exec_type, "Untested", "", "", "", "Untested", "", "", "", "", ""]
            ws.append(row_data)

            dv_action.add(ws.cell(row=current_row, column=12))
            dv_auto.add(ws.cell(row=current_row, column=13))
            dv_auto.add(ws.cell(row=current_row, column=14))
            dv_result.add(ws.cell(row=current_row, column=15))
            dv_result.add(ws.cell(row=current_row, column=19))

            for col_idx in range(1, len(sub_headers) + 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = full_border
                cell.font = font_body
                h = sub_headers[col_idx - 1]
                if h in ["Steps", "Expected_Result", "Precondition", "Title", "Test_Data", "Notes\n備考", "Evidence\n証拠"]:
                    cell.alignment = wrap_align
                else:
                    cell.alignment = center_align
                if h == "Priority":
                    cell.font = get_priority_font(priority)
                if h == "Type":
                    cell.font = get_type_font(tc_type)

            current_row += 1
            total += 1

    # Column widths
    for i, header in enumerate(sub_headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = column_widths.get(header, 15)

    return sheet_name, total


def main():
    if not os.path.exists(SOURCE_FILE):
        print(f"❌ Source file not found: {SOURCE_FILE}")
        return

    print(f"📖 Loading {SOURCE_FILE}...")
    shutil.copy2(SOURCE_FILE, OUTPUT_FILE)
    wb = openpyxl.load_workbook(OUTPUT_FILE)

    print(f"📊 Existing sheets: {wb.sheetnames}")

    # Create new feature sheets
    print("➕ Creating HOME PAGE sheet...")
    hp_sheet, hp_total = create_feature_sheet(wb, "HOME PAGE", HOME_PAGE_TCS)
    print(f"   ✅ {hp_total} test cases")

    print("➕ Creating DESIGN STUDIO sheet...")
    ds_sheet, ds_total = create_feature_sheet(wb, "DESIGN STUDIO", DESIGN_STUDIO_TCS)
    print(f"   ✅ {ds_total} test cases")

    # Update Cover Page version
    if "Cover Page" in wb.sheetnames:
        cover = wb["Cover Page"]
        for row in cover.iter_rows(min_row=1, max_row=20, min_col=2, max_col=5):
            for cell in row:
                if cell.value and "DOCUMENT VERSION" in str(cell.value).upper():
                    val_cell = cover.cell(row=cell.row, column=3)
                    val_cell.value = f"v{VERSION}.0 (Excel formatted)"
                    print(f"   📝 Updated Cover Page version to v{VERSION}.0")
                if cell.value and "MODULE SCOPE" in str(cell.value).upper():
                    val_cell = cover.cell(row=cell.row, column=3)
                    val_cell.value = "Auth, Gallery, Editor, Checkout, CMS, AI, Credits, Home Page, Design Studio"
                if cell.value and "GENERATED DATE" in str(cell.value).upper():
                    val_cell = cover.cell(row=cell.row, column=3)
                    val_cell.value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Update Change History
    if "Change History" in wb.sheetnames:
        hst = wb["Change History"]
        last_row = hst.max_row + 1
        new_entry = [f"v{VERSION}.0", today_str, "Added HOME PAGE + DESIGN STUDIO test cases from UI/UX research", "QA Team"]
        for col_idx, val in enumerate(new_entry, 2):
            c = hst.cell(row=last_row, column=col_idx, value=val)
            c.border = full_border
            c.font = font_body
            c.alignment = Alignment(horizontal="left" if col_idx == 4 else "center", vertical="center", wrap_text=True)
        print(f"   📝 Updated Change History with v{VERSION}.0 entry")

    # Update Execution Summary
    if "Execution Summary" in wb.sheetnames:
        summary = wb["Execution Summary"]
        max_row = summary.max_row

        # Find the last "Round 1" total row to know where to insert
        # We'll append the new features before the Round 1 total row
        r1_total_row = None
        r2_start_row = None
        for row_idx in range(4, max_row + 1):
            cell_val = str(summary.cell(row=row_idx, column=2).value or "")
            if "Round 1" in cell_val:
                r1_total_row = row_idx
            elif "Round 2" in cell_val and r1_total_row:
                r2_start_row = row_idx
                break

        # Instead of complex insertion, add at the bottom with clear labeling
        new_row = max_row + 2
        # Add separator
        summary.cell(row=new_row, column=2, value="── New Features (v22) ──").font = font_bold
        new_row += 1

        new_features = [
            {"Feature": "HOME PAGE", "SheetName": hp_sheet, "Total": hp_total},
            {"Feature": "DESIGN STUDIO", "SheetName": ds_sheet, "Total": ds_total},
        ]

        for data in new_features:
            sheet = data['SheetName']
            total = data['Total']
            f_pass = f"=COUNTIF('{sheet}'!O:O, \"Pass\")"
            f_fail = f"=COUNTIF('{sheet}'!O:O, \"Fail\")"
            f_na = f"=COUNTIF('{sheet}'!O:O, \"N/A\")"
            f_untested = f"=G{new_row} - C{new_row} - D{new_row} - F{new_row}"
            f_prog_test = f"=IF(G{new_row}>0, (C{new_row}+D{new_row})/G{new_row}, 0)"
            f_prog_pass = f"=IF(G{new_row}>0, C{new_row}/G{new_row}, 0)"

            row_data = [f"★ {data['Feature']}", f_pass, f_fail, f_untested, f_na, total, f_prog_test, f_prog_pass]
            for col_idx, val in enumerate(row_data, 2):
                cell = summary.cell(row=new_row, column=col_idx, value=val)
                cell.border = full_border
                cell.font = font_body
                cell.alignment = Alignment(horizontal="left", vertical="center") if col_idx == 2 else center_align
                if col_idx in [8, 9]:
                    cell.number_format = '0%'
            new_row += 1

        # Total row for new features
        tot_data = [
            "v22 Subtotal",
            f"=SUM(C{new_row-2}:C{new_row-1})",
            f"=SUM(D{new_row-2}:D{new_row-1})",
            f"=SUM(E{new_row-2}:E{new_row-1})",
            f"=SUM(F{new_row-2}:F{new_row-1})",
            f"=SUM(G{new_row-2}:G{new_row-1})",
            f"=IF(G{new_row}>0, (C{new_row}+D{new_row})/G{new_row}, 0)",
            f"=IF(G{new_row}>0, C{new_row}/G{new_row}, 0)"
        ]
        for col_idx, val in enumerate(tot_data, 2):
            cell = summary.cell(row=new_row, column=col_idx, value=val)
            cell.border = full_border
            cell.font = font_bold
            cell.fill = total_fill
            cell.alignment = center_align
            if col_idx in [8, 9]:
                cell.number_format = '0%'

        print(f"   📝 Updated Execution Summary with 2 new feature rows")

    wb.save(OUTPUT_FILE)
    print(f"\n🎉 v{VERSION} Excel generated: {OUTPUT_FILE}")
    print(f"   📋 New TCs: HOME PAGE ({hp_total}) + DESIGN STUDIO ({ds_total}) = {hp_total + ds_total} total")


if __name__ == "__main__":
    main()
