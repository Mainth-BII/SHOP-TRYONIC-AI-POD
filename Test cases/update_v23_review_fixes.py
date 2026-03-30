"""
Update Excel v22 → v23: Fix all Agent A review issues
- 5 existing TC text fixes
- 17 new TCs added (responsive DS, Undo/Redo/Zoom, Color/Size, BVA, XSS, State Transition)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os, datetime, shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
SOURCE = os.path.join(BASE_DIR, "TC_POD-TShirt-Platform_ExecutionSummary_v22_2026-03-19.xlsx")
today = datetime.datetime.now().strftime("%Y-%m-%d")
OUTPUT = os.path.join(BASE_DIR, f"TC_POD-TShirt-Platform_ExecutionSummary_v23_{today}.xlsx")
VER = 23

# Styles
FF = "Calibri"; SZ = 11
fb = Font(name=FF, size=SZ)
fbd = Font(name=FF, size=SZ, bold=True)
fh = Font(name=FF, size=11, bold=True, color="FFFFFF")
hf = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
rhf = PatternFill(start_color="3B73B9", end_color="3B73B9", fill_type="solid")
cf = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
tf = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
bt = Side(border_style="thin", color="BFBFBF")
bdr = Border(left=bt, right=bt, top=bt, bottom=bt)
wa = Alignment(wrap_text=True, vertical="top")
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
fcat = Font(name=FF, size=12, bold=True, color="1F4E78")
fp0 = Font(name=FF, size=SZ, color="C00000", bold=True)
fp1 = Font(name=FF, size=SZ, color="ED7D31", bold=True)
fp2 = Font(name=FF, size=SZ, color="70AD47", bold=True)
fui = Font(name=FF, size=SZ, color="7030A0", bold=True)
fneg = Font(name=FF, size=SZ, color="C00000")
fpos = Font(name=FF, size=SZ, color="0070C0")

sub_h = ["TC_ID","US_Mapping","Feature","Module","Title","Type","Priority","Precondition","Test_Data","Steps","Expected_Result",
    "Action Type\n\u30a2\u30af\u30b7\u30e7\u30f3","Create TCs Type","Execution Type\n\u5b9f\u884c\u30bf\u30a4\u30d7",
    "Result\n\u7d50\u679c","Test date\n\u30c6\u30b9\u30c8\u65e5","Tester\n\u30c6\u30b9\u30bf\u30fc","ID Bug\n\u30d0\u30b0ID",
    "Result\n\u7d50\u679c","Test date\n\u30c6\u30b9\u30c8\u65e5","Tester\n\u30c6\u30b9\u30bf\u30fc","ID Bug\n\u30d0\u30b0ID",
    "Evidence\n\u8a3c\u62e0","Notes\n\u5099\u8003"]

cw = {"TC_ID":15,"US_Mapping":11,"Feature":12,"Module":12,"Title":30,"Type":12,"Priority":10,"Precondition":20,
    "Test_Data":15,"Steps":45,"Expected_Result":45}

def pfont(p):
    return fp0 if p=="P0" else fp1 if p=="P1" else fp2 if p=="P2" else fb

def tfont(t):
    return fui if t=="UI/UX" else fneg if t=="Negative" else fpos if t=="Positive" else fb

# ---- FIXES for existing sheets ----
HOME_FIXES = {
    "TC_HOME_005": ("Tag \u0111\u01b0\u1ee3c ch\u1ecdn c\u00f3 highlight (background \u0111\u1ed5i m\u00e0u ho\u1eb7c border \u0111\u1eadm). Ch\u1ec9 1 tag active c\u00f9ng l\u00fac \u2014 tag tr\u01b0\u1edbc \u0111\u00f3 t\u1ef1 \u0111\u1ed9ng b\u1ecf highlight khi ch\u1ecdn tag m\u1edbi"),
    "TC_HOME_006": ("M\u1edf modal th\u01b0 vi\u1ec7n m\u1eabu hi\u1ec3n th\u1ecb danh s\u00e1ch templates c\u00f3 s\u1eb5n. Modal c\u00f3 n\u00fat \u0111\u00f3ng (X) v\u00e0 cho ph\u00e9p ch\u1ecdn template"),
    "TC_HOME_007": ("M\u1edf native file picker (OS dialog) cho ph\u00e9p ch\u1ecdn file \u1ea3nh t\u1eeb m\u00e1y. Ch\u1ea5p nh\u1eadn formats: PNG, JPG, SVG"),
    "TC_HOME_008": ("Form submit th\u00e0nh c\u00f4ng \u2014 chuy\u1ec3n sang Design Studio ho\u1eb7c hi\u1ec3n th\u1ecb loading state. Behavior t\u01b0\u01a1ng \u0111\u01b0\u01a1ng click n\u00fat Generate"),
}

DS_FIXES = {
    "TC_DS_012": {
        "steps": "Precondition: T\u00e0i kho\u1ea3n c\u00f3 0 credits \n1. Truy c\u1eadp v\u00e0o trang \n2. M\u1edf Design Studio\n3. Nh\u1eadp m\u00f4 t\u1ea3 h\u1ee3p l\u1ec7 v\u00e0o textarea\n4. Click '\u2728 T\u1ea1o Artwork M\u1edbi'",
        "expected": "Hi\u1ec3n th\u1ecb: 'H\u1ebft credits'. G\u1ee3i \u00fd n\u1ea1p th\u00eam. Kh\u00f4ng cho ph\u00e9p t\u1ea1o artwork. Kh\u00f4ng tr\u1eeb credits"
    }
}

# ---- NEW TCs ----
NEW_HOME_TCS = [
    ("TC_HOME_018","US-HP-03","HOME PAGE","AI Input","AI prompt ch\u1ec9 c\u00f3 1 k\u00fd t\u1ef1","Negative","P2","","","1. Truy c\u1eadp v\u00e0o trang \n2. Nh\u1eadp: 'A' (1 k\u00fd t\u1ef1)\n3. Click 'Generate'","Hi\u1ec3n th\u1ecb validation: 'M\u00f4 t\u1ea3 qu\u00e1 ng\u1eafn' ho\u1eb7c ch\u1ea5p nh\u1eadn v\u00e0 x\u1eed l\u00fd. Kh\u00f4ng crash"),
    ("TC_HOME_019","US-HP-03","HOME PAGE","AI Input","AI prompt c\u1ef1c d\u00e0i (5000+ k\u00fd t\u1ef1)","Negative","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Nh\u1eadp 5000+ k\u00fd t\u1ef1\n3. Click 'Generate'","Input truncate ho\u1eb7c validation: 'M\u00f4 t\u1ea3 qu\u00e1 d\u00e0i'. Kh\u00f4ng g\u1eedi payload qu\u00e1 l\u1edbn"),
    ("TC_HOME_020","US-HP-03","HOME PAGE","AI Input","AI prompt ch\u1ec9 c\u00f3 whitespace","Negative","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Nh\u1eadp '     ' (spaces)\n3. Click 'Generate'","Validation t\u01b0\u01a1ng t\u1ef1 input r\u1ed7ng. Trim whitespace"),
]

NEW_DS_TCS = [
    ("TC_DS_013","US-DS-04","DESIGN STUDIO","Panel","Nh\u1eadp XSS v\u00e0o textarea M\u00d4 T\u1ea2","Negative","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. M\u1edf Design Studio\n3. Nh\u1eadp '<script>alert(1)</script>'\n4. Click T\u1ea1o Artwork","Sanitize input. Kh\u00f4ng execute script"),
    ("TC_DS_014","US-DS-02","DESIGN STUDIO","Sidebar","Click Undo sau khi th\u00eam element","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Th\u00eam element l\u00ean canvas\n3. Click Ho\u00e0n T\u00e1c","Element bi\u1ebfn m\u1ea5t. Canvas quay v\u1ec1 tr\u1ea1ng th\u00e1i tr\u01b0\u1edbc"),
    ("TC_DS_015","US-DS-02","DESIGN STUDIO","Sidebar","Click Redo sau khi Undo","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Th\u00eam element \u2192 Undo\n3. Click L\u00e0m L\u1ea1i","Element xu\u1ea5t hi\u1ec7n tr\u1edf l\u1ea1i"),
    ("TC_DS_016","US-DS-02","DESIGN STUDIO","Sidebar","Click Undo khi kh\u00f4ng c\u00f3 action","Negative","P2","","","1. Truy c\u1eadp v\u00e0o trang \n2. Canvas tr\u1ed1ng\n3. Click Ho\u00e0n T\u00e1c","N\u00fat disabled ho\u1eb7c kh\u00f4ng ph\u1ea3n h\u1ed3i. Kh\u00f4ng crash"),
    ("TC_DS_017","US-DS-02","DESIGN STUDIO","Sidebar","Thu Ph\u00f3ng zoom in/out canvas","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click Thu Ph\u00f3ng Zoom In\n3. Click Zoom Out","Canvas ph\u00f3ng to/thu nh\u1ecf. Gi\u1edbi h\u1ea1n zoom min/max"),
    ("TC_DS_018","US-DS-05","DESIGN STUDIO","Bottom Bar","\u0110\u1ed5i m\u00e0u \u00e1o tr\u00ean bottom bar","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click toggle m\u00e0u (Tr\u1eafng \u2192 \u0110en)","Canvas mockup c\u1eadp nh\u1eadt \u0111\u00fang m\u00e0u m\u1edbi"),
    ("TC_DS_019","US-DS-05","DESIGN STUDIO","Bottom Bar","\u0110\u1ed5i size \u00e1o tr\u00ean bottom bar","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. \u0110\u1ed5i size L \u2192 S","Size c\u1eadp nh\u1eadt tr\u00ean bottom bar"),
    ("TC_DS_020","US-DS-03","DESIGN STUDIO","Canvas","Canvas empty \u2192 Th\u00eam AI Artwork","Positive","P0","","","1. Truy c\u1eadp v\u00e0o trang \n2. Canvas tr\u1ed1ng\n3. T\u1ea1o AI Artwork","Artwork hi\u1ec3n th\u1ecb. Text h\u01b0\u1edbng d\u1eabn bi\u1ebfn m\u1ea5t. Drag/resize \u0111\u01b0\u1ee3c"),
    ("TC_DS_021","US-DS-03","DESIGN STUDIO","Canvas","Canvas c\u00f3 artwork \u2192 X\u00f3a \u2192 tr\u1ed1ng","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Ch\u1ecdn artwork \u2192 X\u00f3a","Artwork bi\u1ebfn m\u1ea5t. Text h\u01b0\u1edbng d\u1eabn xu\u1ea5t hi\u1ec7n l\u1ea1i"),
    ("TC_DS_UI_860","Global","DESIGN STUDIO","Responsive & Zoom","DS: Browser Zoom 50%-200%","UI/UX","P1","","","1. Zoom In 200%\n2. Zoom Out 50%","Canvas, sidebar, panel \u0111\u00fang t\u1ef7 l\u1ec7"),
    ("TC_DS_UI_861","Global","DESIGN STUDIO","Responsive (iPhone)","DS: iPhone Portrait","UI/UX","P0","","","1. iPhone Portrait\n2. Thao t\u00e1c canvas, panel","Sidebar collapse. Panel bottom sheet. Canvas th\u00e1o t\u00e1c \u0111\u01b0\u1ee3c"),
    ("TC_DS_UI_862","Global","DESIGN STUDIO","Responsive (Android)","DS: Android Portrait","UI/UX","P0","","","1. Android Portrait\n2. Focus textarea","T\u01b0\u01a1ng t\u1ef1 iPhone. B\u00e0n ph\u00edm kh\u00f4ng che n\u00fat"),
    ("TC_DS_UI_863","Global","DESIGN STUDIO","Responsive (iPad)","DS: iPad Portrait","UI/UX","P1","","","1. iPad Portrait\n2. Quan s\u00e1t layout","Layout 3 c\u1ed9t gi\u1eef nguy\u00ean ho\u1eb7c 2 c\u1ed9t h\u1ee3p l\u00fd"),
    ("TC_DS_UI_864","Global","DESIGN STUDIO","Responsive (Android Tablet)","DS: Android Tablet Portrait","UI/UX","P1","","","1. Tablet Android Portrait","T\u01b0\u01a1ng t\u1ef1 iPad. Touch targets \u0111\u1ee7 l\u1edbn"),
    ("TC_DS_UI_865","Global","DESIGN STUDIO","Responsive (Landscape)","DS: Landscape","UI/UX","P0","","","1. Xoay ngang\n2. Quan s\u00e1t layout","Canvas resize. Bottom bar \u0111\u1ea7y \u0111\u1ee7"),
]

def apply_row_style(ws, row_idx, tc_type, priority):
    for col_idx in range(1, len(sub_h)+1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = bdr
        cell.font = fb
        h = sub_h[col_idx-1]
        if h in ["Steps","Expected_Result","Precondition","Title","Test_Data","Notes\n\u5099\u8003","Evidence\n\u8a3c\u62e0"]:
            cell.alignment = wa
        else:
            cell.alignment = ca
        if h == "Priority": cell.font = pfont(priority)
        if h == "Type": cell.font = tfont(tc_type)

def add_category_header(ws, row_idx, label):
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(sub_h))
    cell = ws.cell(row=row_idx, column=1)
    cell.value = f"\U0001f4cc {label}"
    cell.font = fcat; cell.fill = cf
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, len(sub_h)+1):
        ws.cell(row=row_idx, column=c).border = bdr

def main():
    if not os.path.exists(SOURCE):
        print(f"\u274c {SOURCE} not found!"); return
    print(f"\U0001f4d6 Loading v22...")
    shutil.copy2(SOURCE, OUTPUT)
    wb = openpyxl.load_workbook(OUTPUT)
    print(f"\U0001f4ca Sheets: {wb.sheetnames}")

    # --- FIX HOME PAGE existing rows ---
    if "HOME PAGE" in wb.sheetnames:
        ws = wb["HOME PAGE"]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=11):
            tc_id = str(row[0].value or "")
            if tc_id in HOME_FIXES:
                row[10].value = HOME_FIXES[tc_id]
                print(f"   \u2705 Fixed {tc_id}")

        # Add new HOME TCs (BVA) at end
        r = ws.max_row + 1
        add_category_header(ws, r, "Validation \u2014 BVA (AI Prompt)"); r += 1
        # Add header row for new section
        for tc in NEW_HOME_TCS:
            tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp = tc
            exec_t = "Manual" if tp=="UI/UX" else "Auto"
            data = [tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp,"Add new","By AI",exec_t,"Untested","","","","Untested","","","","",""]
            for ci, val in enumerate(data, 1):
                ws.cell(row=r, column=ci, value=val)
            apply_row_style(ws, r, tp, pri)
            r += 1
        print(f"   \u2795 Added {len(NEW_HOME_TCS)} BVA TCs to HOME PAGE")

    # --- FIX + ADD DESIGN STUDIO ---
    if "DESIGN STUDIO" in wb.sheetnames:
        ws = wb["DESIGN STUDIO"]
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=11):
            tc_id = str(row[0].value or "")
            if tc_id in DS_FIXES:
                row[9].value = DS_FIXES[tc_id]["steps"]
                row[10].value = DS_FIXES[tc_id]["expected"]
                print(f"   \u2705 Fixed {tc_id}")

        # Group new DS TCs by category
        categories = {
            "Validation \u2014 Security": [t for t in NEW_DS_TCS if "XSS" in t[4]],
            "Functional \u2014 Sidebar Tools": [t for t in NEW_DS_TCS if "Undo" in t[4] or "Redo" in t[4] or "Zoom" in t[4] or "Thu" in t[4]],
            "Functional \u2014 Bottom Bar": [t for t in NEW_DS_TCS if "m\u00e0u" in t[4] or "size" in t[4]],
            "Functional \u2014 Canvas State": [t for t in NEW_DS_TCS if "Canvas" in t[4] or "empty" in t[4] or "artwork" in t[4]],
            "Responsive \u2014 Design Studio": [t for t in NEW_DS_TCS if "DS:" in t[4] or "Responsive" in t[3] or "Zoom" in t[3]],
        }

        r = ws.max_row + 1
        total_added = 0
        for cat_name, tcs in categories.items():
            if not tcs: continue
            add_category_header(ws, r, cat_name); r += 1
            for tc in tcs:
                tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp = tc
                exec_t = "Manual" if tp=="UI/UX" else "Auto"
                data = [tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp,"Add new","By AI",exec_t,"Untested","","","","Untested","","","","",""]
                for ci, val in enumerate(data, 1):
                    ws.cell(row=r, column=ci, value=val)
                apply_row_style(ws, r, tp, pri)
                r += 1
                total_added += 1
        print(f"   \u2795 Added {total_added} new TCs to DESIGN STUDIO")

    # --- Update Cover Page ---
    if "Cover Page" in wb.sheetnames:
        cover = wb["Cover Page"]
        for row in cover.iter_rows(min_row=1, max_row=20, min_col=2, max_col=5):
            for cell in row:
                if cell.value and "DOCUMENT VERSION" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = f"v{VER}.0 (Excel formatted)"
                if cell.value and "GENERATED DATE" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # --- Update Change History ---
    if "Change History" in wb.sheetnames:
        hst = wb["Change History"]
        r = hst.max_row + 1
        for ci, val in enumerate([f"v{VER}.0", today, "Agent A Review Fixes: 5 TC text fixes + 17 new TCs (responsive DS, Undo/Redo/Zoom, Color/Size, BVA, XSS, State Transition)", "QA Team"], 2):
            c = hst.cell(row=r, column=ci, value=val)
            c.border = bdr; c.font = fb

    # --- Update Execution Summary ---
    if "Execution Summary" in wb.sheetnames:
        s = wb["Execution Summary"]
        r = s.max_row + 2
        s.cell(row=r, column=2, value="── v23 Fixes (Agent A Review) ──").font = fbd
        r += 1
        for feat in [("HOME PAGE", len(NEW_HOME_TCS)), ("DESIGN STUDIO", len(NEW_DS_TCS))]:
            name, cnt = feat
            sheet = name
            vals = [f"+ {name} (new TCs)",
                f"=COUNTIF('{sheet}'!O:O, \"Pass\")",
                f"=COUNTIF('{sheet}'!O:O, \"Fail\")",
                f"=G{r} - C{r} - D{r} - F{r}",
                f"=COUNTIF('{sheet}'!O:O, \"N/A\")",
                f"=COUNTA('{sheet}'!A:A)-2",
                f"=IF(G{r}>0, (C{r}+D{r})/G{r}, 0)",
                f"=IF(G{r}>0, C{r}/G{r}, 0)"]
            for ci, val in enumerate(vals, 2):
                c = s.cell(row=r, column=ci, value=val)
                c.border = bdr; c.font = fb
                c.alignment = Alignment(horizontal="left", vertical="center") if ci==2 else ca
                if ci in [8,9]: c.number_format = '0%'
            r += 1

    wb.save(OUTPUT)
    print(f"\n\U0001f389 v{VER} Excel generated: {OUTPUT}")
    print(f"   Fixed: 5 TCs | Added: {len(NEW_HOME_TCS) + len(NEW_DS_TCS)} new TCs")

if __name__ == "__main__":
    main()
