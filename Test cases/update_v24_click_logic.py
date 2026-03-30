"""
Update Excel v23 → v24: Add 17 missing click logic TCs
- HOME PAGE: +6 (style tags, generate with style, badge)
- DESIGN STUDIO: +11 (tabs, ref images, logo, style cards, artwork happy path)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os, datetime, shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Find latest v23
SOURCE = None
for f in sorted(os.listdir(BASE_DIR), reverse=True):
    if f.startswith("TC_POD") and f.endswith(".xlsx") and "v23" in f:
        SOURCE = os.path.join(BASE_DIR, f)
        break
if not SOURCE:
    for f in sorted(os.listdir(BASE_DIR), reverse=True):
        if f.startswith("TC_POD") and f.endswith(".xlsx"):
            SOURCE = os.path.join(BASE_DIR, f)
            break

today = datetime.datetime.now().strftime("%Y-%m-%d")
OUTPUT = os.path.join(BASE_DIR, f"TC_POD-TShirt-Platform_ExecutionSummary_v24_{today}.xlsx")
VER = 24

FF = "Calibri"; SZ = 11
fb = Font(name=FF, size=SZ)
fbd = Font(name=FF, size=SZ, bold=True)
fh = Font(name=FF, size=11, bold=True, color="FFFFFF")
bt = Side(border_style="thin", color="BFBFBF")
bdr = Border(left=bt, right=bt, top=bt, bottom=bt)
wa = Alignment(wrap_text=True, vertical="top")
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
fp0 = Font(name=FF, size=SZ, color="C00000", bold=True)
fp1 = Font(name=FF, size=SZ, color="ED7D31", bold=True)
fp2 = Font(name=FF, size=SZ, color="70AD47", bold=True)
fpos = Font(name=FF, size=SZ, color="0070C0")
fcat = Font(name=FF, size=12, bold=True, color="1F4E78")
cf = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

sub_h = ["TC_ID","US_Mapping","Feature","Module","Title","Type","Priority","Precondition","Test_Data","Steps","Expected_Result",
    "Action Type\n\u30a2\u30af\u30b7\u30e7\u30f3","Create TCs Type","Execution Type\n\u5b9f\u884c\u30bf\u30a4\u30d7",
    "Result\n\u7d50\u679c","Test date\n\u30c6\u30b9\u30c8\u65e5","Tester\n\u30c6\u30b9\u30bf\u30fc","ID Bug\n\u30d0\u30b0ID",
    "Result\n\u7d50\u679c","Test date\n\u30c6\u30b9\u30c8\u65e5","Tester\n\u30c6\u30b9\u30bf\u30fc","ID Bug\n\u30d0\u30b0ID",
    "Evidence\n\u8a3c\u62e0","Notes\n\u5099\u8003"]

def pfont(p):
    return fp0 if p=="P0" else fp1 if p=="P1" else fp2

def apply_row_style(ws, row_idx, tc_type, priority):
    for col_idx in range(1, len(sub_h)+1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = bdr
        cell.font = fb
        h = sub_h[col_idx-1]
        if h in ["Steps","Expected_Result","Precondition","Title","Test_Data"]:
            cell.alignment = wa
        else:
            cell.alignment = ca
        if h == "Priority": cell.font = pfont(priority)
        if h == "Type": cell.font = fpos

def add_category_header(ws, row_idx, label):
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(sub_h))
    cell = ws.cell(row=row_idx, column=1)
    cell.value = f"\U0001f4cc {label}"
    cell.font = fcat; cell.fill = cf
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for c in range(1, len(sub_h)+1):
        ws.cell(row=row_idx, column=c).border = bdr

# NEW HOME TCs
NEW_HOME = [
    ("TC_HOME_021","US-HP-03","HOME PAGE","AI Input","Click tag 'Streetwear' v\u00e0 verify","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click tag 'Streetwear' (\ud83d\udd25)","Tag 'Streetwear' highlight. C\u00e1c tag kh\u00e1c b\u1ecf highlight"),
    ("TC_HOME_022","US-HP-03","HOME PAGE","AI Input","Click tag 'Vintage' v\u00e0 verify","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click tag 'Vintage' (\ud83d\udcf7)","Tag 'Vintage' highlight. C\u00e1c tag kh\u00e1c b\u1ecf highlight"),
    ("TC_HOME_023","US-HP-03","HOME PAGE","AI Input","Click tag 'Y2K' v\u00e0 verify","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click tag 'Y2K' (\u2728)","Tag 'Y2K' highlight. C\u00e1c tag kh\u00e1c b\u1ecf highlight"),
    ("TC_HOME_024","US-HP-03","HOME PAGE","AI Input","Click tag 'Abstract Art' v\u00e0 verify","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click tag 'Abstract Art' (\ud83d\udd8c\ufe0f)","Tag 'Abstract Art' highlight. C\u00e1c tag kh\u00e1c b\u1ecf highlight"),
    ("TC_HOME_025","US-HP-03","HOME PAGE","AI Input","Click Generate sau khi ch\u1ecdn Style Tag","Positive","P0","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click tag 'Streetwear'\n3. Nh\u1eadp: '\u00c1o ph\u1ecbng c\u00e1ch \u0111\u01b0\u1eddng ph\u1ed1'\n4. Click 'Generate'","Prompt g\u1eedi k\u00e8m style. Loading state. Chuy\u1ec3n DS"),
    ("TC_HOME_026","US-HP-02","HOME PAGE","Hero","Click badge 'AI-Powered Design'","Positive","P2","","","1. Truy c\u1eadp v\u00e0o trang \n2. Click badge '\u2728 AI-Powered Design'","Badge kh\u00f4ng redirect ho\u1eb7c scroll xu\u1ed1ng AI Input"),
]

NEW_DS = [
    ("TC_DS_022","US-DS-04","DESIGN STUDIO","Panel","Click tab 'S\u1ea2N PH\u1ea8M' xem n\u1ed9i dung","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. M\u1edf Design Studio\n3. Click tab 'S\u1ea2N PH\u1ea8M' (\ud83d\udc55)","Tab 'S\u1ea2N PH\u1ea8M' active. Danh s\u00e1ch lo\u1ea1i \u00e1o hi\u1ec3n th\u1ecb"),
    ("TC_DS_023","US-DS-04","DESIGN STUDIO","Panel","Click tab '\u1ea2NH C\u1ee6A B\u1ea0N' xem n\u1ed9i dung","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. M\u1edf Design Studio\n3. Click tab '\u1ea2NH C\u1ee6A B\u1ea0N' (\ud83d\udcf7)","Tab '\u1ea2NH C\u1ee6A B\u1ea0N' active. Upload area hi\u1ec3n th\u1ecb"),
    ("TC_DS_024","US-DS-04","DESIGN STUDIO","Panel","Click tab 'TH\u01af VI\u1ec6N' xem n\u1ed9i dung","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. M\u1edf Design Studio\n3. Click tab 'TH\u01af VI\u1ec6N' (\ud83d\udcc2)","Tab 'TH\u01af VI\u1ec6N' active. Th\u01b0 vi\u1ec7n m\u1eabu hi\u1ec3n th\u1ecb"),
    ("TC_DS_025","US-DS-04","DESIGN STUDIO","Panel","Click tab '\u0110\u1eb6T H\u00c0NG' xem n\u1ed9i dung","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. M\u1edf Design Studio\n3. Click tab '\u0110\u1eb6T H\u00c0NG' (\ud83d\uded2)","Tab '\u0110\u1eb6T H\u00c0NG' active. Form \u0111\u1eb7t h\u00e0ng hi\u1ec3n th\u1ecb"),
    ("TC_DS_026","US-DS-04","DESIGN STUDIO","Panel","Click 'Ch\u1ecdn t\u1eeb th\u01b0 vi\u1ec7n' trong \u1ea2NH THAM KH\u1ea2O","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. DS \u2192 Tab 'T\u1ea0O \u1ea2NH AI'\n3. Click 'Ch\u1ecdn t\u1eeb th\u01b0 vi\u1ec7n'","M\u1edf modal gallery \u1ea3nh tham kh\u1ea3o"),
    ("TC_DS_027","US-DS-04","DESIGN STUDIO","Panel","Click 'T\u1ea3i \u1ea3nh l\u00ean' trong \u1ea2NH THAM KH\u1ea2O","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. DS \u2192 Tab 'T\u1ea0O \u1ea2NH AI'\n3. Click 'T\u1ea3i \u1ea3nh l\u00ean'","M\u1edf file picker. Ch\u1ea5p nh\u1eadn PNG/JPG. Preview hi\u1ec3n th\u1ecb"),
    ("TC_DS_028","US-DS-01","DESIGN STUDIO","Header","Click logo 'Tryonic' tr\u00ean DS header","Positive","P2","","","1. Truy c\u1eadp v\u00e0o trang \n2. M\u1edf Design Studio\n3. Click logo 'Tryonic'","Redirect v\u1ec1 /home/ ho\u1eb7c kh\u00f4ng ph\u1ea3n h\u1ed3i"),
    ("TC_DS_029","US-DS-04","DESIGN STUDIO","Panel","Click style card 'Line Art'","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. DS \u2192 Tab 'T\u1ea0O \u1ea2NH AI'\n3. Click 'Line Art'","Card 'Line Art' highlight. C\u00e1c card kh\u00e1c b\u1ecf"),
    ("TC_DS_030","US-DS-04","DESIGN STUDIO","Panel","Click style card 'Grunge'","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. DS \u2192 Tab 'T\u1ea0O \u1ea2NH AI'\n3. Click 'Grunge'","Card 'Grunge' highlight"),
    ("TC_DS_031","US-DS-04","DESIGN STUDIO","Panel","Click style card 'Flat Design'","Positive","P1","","","1. Truy c\u1eadp v\u00e0o trang \n2. DS \u2192 Tab 'T\u1ea0O \u1ea2NH AI'\n3. Click 'Flat Design'","Card 'Flat Design' highlight"),
    ("TC_DS_032","US-DS-04","DESIGN STUDIO","Panel","T\u1ea1o Artwork happy path \u0111\u1ea7y \u0111\u1ee7","Positive","P0","","","1. Truy c\u1eadp v\u00e0o trang \n2. DS \u2192 Tab 'T\u1ea0O \u1ea2NH AI'\n3. Nh\u1eadp: 'R\u1ed3ng Vi\u1ec7t Nam watercolor'\n4. Ch\u1ecdn 'Watercolor'\n5. Click 'T\u1ea1o Artwork M\u1edbi'","Loading. Artwork tr\u00ean canvas. Credits gi\u1ea3m 3"),
]

def main():
    if not SOURCE or not os.path.exists(SOURCE):
        print(f"\u274c Source not found!"); return
    print(f"\U0001f4d6 Loading {os.path.basename(SOURCE)}...")
    shutil.copy2(SOURCE, OUTPUT)
    wb = openpyxl.load_workbook(OUTPUT)
    print(f"\U0001f4ca Sheets: {wb.sheetnames}")

    # HOME PAGE - append new TCs
    if "HOME PAGE" in wb.sheetnames:
        ws = wb["HOME PAGE"]
        r = ws.max_row + 1
        add_category_header(ws, r, "Functional \u2014 Click Logic (Style Tags \u0026 Badge)"); r += 1
        for tc in NEW_HOME:
            tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp = tc
            data = [tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp,"Add new","By AI","Auto","Untested","","","","Untested","","","","",""]
            for ci, val in enumerate(data, 1):
                ws.cell(row=r, column=ci, value=val)
            apply_row_style(ws, r, tp, pri)
            r += 1
        print(f"   \u2795 Added {len(NEW_HOME)} TCs to HOME PAGE")

    # DESIGN STUDIO - append new TCs
    if "DESIGN STUDIO" in wb.sheetnames:
        ws = wb["DESIGN STUDIO"]
        r = ws.max_row + 1
        add_category_header(ws, r, "Functional \u2014 Click Logic (Tabs, Ref Images, Style Cards)"); r += 1
        for tc in NEW_DS:
            tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp = tc
            data = [tc_id,us,feat,mod,title,tp,pri,pre,td,steps,exp,"Add new","By AI","Auto","Untested","","","","Untested","","","","",""]
            for ci, val in enumerate(data, 1):
                ws.cell(row=r, column=ci, value=val)
            apply_row_style(ws, r, tp, pri)
            r += 1
        print(f"   \u2795 Added {len(NEW_DS)} TCs to DESIGN STUDIO")

    # Update Cover Page version
    if "Cover Page" in wb.sheetnames:
        cover = wb["Cover Page"]
        for row in cover.iter_rows(min_row=1, max_row=20, min_col=2, max_col=5):
            for cell in row:
                if cell.value and "DOCUMENT VERSION" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = f"v{VER}.0 (Excel formatted)"
                if cell.value and "GENERATED DATE" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Update Change History
    if "Change History" in wb.sheetnames:
        hst = wb["Change History"]
        r = hst.max_row + 1
        for ci, val in enumerate([f"v{VER}.0", today, f"Click Logic Audit: +{len(NEW_HOME)} HOME TCs (style tags, badge) + {len(NEW_DS)} DS TCs (tabs, ref images, style cards, artwork path)", "QA Team"], 2):
            c = hst.cell(row=r, column=ci, value=val)
            c.border = bdr; c.font = fb

    wb.save(OUTPUT)
    print(f"\n\U0001f389 v{VER} Excel generated: {OUTPUT}")
    print(f"   Added: {len(NEW_HOME) + len(NEW_DS)} new TCs")

if __name__ == "__main__":
    main()
