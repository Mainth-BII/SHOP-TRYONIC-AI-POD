"""
Update Excel v25 → v26: Apply all Review_Manual feedback fixes
Changes:
- HOME PAGE: 6 ADD (font size, active state), 5 FIX (expected results)
- DESIGN STUDIO: 7 FIX (icon colors, steps, expected results)
- TC_DS_008 split into TC_DS_008a + TC_DS_008b
- 3 new TCs for 'Chia sẻ thiết kế' screen (TC_DS_033, 034, 035)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os, datetime, shutil

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
today = datetime.datetime.now().strftime("%Y-%m-%d")

SOURCE = None
for f in sorted(os.listdir(BASE_DIR), reverse=True):
    if f.startswith("TC_POD") and f.endswith(".xlsx") and "v25" in f and not f.startswith("~$"):
        SOURCE = os.path.join(BASE_DIR, f)
        break

OUTPUT = os.path.join(BASE_DIR, f"TC_POD-TShirt-Platform_ExecutionSummary_v26_{today}.xlsx")
VER = 26

FF = "Calibri"; SZ = 11
fb = Font(name=FF, size=SZ)
fh = Font(name=FF, size=11, bold=True, color="FFFFFF")
hf = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
hf2 = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
bt = Side(border_style="thin", color="BFBFBF")
bdr = Border(left=bt, right=bt, top=bt, bottom=bt)
wa = Alignment(wrap_text=True, vertical="top")
ca = Alignment(horizontal="center", vertical="center", wrap_text=True)
fcat = Font(name=FF, size=12, bold=True, color="1F4E78")
cf = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
fp0 = Font(name=FF, size=SZ, color="C00000", bold=True)
fp1 = Font(name=FF, size=SZ, color="ED7D31", bold=True)
fp2 = Font(name=FF, size=SZ, color="70AD47", bold=True)
fui = Font(name=FF, size=SZ, color="7030A0", bold=True)
fneg = Font(name=FF, size=SZ, color="C00000")
fpos = Font(name=FF, size=SZ, color="0070C0")

# Review_Manual column style
frv = Font(name=FF, size=11, bold=True, color="000000")
hf_review = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

HEADERS_COUNT = 25  # columns A-Y

def pfont(p):
    if "P0" in str(p): return fp0
    if "P1" in str(p): return fp1
    return fp2

def tfont(t):
    if "UI/UX" in str(t): return fui
    if "Negative" in str(t): return fneg
    return fpos


# ---- HOME PAGE FIXES ----
HOME_EXPECTED_FIXES = {
    "TC_HOME_UI_002": "Hiển thị đầy đủ 4 menu items: 'Trang chủ', 'Sản phẩm', 'Dịch vụ', 'Liên hệ'. Font sans-serif, text đậm vừa. Menu 'Trang chủ' có trạng thái active/focus khi đang ở trang Home",
    "TC_HOME_UI_007": "Subtitle hiển thị: 'Chỉ cần mô tả — AI sẽ thiết kế cho bạn. Chất liệu premium, giao tận nơi.' Text xám, italic, centered. Font size ~16-18px",
    "TC_HOME_UI_010": "Hiển thị đầy đủ 6 tags: 'Minimalist', 'Streetwear', 'Anime', 'Vintage', 'Y2K', 'Abstract Art'. Mỗi tag có icon riêng, bo tròn full, border 1px. Font size ~14px",
    "TC_HOME_UI_011": "Nút 'Chọn từ mẫu có sẵn' hiển thị: Grid icon tím, subtitle 'Khám phá thư viện mẫu', background trắng, border tím. Font size title ~16px, subtitle ~13px",
    "TC_HOME_UI_012": "Nút 'Tải lên ảnh của bạn' hiển thị: Upload icon tím, subtitle 'Sử dụng file thiết kế riêng', background trắng, border tím. Font size title ~16px, subtitle ~13px",
    "TC_HOME_UI_013": "Hiển thị 3 markers: '✅ Thanh toán an toàn', '✅ Giao hàng toàn quốc', '✅ Đổi trả 7 ngày'. Mỗi marker có icon check xanh, phân cách bởi dấu chấm. Font size ~14px",
    "TC_HOME_004": "Chuyển sang Design Studio và hiển thị kết quả AI. Prompt được xử lý. Loading state hiển thị",
    "TC_HOME_006": "Chuyển sang Design Studio và mở tab thư viện mẫu hiển thị danh sách templates có sẵn",
    "TC_HOME_007": "Chuyển sang Design Studio và mở tab 'ẢNH CỦA BẠN' cho phép chọn file ảnh để tải lên. Chấp nhận formats: PNG, JPG, SVG",
    "TC_HOME_026": "Badge không phải link → không redirect",
    "TC_HOME_009": "Không validate. Chuyển sang Design Studio với thiết kế mặc định",
}

# ---- DESIGN STUDIO FIXES ----
DS_EXPECTED_FIXES = {
    "TC_DS_UI_001": "Header hiển thị: ← 'Quay lại' (trái), Logo 'Tryonic' icon (trái), text 'Design Studio' (giữa), Credits + icon User + icon Giỏ hàng (phải)",
    "TC_DS_UI_008": "Hiển thị: 'Áo Thun Cotton Gildan 5000', Màu: Trắng (toggle tròn), Size: L, 'Tạm tính: 150.000đ', text 'Giá chưa bao gồm phí in', nút 'Đặt hàng'",
    "TC_DS_004": "Mở đến màn hình 'Chia sẻ thiết kế'",
    "TC_DS_010": "Hiển thị tooltip thông tin credits. Không mở trang nạp riêng",
}

DS_STEPS_FIXES = {
    "TC_DS_007": "1. Truy cập vào trang \n2. Mở Design Studio\n3. Click nút 'Đặt hàng' ở góc phải bottom bar",
}

DS_TITLE_FIXES = {
    "TC_DS_004": "Click 'Chia Sẻ' mở màn hình chia sẻ",
    "TC_DS_018": "Đổi màu áo qua tab Sản phẩm",
    "TC_DS_019": "Đổi size áo qua tab Sản phẩm",
}

DS_MODULE_FIXES = {
    "TC_DS_018": "Panel",
    "TC_DS_019": "Panel",
}

DS_FULL_FIXES = {
    "TC_DS_018": {
        "steps": "1. Truy cập vào trang \n2. Mở Design Studio\n3. Click tab 'SẢN PHẨM'\n4. Đổi màu (Đen hoặc màu khác)",
        "expected": "Canvas mockup cập nhật hiển thị áo đúng màu mới. Giá tạm tính không thay đổi (cùng loại áo)"
    },
    "TC_DS_019": {
        "steps": "1. Truy cập vào trang \n2. Mở Design Studio\n3. Click tab 'SẢN PHẨM'\n4. Đổi size từ L → S hoặc XL",
        "expected": "Size hiển thị cập nhật trên bottom bar. Giá tạm tính có thể thay đổi hoặc giữ nguyên tùy chính sách"
    },
}

# New TCs for DS (TC_DS_008 split + share screen)
NEW_DS_TCS = [
    # TC_DS_008a: logged in
    ("TC_DS_008a","US-DS-01","DESIGN STUDIO","Header","Click icon User khi đã đăng nhập","Positive","P2",
     "Tài khoản đã đăng nhập","",
     "1. Truy cập vào trang (tài khoản đã đăng nhập)\n2. Mở Design Studio\n3. Click icon User (góc phải header)",
     "Mở trang profile/account. Hiển thị thông tin tài khoản"),
    # TC_DS_008b: guest
    ("TC_DS_008b","US-DS-01","DESIGN STUDIO","Header","Click icon User khi chưa đăng nhập","Positive","P2",
     "Chưa đăng nhập / Guest","",
     "1. Truy cập vào trang (chưa đăng nhập / Guest)\n2. Mở Design Studio\n3. Click icon User (góc phải header)",
     "Hiển thị popup login. Cho phép đăng nhập hoặc đăng ký"),
    # TC_DS_033: share screen content
    ("TC_DS_033","US-DS-02","DESIGN STUDIO","Sidebar","Màn hình 'Chia sẻ thiết kế' hiển thị đầy đủ","Positive","P2",
     "","",
     "1. Truy cập vào trang \n2. Mở Design Studio\n3. Click icon 'Chia Sẻ'\n4. Quan sát màn hình 'Chia sẻ thiết kế'",
     "Màn hình 'Chia sẻ thiết kế' hiển thị đầy đủ các tùy chọn chia sẻ (link, social media). Có nút đóng để quay lại Design Studio"),
    # TC_DS_034: copy link
    ("TC_DS_034","US-DS-02","DESIGN STUDIO","Sidebar","Copy link chia sẻ từ màn hình 'Chia sẻ thiết kế'","Positive","P2",
     "","",
     "1. Truy cập vào trang \n2. Mở màn hình 'Chia sẻ thiết kế'\n3. Click nút 'Copy link'",
     "Link được copy vào clipboard. Hiển thị thông báo 'Đã sao chép'"),
    # TC_DS_035: social share
    ("TC_DS_035","US-DS-02","DESIGN STUDIO","Sidebar","Chia sẻ qua social media từ màn hình 'Chia sẻ thiết kế'","Positive","P2",
     "","",
     "1. Truy cập vào trang \n2. Mở màn hình 'Chia sẻ thiết kế'\n3. Click icon social media (Facebook/Zalo/...)",
     "Mở cửa sổ chia sẻ của mạng xã hội tương ứng với link thiết kế"),
]


def apply_row_style(ws, row_idx, tc_type, priority, max_col=25):
    for col_idx in range(1, max_col+1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.border = bdr
        cell.font = fb
        if col_idx in [5, 8, 9, 10, 11, 23, 24]:
            cell.alignment = wa
        else:
            cell.alignment = ca
        if col_idx == 7: cell.font = pfont(priority)
        if col_idx == 6: cell.font = tfont(tc_type)


def main():
    if not SOURCE:
        print("❌ No v25 source Excel found!"); return
    print(f"📖 Source: {os.path.basename(SOURCE)}")

    shutil.copy2(SOURCE, OUTPUT)
    wb = openpyxl.load_workbook(OUTPUT)
    print(f"📊 Sheets: {wb.sheetnames}")

    # ---- Fix HOME PAGE ----
    if "HOME PAGE" in wb.sheetnames:
        ws = wb["HOME PAGE"]
        fixed = 0
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=HEADERS_COUNT):
            tc_id = str(row[0].value or "").strip()
            if not tc_id or tc_id.startswith("📌"):
                continue

            # Expected result fixes (col 11 = index 10)
            if tc_id in HOME_EXPECTED_FIXES:
                row[10].value = HOME_EXPECTED_FIXES[tc_id]
                # Clear review column after applying fix
                if len(row) >= 25:
                    row[24].value = f"✅ Fixed v26"
                print(f"   ✅ Fixed {tc_id} (expected)")
                fixed += 1
        
        print(f"   HOME PAGE: {fixed} TCs fixed")

    # ---- Fix DESIGN STUDIO ----
    if "DESIGN STUDIO" in wb.sheetnames:
        ws = wb["DESIGN STUDIO"]
        fixed = 0
        ds_008_row = None
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, max_col=HEADERS_COUNT):
            tc_id = str(row[0].value or "").strip()
            if not tc_id or tc_id.startswith("📌"):
                continue

            # Expected result fixes
            if tc_id in DS_EXPECTED_FIXES:
                row[10].value = DS_EXPECTED_FIXES[tc_id]
                if len(row) >= 25:
                    row[24].value = f"✅ Fixed v26"
                print(f"   ✅ Fixed {tc_id} (expected)")
                fixed += 1

            # Steps fixes
            if tc_id in DS_STEPS_FIXES:
                row[9].value = DS_STEPS_FIXES[tc_id]
                print(f"   ✅ Fixed {tc_id} (steps)")
                fixed += 1

            # Title fixes
            if tc_id in DS_TITLE_FIXES:
                row[4].value = DS_TITLE_FIXES[tc_id]
                print(f"   ✅ Fixed {tc_id} (title)")

            # Module fixes
            if tc_id in DS_MODULE_FIXES:
                row[3].value = DS_MODULE_FIXES[tc_id]
                print(f"   ✅ Fixed {tc_id} (module)")

            # Full fixes (steps + expected)
            if tc_id in DS_FULL_FIXES:
                row[9].value = DS_FULL_FIXES[tc_id]["steps"]
                row[10].value = DS_FULL_FIXES[tc_id]["expected"]
                if len(row) >= 25:
                    row[24].value = f"✅ Fixed v26"
                print(f"   ✅ Fixed {tc_id} (full)")
                fixed += 1

            # Also fix TC_DS_028 logo color
            if tc_id == "TC_DS_028":
                steps_val = str(row[9].value or "")
                if "icon hồng" in steps_val:
                    row[9].value = steps_val.replace("icon hồng", "icon")
                    print(f"   ✅ Fixed {tc_id} (logo color)")

            # Track TC_DS_008 row for replacement
            if tc_id == "TC_DS_008":
                ds_008_row = row[0].row

        # Replace TC_DS_008 with TC_DS_008a and add TC_DS_008b
        if ds_008_row:
            # Update existing row to TC_DS_008a
            ws.cell(row=ds_008_row, column=1).value = "TC_DS_008a"
            ws.cell(row=ds_008_row, column=4).value = "Header"
            ws.cell(row=ds_008_row, column=5).value = "Click icon User khi đã đăng nhập"
            ws.cell(row=ds_008_row, column=8).value = "Tài khoản đã đăng nhập"
            ws.cell(row=ds_008_row, column=10).value = "1. Truy cập vào trang (tài khoản đã đăng nhập)\n2. Mở Design Studio\n3. Click icon User (góc phải header)"
            ws.cell(row=ds_008_row, column=11).value = "Mở trang profile/account. Hiển thị thông tin tài khoản"
            if ws.cell(row=ds_008_row, column=25).value:
                ws.cell(row=ds_008_row, column=25).value = "✅ Split v26"
            print(f"   ✅ Replaced TC_DS_008 → TC_DS_008a")

            # Insert new row for TC_DS_008b right after
            ws.insert_rows(ds_008_row + 1)
            new_r = ds_008_row + 1
            data_8b = ["TC_DS_008b","US-DS-01","DESIGN STUDIO","Header",
                "Click icon User khi chưa đăng nhập","Positive","P2",
                "Chưa đăng nhập / Guest","",
                "1. Truy cập vào trang (chưa đăng nhập / Guest)\n2. Mở Design Studio\n3. Click icon User (góc phải header)",
                "Hiển thị popup login. Cho phép đăng nhập hoặc đăng ký",
                "Add new","By AI","Manual",
                "Untested","","","",
                "Untested","","","",
                "","",
                "✅ New v26"]
            for ci, val in enumerate(data_8b, 1):
                ws.cell(row=new_r, column=ci, value=val)
            apply_row_style(ws, new_r, "Positive", "P2")
            print(f"   ➕ Added TC_DS_008b")

        # Add new share screen TCs at end
        r = ws.max_row + 1
        # Category header
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=HEADERS_COUNT)
        cell = ws.cell(row=r, column=1, value="📌 Functional — Chia sẻ thiết kế")
        cell.font = fcat; cell.fill = cf
        cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        for c in range(1, HEADERS_COUNT+1):
            ws.cell(row=r, column=c).border = bdr
            ws.cell(row=r, column=c).fill = cf
        r += 1

        for tc in NEW_DS_TCS:
            if tc[0].startswith("TC_DS_008"):
                continue  # Already handled above
            tc_id, us, feat, mod, title, tp, pri, pre, td, steps, exp = tc
            exec_t = "Manual" if tp == "UI/UX" else "Auto"
            data = [tc_id, us, feat, mod, title, tp, pri, pre, td, steps, exp,
                    "Add new", "By AI", exec_t,
                    "Untested", "", "", "",
                    "Untested", "", "", "",
                    "", "",
                    "✅ New v26"]
            for ci, val in enumerate(data, 1):
                ws.cell(row=r, column=ci, value=val)
            apply_row_style(ws, r, tp, pri)
            r += 1
        print(f"   ➕ Added 3 new share screen TCs")
        print(f"   DESIGN STUDIO: {fixed} TCs fixed total")

    # ---- Update Cover Page ----
    if "Cover Page" in wb.sheetnames:
        cover = wb["Cover Page"]
        for row in cover.iter_rows(min_row=1, max_row=20, min_col=2, max_col=5):
            for cell in row:
                if cell.value and "DOCUMENT VERSION" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = f"v{VER}.0"
                if cell.value and "GENERATED DATE" in str(cell.value).upper():
                    cover.cell(row=cell.row, column=3).value = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ---- Update Change History ----
    if "Change History" in wb.sheetnames:
        hst = wb["Change History"]
        r = hst.max_row + 1
        hist_data = [f"v{VER}.0", today, 
            f"Review Manual fixes: HOME(11 fixes: font size, active state, expected results), DS(9 fixes: icon color, button, share screen, tab nav). Split TC_DS_008→2 cases. +3 new TCs (share screen).", 
            "QA Team"]
        for ci, val in enumerate(hist_data, 2):
            c = hst.cell(row=r, column=ci, value=val)
            c.border = bdr; c.font = fb

    wb.save(OUTPUT)
    print(f"\n🎉 v{VER} saved: {os.path.basename(OUTPUT)}")


if __name__ == "__main__":
    main()
