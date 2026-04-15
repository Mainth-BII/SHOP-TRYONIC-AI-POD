"""Report writer — saves test results to CSV and Excel."""

import csv
import os
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class ReportWriter:
    """Collects test results and writes them to CSV (and optionally Excel)."""

    HEADERS = [
        "TC_ID", "Screen", "Module", "Title", "Type", "Priority",
        "Steps", "Expected_Result", "Actual_Result",
        "Result_R1", "Test_Date_R1", "Tester_R1",
        "Bug_ID_R1", "Bug_Desc_R1", "Evidence", "Notes",
    ]

    def __init__(self, output_dir: str = "test_reports"):
        self.results: list[dict] = []
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)
        self._date = datetime.now().strftime("%Y-%m-%d")
        self._run_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # ── Public API ────────────────────────────────────────────────────────────

    def add(
        self,
        tc_id: str,
        status: str,                    # "PASS" | "FAIL" | "N/A"
        *,
        screen: str = "",
        module: str = "",
        title: str = "",
        tc_type: str = "",
        priority: str = "",
        steps: str = "",
        expected: str = "",
        actual: str = "",
        error: str = "",
        screenshot: str = "",
        bug_id: str = "",
        bug_desc: str = "",
        notes: str = "",
    ) -> None:
        self.results.append({
            "TC_ID": tc_id,
            "Screen": screen,
            "Module": module,
            "Title": title,
            "Type": tc_type,
            "Priority": priority,
            "Steps": steps,
            "Expected_Result": expected,
            "Actual_Result": actual or error,
            "Result_R1": status,
            "Test_Date_R1": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "Tester_R1": "Playwright CI",
            "Bug_ID_R1": bug_id,
            "Bug_Desc_R1": bug_desc,
            "Evidence": screenshot,
            "Notes": notes,
        })

    def save(self) -> dict[str, str]:
        """Write CSV and Excel reports; return dict of file paths."""
        paths: dict[str, str] = {}
        paths["csv"] = self._save_csv()
        if HAS_OPENPYXL:
            paths["excel"] = self._save_excel()
        return paths

    def summary(self) -> dict:
        total = len(self.results)
        passed = sum(1 for r in self.results if r["Result_R1"] == "PASS")
        failed = sum(1 for r in self.results if r["Result_R1"] == "FAIL")
        return {
            "total": total,
            "passed": passed,
            "failed": failed,
            "skipped": total - passed - failed,
            "pass_rate": f"{(passed / total * 100):.1f}%" if total else "0%",
            "run_time": self._run_time,
        }

    # ── Private ───────────────────────────────────────────────────────────────

    def _save_csv(self) -> str:
        path = os.path.join(self.output_dir, f"RESULT_{self._date}.csv")
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=self.HEADERS)
            writer.writeheader()
            writer.writerows(self.results)
        print(f"[Report] CSV saved → {path}")
        return path

    def _save_excel(self) -> str:
        path = os.path.join(self.output_dir, f"RESULT_{self._date}.xlsx")
        wb = openpyxl.Workbook()

        # ── Summary sheet ──────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        s = self.summary()
        ws_summary.append(["Tryonic AI — Test Execution Report"])
        ws_summary.append(["Run Time", s["run_time"]])
        ws_summary.append(["Total", s["total"]])
        ws_summary.append(["PASS", s["passed"]])
        ws_summary.append(["FAIL", s["failed"]])
        ws_summary.append(["Pass Rate", s["pass_rate"]])

        # ── Results sheet ──────────────────────────────────────────────────
        ws = wb.create_sheet("Results")
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        header_font = Font(bold=True, color="FFFFFF")

        ws.append(self.HEADERS)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        pass_fill = PatternFill("solid", fgColor="C6EFCE")
        fail_fill = PatternFill("solid", fgColor="FFC7CE")

        for row_data in self.results:
            row = [row_data.get(h, "") for h in self.HEADERS]
            ws.append(row)
            last_row = ws.max_row
            status = row_data.get("Result_R1", "")
            fill = pass_fill if status == "PASS" else (fail_fill if status == "FAIL" else None)
            if fill:
                for cell in ws[last_row]:
                    cell.fill = fill

            # Make Evidence cell a hyperlink
            ev_col = self.HEADERS.index("Evidence") + 1
            ev_cell = ws.cell(row=last_row, column=ev_col)
            if ev_cell.value and os.path.exists(str(ev_cell.value)):
                ev_cell.hyperlink = str(ev_cell.value)
                ev_cell.font = Font(color="0563C1", underline="single")

        # Auto-fit columns
        for col_idx, col in enumerate(ws.columns, 1):
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

        wb.save(path)
        print(f"[Report] Excel saved → {path}")
        return path
